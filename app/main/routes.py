from flask import (
    Blueprint,
    render_template,
    session,
    redirect,
    url_for,
    abort,
    request,
    jsonify,
    current_app,
    send_file,
    flash,
    g,
    make_response,
)
from functools import wraps
import csv
import io
import os
from pathlib import Path
from urllib.parse import urlparse
import re
import json
import sqlite3
from datetime import datetime, date, timezone, timedelta
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError
from openpyxl import load_workbook
import xlrd
import base64
import math
from werkzeug.security import generate_password_hash
try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
except Exception:  # pragma: no cover
    matplotlib = None
    plt = None

from app.db import (
    fetch_app_versions,
    delete_app_user,
    fetch_aoi_reports,
    fetch_combined_reports,
    fetch_defect_catalog,
    fetch_app_user_credentials,
    fetch_app_users,
    fetch_fi_reports,
    query_aoi_base_daily,
    fetch_moat,
    fetch_moat_dpm,
    fetch_recent_moat,
    fetch_recent_moat_dpm,
    fetch_saved_queries,
    fetch_dpm_saved_queries,
    fetch_saved_aoi_queries,
    fetch_saved_fi_queries,
    fetch_part_results,
    insert_saved_query,
    insert_dpm_saved_query,
    update_saved_query,
    update_dpm_saved_query,
    insert_saved_aoi_query,
    update_saved_aoi_query,
    insert_saved_fi_query,
    update_saved_fi_query,
    insert_aoi_report,
    insert_aoi_reports_bulk,
    insert_app_user,
    insert_bug_report,
    insert_fi_report,
    insert_moat,
    insert_moat_bulk,
    insert_moat_dpm_bulk,
    fetch_bug_reports,
    update_bug_report_status,
    fetch_feature_states,
    fetch_feature_states_for_bug,
    upsert_feature_state,
)

_ORIGINAL_AOI_QUERY = query_aoi_base_daily
from app.grades import calculate_aoi_grades
from app.main.pdf_utils import PdfGenerationError, render_html_to_pdf
from app.auth import routes as auth_routes
from fi_utils import parse_fi_rejections

# Helpers for AOI Grades analytics
from collections import defaultdict, Counter
from itertools import combinations
from statistics import mean, pstdev


def _load_report_css() -> str:
    """Load the shared report stylesheet so it can be inlined."""

    static_folder = current_app.static_folder or ''
    css_path = Path(static_folder) / 'css' / 'report.css'
    try:
        return css_path.read_text(encoding='utf-8')
    except OSError as exc:  # pragma: no cover - log & fall back to default styling
        current_app.logger.warning("Unable to load report CSS: %s", exc)
    return ""


def _report_timezone():
    """Return the timezone used for report timestamps.

    Prefers the configured ``LOCAL_TIMEZONE`` (defaulting to America/New_York)
    and falls back to UTC if the zone cannot be loaded.
    """

    tz_name = current_app.config.get("LOCAL_TIMEZONE") or "America/New_York"
    try:
        return ZoneInfo(tz_name)
    except ZoneInfoNotFoundError:
        current_app.logger.warning(
            "Timezone %s unavailable; falling back to UTC", tz_name
        )
    except Exception as exc:  # pragma: no cover - unexpected zoneinfo failures
        current_app.logger.warning(
            "Error loading timezone %s: %s; falling back to UTC", tz_name, exc
        )
    return timezone.utc


def _load_local_dpm_saved_charts() -> list[dict]:
    """Return built-in DPM saved chart definitions for offline use."""

    try:
        config_dir = Path(current_app.root_path).parent / 'config'
        json_path = config_dir / 'dpm_saved_charts.json'
        if not json_path.exists():
            return []
        with json_path.open(encoding='utf-8') as handle:
            payload = json.load(handle)
        if isinstance(payload, list):
            return payload
        current_app.logger.warning(
            "dpm_saved_charts.json is not a list; ignoring fallback definitions",
        )
        return []
    except OSError as exc:  # pragma: no cover - filesystem issues
        current_app.logger.warning("Unable to load local DPM saved charts: %s", exc)
        return []


def _normalize_header(value: str | None) -> str:
    """Normalize a CSV header by trimming whitespace and lowercasing."""

    return (value or "").strip().lower()


def _compare_headers(
    fieldnames: list[str] | None,
    ordered_columns: list[str],
    optional_columns: list[str] | None = None,
) -> tuple[list[str], list[str], list[str], dict[str, str | None]]:
    """Compare CSV headers against the required column list.

    Returns lists of missing, unexpected, and out-of-order columns along with a
    mapping from required column names to the actual header present in the CSV
    (if any). Comparison is performed after trimming whitespace and
    normalising casing so that minor formatting issues do not prevent uploads.
    """

    actual_headers = fieldnames or []
    optional_columns = optional_columns or []
    normalized_actual = [_normalize_header(name) for name in actual_headers]
    normalized_order = [_normalize_header(name) for name in ordered_columns]

    actual_lookup: dict[str, str] = {}
    for original, normalized in zip(actual_headers, normalized_actual):
        if normalized and normalized not in actual_lookup:
            actual_lookup[normalized] = original

    expected_set = set(normalized_order)

    missing = [
        column
        for column, normalized in zip(ordered_columns, normalized_order)
        if column not in optional_columns and normalized not in actual_lookup
    ]

    unexpected = [
        original
        for original, normalized in zip(actual_headers, normalized_actual)
        if normalized not in expected_set
    ]

    out_of_order: list[str] = []
    filtered_actual = [
        (original, normalized)
        for original, normalized in zip(actual_headers, normalized_actual)
        if normalized in expected_set
    ]
    expected_order = [
        normalized
        for normalized in normalized_order
        if normalized in actual_lookup
    ]
    if (
        not missing
        and not unexpected
        and len(filtered_actual) == len(expected_order)
    ):
        for (original, actual_norm), required_norm in zip(
            filtered_actual, expected_order
        ):
            if actual_norm != required_norm:
                out_of_order.append(original)

    mapping = {
        column: actual_lookup.get(normalized)
        for column, normalized in zip(ordered_columns, normalized_order)
    }

    return missing, unexpected, out_of_order, mapping

def _parse_date(val):
    if not val:
        return None

    if isinstance(val, datetime):
        return val.date()

    if isinstance(val, date):
        return val

    text = str(val).strip()
    if not text:
        return None

    if text.endswith("Z"):
        text = text[:-1] + "+00:00"

    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        return None

    if isinstance(parsed, datetime):
        return parsed.date()

    return parsed


def _coerce_number(value, *, default=0.0):
    """Convert Excel cell values to floats, stripping formatting."""

    if value is None:
        return default

    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return default
        return float(value)

    text = str(value).strip()
    if not text:
        return default

    text = text.replace(',', '')
    if text.endswith('%'):
        text = text[:-1]
    text = text.strip()
    if not text:
        return default

    lowered = text.lower()
    invalid_tokens = {
        'nan',
        '+nan',
        '-nan',
        'inf',
        '+inf',
        '-inf',
        'infinity',
        '+infinity',
        '-infinity',
    }
    if lowered in invalid_tokens:
        return default

    try:
        number = float(text)
    except (TypeError, ValueError):
        return default

    if math.isnan(number) or math.isinf(number):
        return default

    return number


def _aoi_passed(row):
    ins = float(row.get('aoi_Quantity Inspected') or row.get('Quantity Inspected') or 0)
    rej = float(row.get('aoi_Quantity Rejected') or row.get('Quantity Rejected') or 0)
    v = ins - rej
    return v if v > 0 else 0


def _fi_rejected(row):
    return float(row.get('fi_Quantity Rejected') or 0)


def _fi_inspected(row):
    return float(row.get('fi_Quantity Inspected') or 0)


def _gap_days(row):
    a = _parse_date(row.get('aoi_Date'))
    f = _parse_date(row.get('fi_Date'))
    if a and f:
        return (f - a).days
    return None


def _predict_counts(inspected: float, rejected: float, boards: float) -> dict[str, float]:
    """Compute predicted reject count and yield based on historical rates."""
    reject_rate = (rejected / inspected) if inspected else 0.0
    predicted_rejects = reject_rate * boards
    predicted_yield = (
        (boards - predicted_rejects) / boards * 100.0 if boards else 0.0
    )
    return {"predictedRejects": predicted_rejects, "predictedYield": predicted_yield}


def _split_model_name(name: str) -> tuple[str, str]:
    """Split a MOAT 'Model Name' into assembly and program.

    Examples::
        'Asm1 SMT' -> ('Asm1', 'SMT')

    If a program cannot be determined, the program portion will be an empty
    string. Whitespace is trimmed from both parts.
    """
    if not name:
        return "", ""
    parts = re.split(r"\s+", str(name).strip())
    if len(parts) >= 2:
        return " ".join(parts[:-1]), parts[-1]
    return str(name).strip(), ""


def _norm(value: str) -> str:
    """Normalize an assembly or program string for comparisons."""
    return re.sub(r'[-\s]+', ' ', value).strip().lower() if value else ''


def _aggregate_forecast(
    assemblies: list[str], moat_rows: list[dict], aoi_rows: list[dict]
) -> list[dict]:
    """Aggregate MOAT and AOI data for selected assemblies.

    MOAT rows may contain a combined "Model Name" field where the final token
    represents the program. We split that field into separate assembly/program
    pieces so that MOAT and AOI records can be joined on both attributes.
    Assembly and program strings are normalised to lower-case without leading or
    trailing whitespace prior to comparison. The returned dictionaries include a
    ``missing`` boolean flag indicating when an assembly does not appear in
    either dataset.
    """

    by_name = {_norm(a): a for a in assemblies if a}

    assembly_customer: dict[str, str] = {}
    customer_totals: dict[str, dict[str, float]] = defaultdict(
        lambda: {"inspected": 0.0, "rejected": 0.0}
    )

    moat_map: dict[tuple[str, str], dict[str, float]] = defaultdict(
        lambda: {"boards": 0.0, "falseCalls": 0.0}
    )
    for row in moat_rows or []:
        model_name = row.get("Model Name") or ""
        model_norm = _norm(model_name)
        asm_guess, prog_guess = _split_model_name(model_name)
        asm_raw = row.get("Assembly") or row.get("Model") or asm_guess
        prog_raw = row.get("Program") or row.get("program") or prog_guess
        cust_raw = row.get("Customer") or row.get("customer") or ""

        matched_asm = next(
            (a for a in sorted(by_name.keys(), key=len, reverse=True) if a in model_norm),
            None,
        )
        asm_key = matched_asm or _norm(asm_raw)
        prog_key = _norm(prog_raw)
        if asm_key and cust_raw and asm_key not in assembly_customer:
            assembly_customer[asm_key] = cust_raw
        if not asm_key:
            continue
        try:
            boards = float(row.get("Total Boards") or row.get("total_boards") or 0)
            false_calls = float(
                row.get("FalseCall Parts") or row.get("falsecall_parts") or 0
            )
        except (TypeError, ValueError):
            continue
        moat_map[(asm_key, prog_key)]["boards"] += boards
        moat_map[(asm_key, prog_key)]["falseCalls"] += false_calls

    aoi_map: dict[tuple[str, str], dict[str, float]] = defaultdict(
        lambda: {"inspected": 0.0, "rejected": 0.0}
    )
    for row in aoi_rows or []:
        asm_raw = row.get("Assembly") or row.get("aoi_Assembly") or ""
        prog_raw = row.get("Program") or row.get("aoi_Program") or ""
        cust_raw = row.get("Customer") or row.get("aoi_Customer") or ""
        asm_key = _norm(asm_raw)
        prog_key = _norm(prog_raw)
        if asm_key and cust_raw and asm_key not in assembly_customer:
            assembly_customer[asm_key] = cust_raw
        if not asm_key:
            continue
        try:
            inspected = float(
                row.get("Quantity Inspected")
                or row.get("aoi_Quantity Inspected")
                or 0
            )
            rejected = float(
                row.get("Quantity Rejected")
                or row.get("aoi_Quantity Rejected")
                or 0
            )
        except (TypeError, ValueError):
            continue
        aoi_map[(asm_key, prog_key)]["inspected"] += inspected
        aoi_map[(asm_key, prog_key)]["rejected"] += rejected
        if cust_raw:
            ck = _norm(cust_raw)
            customer_totals[ck]["inspected"] += inspected
            customer_totals[ck]["rejected"] += rejected

    customer_yields = {
        c: (
            (vals["inspected"] - vals["rejected"]) / vals["inspected"] * 100.0
            if vals["inspected"]
            else 0.0
        )
        for c, vals in customer_totals.items()
    }

    moat_asms = {a for a, _ in moat_map.keys()}
    aoi_asms = {a for a, _ in aoi_map.keys()}

    results: list[dict] = []
    for asm_key, original in by_name.items():
        boards = 0.0
        false_calls = 0.0
        inspected = 0.0
        rejected = 0.0
        prog_keys = (
            {p for a, p in moat_map.keys() if a == asm_key}
            | {p for a, p in aoi_map.keys() if a == asm_key}
        )
        for prog in prog_keys:
            m = moat_map.get((asm_key, prog), {})
            a = aoi_map.get((asm_key, prog), {})
            boards += m.get("boards", 0.0)
            false_calls += m.get("falseCalls", 0.0)
            inspected += a.get("inspected", 0.0)
            rejected += a.get("rejected", 0.0)

        avg_fc = false_calls / boards if boards else 0.0
        predicted_fc = avg_fc * boards
        preds = _predict_counts(inspected, rejected, boards)
        yield_pct = (
            (inspected - rejected) / inspected * 100.0 if inspected else 0.0
        )
        ng_ratio = rejected / inspected * 100.0 if inspected else 0.0
        predicted_ng_per_board = (
            preds.get("predictedRejects", 0.0) / boards if boards else 0.0
        )
        predicted_fc_per_board = predicted_fc / boards if boards else 0.0
        cust_yield = 0.0
        cust_raw = assembly_customer.get(asm_key)
        if cust_raw:
            cust_yield = customer_yields.get(_norm(cust_raw), 0.0)
        missing = not ((asm_key in moat_asms) or (asm_key in aoi_asms))
        results.append(
            {
                "assembly": original,
                "customer": cust_raw,
                "boards": boards,
                "falseCalls": false_calls,
                "avgFalseCalls": avg_fc,
                "predictedFalseCalls": predicted_fc,
                "inspected": inspected,
                "rejected": rejected,
                "yield": yield_pct,
                "ngRatio": ng_ratio,
                "predictedNGsPerBoard": predicted_ng_per_board,
                "predictedFCPerBoard": predicted_fc_per_board,
                "customerYield": cust_yield,
                **preds,
                "missing": missing,
            }
        )
    return results


main_bp = Blueprint('main', __name__)

EMPLOYEE_AREA_OPTIONS = [
    "AOI",
    "SMT",
    "Rework",
    "Hand Assembly",
    "Solder",
    "Test",
    "RMA",
    "Coating",
    "Final Inspect",
    "ICT",
]

EMPLOYEE_SHEET_LABELS = {
    "SMT": "SMT AOI Inspection Data Sheet",
    "TH": "TH AOI Inspection Data Sheet",
}


FEATURE_REGISTRY = [
    {
        "slug": "analysis_ppm",
        "label": "PPM Analysis",
        "category": "analysis",
        "description": "Monitor AOI false-call performance and long-term yield trends.",
        "default_message": "PPM Analysis is temporarily unavailable while we perform maintenance.",
    },
    {
        "slug": "analysis_dpm",
        "label": "DPM Analysis",
        "category": "analysis",
        "description": "Track defect-per-million trends and compare AOI performance by line or model.",
        "default_message": "DPM Analysis is temporarily unavailable while we perform maintenance.",
    },
    {
        "slug": "analysis_aoi_grades",
        "label": "AOI & FI Analysis",
        "category": "analysis",
        "description": "Advanced AOI and FI analytics including escape pareto and learning curves.",
        "default_message": "AOI & FI Analysis is undergoing maintenance. Please check back soon.",
    },
    {
        "slug": "analysis_aoi_daily",
        "label": "AOI Daily Reports",
        "category": "analysis",
        "description": "Review daily AOI production metrics and exportable summaries.",
        "default_message": "AOI Daily Reports are temporarily locked while we investigate an issue.",
    },
    {
        "slug": "analysis_fi_daily",
        "label": "FI Daily Reports",
        "category": "analysis",
        "description": "Inspect final inspection throughput and rejection trends.",
        "default_message": "FI Daily Reports are paused for maintenance.",
    },
    {
        "slug": "reports_integrated",
        "label": "AOI Integrated Report",
        "category": "reports",
        "description": "Combined AOI insights with export capabilities for leadership reviews.",
        "default_message": "The AOI Integrated Report is temporarily unavailable.",
    },
    {
        "slug": "reports_operator",
        "label": "Operator Report",
        "category": "reports",
        "description": "Operator-level KPIs and drill-down performance dashboards.",
        "default_message": "The Operator Report is temporarily unavailable while we resolve a bug.",
    },
    {
        "slug": "reports_line",
        "label": "Line Report",
        "category": "reports",
        "description": "Line-level AOI performance, benchmarking, and synchronization insights.",
        "default_message": "The Line Report is temporarily unavailable while we finish calibrations.",
    },
    {
        "slug": "reports_part",
        "label": "Part Report",
        "category": "reports",
        "description": "Part-level AOI performance, spatial metrics, and false-call intelligence.",
        "default_message": "The Part Report is temporarily unavailable while we verify data sources.",
    },
    {
        "slug": "reports_aoi_daily",
        "label": "AOI Daily Report",
        "category": "reports",
        "description": "Classic AOI daily production summary for distribution.",
        "default_message": "The AOI Daily Report export is momentarily locked.",
    },
    {
        "slug": "tools_assembly_forecast",
        "label": "Assembly Forecast",
        "category": "tools",
        "description": "Forecast SMT assemblies and predict workload based on recent history.",
        "default_message": "Assembly Forecast is temporarily offline while we improve accuracy.",
    },
]

FEATURE_DEFINITIONS = {entry["slug"]: entry for entry in FEATURE_REGISTRY}

FEATURE_STATUS_AVAILABLE = "available"
FEATURE_STATUS_LOCKED = "locked"


def _feature_definition(slug: str) -> dict[str, str]:
    return FEATURE_DEFINITIONS.get(slug, {})


def _normalize_feature_status(status: str | None) -> str:
    normalized = (status or FEATURE_STATUS_AVAILABLE).strip().lower()
    return normalized or FEATURE_STATUS_AVAILABLE


def _get_feature_state_map() -> dict[str, dict[str, object]]:
    cached = getattr(g, "_feature_state_map", None)
    if cached is not None:
        return cached

    records, error = fetch_feature_states()
    mapping: dict[str, dict[str, object]] = {}
    for record in records or []:
        slug = record.get("slug")
        if not slug:
            continue
        key = str(slug)
        mapping[key] = {
            "status": _normalize_feature_status(record.get("status")),
            "message": record.get("message") or None,
            "bug_report_id": record.get("bug_report_id"),
            "updated_at": record.get("updated_at"),
        }

    g._feature_state_map = mapping
    g._feature_state_error = error
    return mapping


def _compose_feature_state(slug: str) -> dict[str, object]:
    definition = _feature_definition(slug)
    merged = {
        "slug": slug,
        "label": definition.get("label", slug.replace("_", " ").title()),
        "category": definition.get("category"),
        "description": definition.get("description"),
    }

    state_map = _get_feature_state_map()
    record = state_map.get(slug, {})
    status = _normalize_feature_status(record.get("status"))
    default_message = definition.get("default_message") or definition.get("description")
    message = record.get("message")
    if not message and status != FEATURE_STATUS_AVAILABLE:
        message = default_message
    merged.update(
        {
            "status": status,
            "message": message or "",
            "bug_report_id": record.get("bug_report_id"),
            "updated_at": record.get("updated_at"),
        }
    )
    return merged


def _feature_locked_response(slug: str):
    state = _compose_feature_state(slug)
    status = state.get("status") or FEATURE_STATUS_LOCKED
    message = state.get("message") or (
        f"{state.get('label', 'This feature')} is currently unavailable."
    )

    wants_json = False
    if request.path.startswith("/api/"):
        wants_json = True
    if request.is_json:
        wants_json = True
    accept = request.accept_mimetypes
    if accept and accept.best == "application/json" and accept["application/json"] > accept["text/html"]:
        wants_json = True

    if wants_json:
        payload = {
            "error": "feature_locked",
            "feature": slug,
            "status": status,
            "message": message,
        }
        return jsonify(payload), 423

    flash(message, "warning")
    return redirect(url_for("main.home"))


def feature_required(slug: str):
    """Decorator enforcing feature availability for non-admin users."""

    def decorator(view):
        @wraps(view)
        def wrapped_view(*args, **kwargs):
            role = (session.get("role") or session.get("username") or "").upper()
            if role == "ADMIN":
                return view(*args, **kwargs)

            state = _compose_feature_state(slug)
            status = _normalize_feature_status(state.get("status"))
            if status == FEATURE_STATUS_AVAILABLE:
                return view(*args, **kwargs)

            return _feature_locked_response(slug)

        return wrapped_view

    return decorator


@main_bp.app_context_processor
def inject_feature_state_context() -> dict[str, object]:
    try:
        context = {entry["slug"]: _compose_feature_state(entry["slug"]) for entry in FEATURE_REGISTRY}
    except Exception:  # pragma: no cover - defensive guard for template rendering
        context = {}
    error = getattr(g, "_feature_state_error", None)
    return {
        "feature_states": context,
        "feature_registry": FEATURE_REGISTRY,
        "feature_state_error": error,
    }


@main_bp.route('/home')
def home():
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    role = session.get('role')
    username = session.get('username')
    if role == 'EMPLOYEE':
        return render_template(
            'employee_home.html',
            username=username,
            areas=EMPLOYEE_AREA_OPTIONS,
            user_role=role,
        )
    code_errors: list[dict[str, str]] = []
    if role == 'ADMIN':
        supabase_status: dict | None = None
        try:
            supabase_status = _summarize_supabase_status()
        except Exception as exc:  # pragma: no cover - defensive gathering
            supabase_status = {
                'error': str(exc),
                'tables': [],
            }

        if supabase_status:
            for table in supabase_status.get('tables', []):
                if (table or {}).get('status') == 'Unavailable':
                    code_errors.append(
                        {
                            'source': 'Supabase',
                            'name': table.get('name') or 'Tracked table',
                            'status': table.get('status') or 'Unavailable',
                            'message': table.get('error')
                            or 'Supabase reported this table as unavailable.',
                            'description': table.get('description'),
                        }
                    )
            if supabase_status.get('error'):
                code_errors.append(
                    {
                        'source': 'Supabase',
                        'name': 'Connection',
                        'status': 'Unavailable',
                        'message': supabase_status['error'],
                        'description': 'Spectra could not reach the configured Supabase project.',
                    }
                )

        feature_state_error = getattr(g, '_feature_state_error', None)
        if feature_state_error:
            code_errors.append(
                {
                    'source': 'Feature service',
                    'name': 'Feature state records',
                    'status': 'Unavailable',
                    'message': feature_state_error,
                    'description': 'Feature availability data may be incomplete until connectivity is restored.',
                }
            )

    return render_template(
        'home.html',
        username=username,
        user_role=role,
        code_errors=code_errors,
    )


def _role_required(allowed_roles: set[str]):
    def decorator(view):
        @wraps(view)
        def wrapped_view(**kwargs):
            role = session.get('role') or session.get('username')
            if role not in allowed_roles:
                abort(403)
            return view(**kwargs)

        return wrapped_view

    return decorator


admin_required = _role_required({'ADMIN'})
employee_portal_required = _role_required({'EMPLOYEE', 'ADMIN'})


@main_bp.route('/admin/employee-portal')
@admin_required
def admin_employee_portal():
    username = session.get('username')
    role = session.get('role') or 'ADMIN'
    return render_template(
        'employee_home.html',
        username=username,
        user_role=role,
        areas=EMPLOYEE_AREA_OPTIONS,
        employee_preview_active=True,
        employee_preview_return_url=request.args.get('return') or url_for('main.home'),
    )


def _require_authenticated_user() -> dict[str, str | None]:
    """Return the current session user or abort if unauthenticated."""

    if "username" not in session:
        abort(401, description="Authentication required")
    return {
        "user_id": session.get("user_id"),
        "username": session.get("username"),
        "role": session.get("role") or session.get("username"),
    }
def _resolve_user_display_name(
    user_id: object,
    user_lookup: dict[str, dict] | None = None,
    fallback: str | None = None,
) -> str | None:
    """Return a human friendly display name for the given Supabase user ID."""

    if user_id in (None, ""):
        return fallback

    lookup_key = str(user_id)
    if user_lookup:
        user_record = user_lookup.get(lookup_key) or user_lookup.get(user_id)
        if user_record:
            for candidate_key in ("label", "display_name", "username"):
                candidate_value = user_record.get(candidate_key)
                if candidate_value:
                    return candidate_value

    if fallback:
        return fallback

    return lookup_key


def _format_bug_report_response(
    record: dict | None,
    assignable_lookup: dict[str, dict] | None = None,
    reporter_display_name: str | None = None,
) -> dict:
    response = dict(record or {})

    resolved_reporter = _resolve_user_display_name(
        response.get("reporter_id"),
        assignable_lookup,
        reporter_display_name or response.get("reporter_name"),
    )

    response["reporter"] = resolved_reporter
    response["reporter_display_name"] = resolved_reporter

    return response


USER_ROLE_LABELS = {
    "ADMIN": "Administrator",
    "USER": "Standard User",
    "EMPLOYEE": "Employee",
    "VIEWER": "Viewer",
    "ANALYST": "Analyst",
    "MANAGER": "Manager",
}


USER_ROLE_CHOICES = [
    ("VIEWER", USER_ROLE_LABELS["VIEWER"]),
    ("ANALYST", USER_ROLE_LABELS["ANALYST"]),
    ("MANAGER", USER_ROLE_LABELS["MANAGER"]),
    ("EMPLOYEE", USER_ROLE_LABELS["EMPLOYEE"]),
    ("USER", USER_ROLE_LABELS["USER"]),
    ("ADMIN", USER_ROLE_LABELS["ADMIN"]),
]


TRACKED_SUPABASE_TABLES = {
    "aoi_reports": "AOI inspection uploads used across the AOI dashboards.",
    "fi_reports": "Final inspection data powering FI quality metrics.",
    "moat": "MOAT data feeding the integrated performance report.",
    "combined_reports": "Joined AOI/FI views surfaced in integrated analytics.",
    "app_users": "Application login accounts managed from this console.",
    "bug_reports": "In-app feedback collected to triage feature issues and bugs.",
    "defects": "Defect catalog entries referenced when analysing bug submissions.",
    "app_versions": "Release ledger aligning desktop and web deployments.",
    "moat_dpm": "MOAT defect-per-million submissions that drive the DPM dashboards.",
    "app_feature_states": "Feature flag states used to lock or reopen capabilities from bug triage.",
    "part_result_table": "Part-level AOI results providing deeper context for line quality checks.",
}


def _get_tracker():
    tracker = current_app.config.get("TRACKER")
    if not tracker:
        abort(503, description="Tracking service unavailable")
    return tracker


def _tracker_local_zone():
    tz_name = current_app.config.get("LOCAL_TIMEZONE", "America/Chicago")
    try:
        return ZoneInfo(tz_name)
    except Exception:  # pragma: no cover - fallback if zone not available
        return timezone.utc


def _tracker_parse_timestamp(value):
    if not value:
        return None
    cleaned = str(value).strip()
    if cleaned.endswith("Z"):
        cleaned = cleaned[:-1] + "+00:00"
    try:
        parsed = datetime.fromisoformat(cleaned)
    except ValueError:
        return None
    return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)


def _tracker_format_timestamp(value, local_zone=None):
    if not value:
        return None
    zone = local_zone or _tracker_local_zone()
    try:
        localized = value.astimezone(zone)
    except Exception:
        localized = value
    return localized.strftime("%Y-%m-%d %H:%M:%S %Z")


def _tracker_format_duration(seconds):
    if seconds is None:
        return "--"
    remaining = max(0.0, float(seconds))
    hours = int(remaining // 3600)
    minutes = int((remaining % 3600) // 60)
    secs = remaining % 60
    parts: list[str] = []
    if hours:
        parts.append(f"{hours}h")
    if minutes:
        parts.append(f"{minutes}m")
    parts.append(f"{secs:.0f}s")
    return " ".join(parts)


def _coerce_float(value):
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _derive_session_end(end_ts, events, *, local_zone=None):
    logout_events = [
        event
        for event in events
        if (event.get("name") or "").lower() in {"session_end", "logout"}
    ]
    derived_end_ts = None
    derived_end_display = None
    if logout_events:
        last_logout = logout_events[-1]
        derived_end_ts = last_logout.get("occurred")
        derived_end_display = last_logout.get("occurred_display")
    if derived_end_ts is None and end_ts is not None:
        derived_end_ts = end_ts
    if derived_end_ts is None and events:
        last_event = events[-1]
        derived_end_ts = last_event.get("occurred")
        derived_end_display = last_event.get("occurred_display")
    if derived_end_ts is not None and derived_end_display is None:
        derived_end_display = _tracker_format_timestamp(derived_end_ts, local_zone)
    return derived_end_ts, derived_end_display


def _calculate_session_duration(start_ts, derived_end_ts, events, stored_duration):
    duration_seconds = None
    if start_ts and derived_end_ts:
        try:
            duration_seconds = max(0.0, (derived_end_ts - start_ts).total_seconds())
        except Exception:
            duration_seconds = None
    if duration_seconds is None:
        duration_seconds = _coerce_float(stored_duration)
    if duration_seconds is None and start_ts:
        reference_ts = derived_end_ts
        if reference_ts is None and events:
            reference_ts = events[-1].get("occurred")
        if reference_ts:
            try:
                duration_seconds = max(
                    0.0, (reference_ts - start_ts).total_seconds()
                )
            except Exception:
                duration_seconds = None
    return duration_seconds


def _summarize_supabase_status():
    """Return metadata about the configured Supabase project."""

    supabase_url = current_app.config.get("SUPABASE_URL") or os.environ.get(
        "SUPABASE_URL"
    )
    supabase = current_app.config.get("SUPABASE")

    project_host = None
    if supabase_url:
        try:
            project_host = urlparse(supabase_url).netloc or supabase_url
        except Exception:  # pragma: no cover - defensive parsing
            project_host = supabase_url

    status = {
        "url": supabase_url,
        "project_host": project_host,
        "status": "Not configured" if not supabase else "Connected",
        "checked_at": datetime.utcnow(),
        "error": None,
        "tables": [],
    }

    if not supabase:
        return status

    for table, description in TRACKED_SUPABASE_TABLES.items():
        try:
            response = supabase.table(table).select("*").limit(1).execute()
            record_count = None
            if hasattr(response, "count") and response.count is not None:
                record_count = response.count
            elif response.data is not None:
                record_count = len(response.data)
            status["tables"].append(
                {
                    "name": table,
                    "description": description,
                    "status": "Available",
                    "records_previewed": record_count,
                    "error": None,
                }
            )
        except Exception as exc:  # pragma: no cover - missing tables or auth errors
            status["tables"].append(
                {
                    "name": table,
                    "description": description,
                    "status": "Unavailable",
                    "records_previewed": None,
                    "error": str(exc),
                }
            )

    if status["tables"] and all(
        entry["status"] != "Available" for entry in status["tables"]
    ):
        status["status"] = "No tracked tables reachable"

    return status


def _fetch_configured_users() -> tuple[list[dict], str | None]:
    """Return all known application users and any Supabase error."""

    users: list[dict] = []
    supabase_error: str | None = None

    try:
        supabase_users, supabase_error = fetch_app_users()
    except Exception as exc:  # pragma: no cover - defensive guard
        supabase_users = None
        supabase_error = str(exc)

    if supabase_users:
        for record in sorted(
            supabase_users,
            key=lambda item: (item.get("username") or "").casefold(),
        ):
            role_code = (record.get("role") or "USER").upper()
            users.append(
                {
                    "id": record.get("id"),
                    "username": record.get("username"),
                    "display_name": record.get("display_name") or record.get("username"),
                    "role": USER_ROLE_LABELS.get(role_code, role_code.title()),
                    "role_code": role_code,
                    "source": "Supabase",
                }
            )

    for username in sorted(auth_routes.ENVIRONMENT_USERS.keys()):
        role_code = username.upper()
        users.append(
            {
                "username": username,
                "display_name": username,
                "role": USER_ROLE_LABELS.get(role_code, "Standard User"),
                "role_code": role_code,
                "source": "Environment variables",
            }
        )

    return users, supabase_error


def _normalize_bug_id(value: object) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    return text or None


def _build_feature_cards(
    bug_records: list[dict] | None,
) -> tuple[list[dict[str, object]], dict[str, dict]]:
    cards: list[dict[str, object]] = []
    bug_lookup: dict[str, dict] = {}

    for bug in bug_records or []:
        bug_id = _normalize_bug_id(bug.get("id"))
        if bug_id is None:
            continue
        bug_lookup[bug_id] = bug

    for entry in FEATURE_REGISTRY:
        state = _compose_feature_state(entry["slug"])
        bug_id = _normalize_bug_id(state.get("bug_report_id"))
        bug_info = bug_lookup.get(bug_id) if bug_id is not None else None
        status = _normalize_feature_status(state.get("status"))

        cards.append(
            {
                "slug": entry["slug"],
                "label": state.get("label"),
                "description": state.get("description"),
                "status": status,
                "message": state.get("message"),
                "bug_report_id": bug_id,
                "bug_summary": (
                    f"#{bug_id} · {(bug_info.get('title') or 'Untitled')}" if bug_info else None
                ),
                "bug_status": (bug_info.get("status") if bug_info else None) or "",
                "updated_at": state.get("updated_at"),
            }
        )

    return cards, bug_lookup


def _build_bug_options(bug_records: list[dict] | None) -> list[dict[str, object]]:
    options: list[dict[str, object]] = []
    for record in bug_records or []:
        bug_id = _normalize_bug_id(record.get("id"))
        if bug_id is None:
            continue
        status = (record.get("status") or "open").replace("_", " ").title()
        title = record.get("title") or "Untitled"
        options.append(
            {
                "id": bug_id,
                "label": f"#{bug_id} — {title} ({status})",
                "status": record.get("status") or "open",
            }
        )

    options.sort(
        key=lambda item: (
            0 if item["status"] == "on_hold" else 1,
            0 if item["status"] == "open" else 1,
            str(item["id"]),
        )
    )
    return options


def _sync_feature_state_from_bug(record: dict | None) -> None:
    if not record:
        return

    bug_id = _normalize_bug_id(record.get("id"))
    if bug_id is None:
        return

    status = _normalize_feature_status(record.get("status"))
    linked_states, error = fetch_feature_states_for_bug(bug_id)
    if error:
        current_app.logger.warning("Failed to load feature state for bug %s: %s", bug_id, error)
        return

    if not linked_states:
        return

    status_label = status.replace("_", " ").title()
    for state in linked_states:
        slug = state.get("slug")
        if not slug:
            continue
        definition = _feature_definition(slug)
        label = definition.get("label", str(slug).replace("_", " ").title())
        if status == "on_hold":
            message = (
                f"{label} is temporarily locked while we investigate bug #{bug_id}."
            )
            _, update_error = upsert_feature_state(
                slug,
                status=FEATURE_STATUS_LOCKED,
                message=message,
                bug_report_id=bug_id,
            )
        else:
            message = (
                f"{label} has been reopened after bug #{bug_id} was marked {status_label}."
            )
            _, update_error = upsert_feature_state(
                slug,
                status=FEATURE_STATUS_AVAILABLE,
                message=message,
                bug_report_id=bug_id,
            )
        if update_error:
            current_app.logger.warning(
                "Failed to update feature '%s' from bug %s: %s", slug, bug_id, update_error
            )

@main_bp.route('/admin')
@admin_required
def admin_panel():
    active_tab = request.args.get('tab') or 'overview'
    supabase_status = _summarize_supabase_status()
    users, supabase_user_error = _fetch_configured_users()
    bug_records: list[dict] | None = None
    feature_bug_error: str | None = None
    try:
        bug_records, feature_bug_error = fetch_bug_reports()
    except Exception as exc:  # pragma: no cover - defensive safeguard
        bug_records = []
        feature_bug_error = str(exc)

    app_versions_data: list[dict] | None = None
    app_version_error: str | None = None
    try:
        app_versions_data, app_version_error = fetch_app_versions()
    except Exception as exc:  # pragma: no cover - defensive safeguard
        app_versions_data = []
        app_version_error = str(exc)

    app_versions = app_versions_data or []
    latest_version = app_versions[0] if app_versions else None

    feature_cards, _ = _build_feature_cards(bug_records)
    feature_state_error = getattr(g, '_feature_state_error', None)
    bug_options = _build_bug_options(bug_records)
    overview = {
        "supabase_status": supabase_status["status"],
        "user_count": len(users),
        "tracked_tables": supabase_status["tables"],
        "last_checked": supabase_status["checked_at"],
        "latest_app_version": latest_version,
    }
    missing_feeds = {
        "roles": (
            "Role definitions are currently hard-coded in the UI. Track them in "
            "a Supabase table such as `roles` to enable real-time edits."
        ),
        "system_policies": (
            "System policy settings are applied manually in environment "
            "configuration. Persist them in Supabase storage or a configuration "
            "service to manage them from this console."
        ),
    }
    if not app_versions:
        missing_feeds["app_versions"] = (
            "Capture release numbers, download links, and checksums in a Supabase "
            "table named `app_versions` so the desktop build stays aligned with "
            "the web deployment, even if binaries live on a shared drive."
        )
    if supabase_user_error:
        flash(supabase_user_error, "warning")
    if app_version_error:
        flash(app_version_error, "warning")

    return render_template(
        'admin.html',
        username=session.get('username'),
        supabase_status=supabase_status,
        users=users,
        overview=overview,
        missing_feeds=missing_feeds,
        user_role_choices=USER_ROLE_CHOICES,
        feature_cards=feature_cards,
        feature_bug_options=bug_options,
        feature_state_error=feature_state_error,
        feature_bug_error=feature_bug_error,
        app_versions=app_versions,
        app_version_error=app_version_error,
        active_tab=active_tab,
    )


@main_bp.route('/admin/data-sources', methods=['POST'])
@admin_required
def admin_data_sources_action():
    requested_action = request.form.get('action') or 'update'
    flash(
        (
            "Data source management changes (" + requested_action + ") are not "
            "automated yet. Track requested connections in a Supabase table such "
            "as `data_sources` or update the deployment configuration directly."
        ),
        'info',
    )
    return redirect(url_for('main.admin_panel'))


@main_bp.route('/admin/users', methods=['POST'])
@admin_required
def admin_user_action():
    action = (request.form.get('action') or 'invite').lower()
    username = (request.form.get('username') or '').strip()
    role = (request.form.get('role') or 'USER').strip().upper()
    temporary_password = (request.form.get('temporary_password') or '').strip()
    display_name = (request.form.get('display_name') or '').strip()

    if role not in USER_ROLE_LABELS:
        role = 'USER'

    if action == 'invite':
        if not username:
            flash('Enter a username to create an account.', 'error')
        elif not temporary_password:
            flash('Provide a temporary password for the new account.', 'error')
        else:
            existing, error = fetch_app_user_credentials(username)
            if error:
                flash(error, 'error')
            elif existing:
                flash(f"User '{username}' already exists.", 'warning')
            else:
                payload = {
                    'username': username,
                    'display_name': display_name or username,
                    'role': role,
                    'password_hash': generate_password_hash(temporary_password),
                }
                inserted, error = insert_app_user(payload)
                if error:
                    flash(error, 'error')
                else:
                    flash(
                        f"User '{username}' has been created with the provided temporary password.",
                        'success',
                    )
    elif action in {'remove', 'delete', 'deactivate'}:
        if not username:
            flash('Specify a username to remove.', 'error')
        else:
            existing, error = fetch_app_user_credentials(username)
            if error:
                flash(error, 'error')
            elif not existing:
                flash(f"User '{username}' was not found in Supabase.", 'warning')
            else:
                deleted, error = delete_app_user(existing['id'])
                if error:
                    flash(error, 'error')
                elif deleted:
                    flash(f"User '{username}' has been removed.", 'success')
                else:
                    flash(f"No changes were applied for '{username}'.", 'info')
    else:
        flash(f"Unrecognised action '{action}'.", 'error')

    return redirect(url_for('main.admin_panel'))


@main_bp.route('/admin/features', methods=['POST'])
@admin_required
def admin_feature_action():
    slug = (request.form.get('slug') or '').strip()
    if not slug or slug not in FEATURE_DEFINITIONS:
        flash('Select a valid application feature to update.', 'error')
        return redirect(url_for('main.admin_panel', tab='features'))

    status = _normalize_feature_status(request.form.get('status'))
    message = (request.form.get('message') or '').strip()
    raw_bug_id = (request.form.get('bug_report_id') or '').strip()
    bug_id = _normalize_bug_id(raw_bug_id)

    if raw_bug_id and bug_id is None:
        flash('Enter a valid bug report identifier to link.', 'error')
        return redirect(url_for('main.admin_panel', tab='features'))

    definition = _feature_definition(slug)
    label = definition.get('label', slug.replace('_', ' ').title())

    if status == FEATURE_STATUS_AVAILABLE and not message:
        if bug_id is not None:
            message = f"{label} has been reopened after bug #{bug_id} was resolved."
        else:
            message = f"{label} is available."
    elif status != FEATURE_STATUS_AVAILABLE and not message:
        if bug_id is not None:
            message = f"{label} is temporarily locked while we investigate bug #{bug_id}."
        else:
            message = definition.get('default_message') or f"{label} is temporarily unavailable."

    _, error = upsert_feature_state(
        slug,
        status=status,
        message=message,
        bug_report_id=bug_id,
    )
    if error:
        flash(error, 'error')
    else:
        if status == FEATURE_STATUS_AVAILABLE:
            flash(f"{label} marked available.", 'success')
        else:
            flash(f"{label} locked for maintenance.", 'info')

    return redirect(url_for('main.admin_panel', tab='features'))


@main_bp.route('/aoi_reports', methods=['GET'])
def get_aoi_reports():
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    data, error = fetch_aoi_reports()
    if error:
        abort(500, description=error)
    return jsonify(data)


@main_bp.route('/aoi_reports', methods=['POST'])
@admin_required
def add_aoi_report():
    payload = request.get_json() or {}
    data, error = insert_aoi_report(payload)
    if error:
        abort(500, description=error)
    return jsonify(data), 201


@main_bp.route('/bug-reports', methods=['POST'])
@main_bp.route('/bug_reports', methods=['POST'])
def submit_bug_report():
    user = _require_authenticated_user()

    if request.is_json:
        payload = request.get_json() or {}
    else:
        payload = request.form.to_dict()

    title = (payload.get('title') or '').strip()
    description = (payload.get('description') or '').strip()
    priority = (payload.get('priority') or '').strip() or None

    if not title or not description:
        abort(400, description='Both title and description are required.')

    reporter_display_name = user.get('username')
    reporter_identifier: str | None = None
    reporter_auth_identifier: str | None = None

    username = user.get('username')
    if username:
        supabase_account, fetch_error = fetch_app_user_credentials(username)
        if fetch_error:  # pragma: no cover - logging only
            current_app.logger.warning(
                "Failed to load Supabase account for %s: %s", username, fetch_error
            )
        if supabase_account:
            reporter_display_name = (
                supabase_account.get('display_name') or reporter_display_name
            )

            account_identifier = supabase_account.get('id')
            if account_identifier not in (None, ''):
                reporter_identifier = str(account_identifier)

            auth_user_id = supabase_account.get('auth_user_id')
            if auth_user_id:
                reporter_auth_identifier = str(auth_user_id)
            else:
                # Some records hydrate the linked auth.users row under ``auth_user``.
                auth_user = supabase_account.get('auth_user')
                if isinstance(auth_user, dict):
                    linked_id = auth_user.get('id') or auth_user.get('uuid')
                    if linked_id:
                        reporter_auth_identifier = str(linked_id)

            if not reporter_auth_identifier:
                for key in (
                    'auth_user_uuid',
                    'auth_uuid',
                ):
                    candidate = supabase_account.get(key)
                    if candidate:
                        reporter_auth_identifier = str(candidate)
                        break

            if reporter_identifier is None and reporter_auth_identifier is not None:
                reporter_identifier = reporter_auth_identifier

    record = {
        'title': title,
        'description': description,
        'priority': priority,
        'status': payload.get('status') or 'open',
        'reporter_name': reporter_display_name,
    }

    if reporter_identifier is None:
        session_user_id = user.get('user_id')
        if session_user_id not in (None, ''):
            reporter_identifier = str(session_user_id)

    if reporter_identifier is not None:
        record['reporter_id'] = str(reporter_identifier)

    created, error = insert_bug_report(record)
    if error:
        abort(503, description=error)
    if not created:
        abort(500, description='Bug report could not be created.')

    response_body = _format_bug_report_response(
        created[0],
        reporter_display_name=reporter_display_name,
    )
    return jsonify(response_body), 201


@main_bp.route('/admin/bug-reports', methods=['GET'])
@admin_required
def list_bug_reports():
    status_filter = request.args.get('status')
    assignee_filter = request.args.get('assignee_id')

    filters = {}
    if status_filter:
        filters['status'] = status_filter
    if assignee_filter:
        filters['assignee_id'] = assignee_filter

    reports, error = fetch_bug_reports(filters or None)
    if error:
        abort(503, description=error)

    return jsonify({'bug_reports': reports or []})


@main_bp.route('/admin/bug-reports/<int:report_id>', methods=['PATCH'])
@admin_required
def update_bug_report(report_id: int):
    payload = request.get_json(silent=True) or {}

    allowed_statuses = {'open', 'in_progress', 'resolved', 'on_hold'}
    updates: dict[str, object] = {}

    if 'status' in payload:
        status_value = (payload.get('status') or '').strip().lower()
        if not status_value:
            abort(400, description='Status is required when provided.')
        if status_value not in allowed_statuses:
            abort(400, description='Invalid status selection.')
        updates['status'] = status_value

    if 'assignee_id' in payload:
        assignee_value = payload.get('assignee_id')
        updates['assignee_id'] = assignee_value if assignee_value not in (None, '') else None

    if 'priority' in payload:
        priority_value = (payload.get('priority') or '').strip()
        updates['priority'] = priority_value or None

    if 'notes' in payload:
        notes_value = payload.get('notes')
        if isinstance(notes_value, str):
            notes_value = notes_value.strip()
        updates['notes'] = notes_value or None

    if not updates:
        abort(400, description='No valid fields to update.')

    updated, error = update_bug_report_status(report_id, updates)
    if error:
        error_message = str(error)
        status_code = 400 if 'No updates supplied' in error_message else 503
        current_app.logger.error(
            'Failed to update bug report %s: %s', report_id, error_message
        )
        payload = {
            'error': 'update_failed',
            'description': error_message,
        }
        return make_response(jsonify(payload), status_code)

    if not updated:
        abort(404, description='Bug report not found.')

    _sync_feature_state_from_bug(updated[0])

    return jsonify({'bug_report': updated[0]})


def _normalize_employee_date(value: str | None) -> str | None:
    if not value:
        return None
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y'):
        try:
            return datetime.strptime(value, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _prepare_employee_aoi_record(
    payload: dict[str, str],
) -> tuple[dict[str, str], dict[str, str], str | None]:
    errors: dict[str, str] = {}
    record: dict[str, str] = {}

    date_raw = (payload.get('date') or '').strip()
    if not date_raw:
        errors['date'] = 'Date is required.'
    else:
        normalized_date = _normalize_employee_date(date_raw)
        if not normalized_date:
            errors['date'] = 'Enter a valid date.'
        else:
            record['Date'] = normalized_date

    text_fields = (
        ('shift', 'Shift'),
        ('operator', 'Operator'),
        ('customer', 'Customer'),
        ('program', 'Program'),
        ('assembly', 'Assembly'),
        ('job_number', 'Job Number'),
    )
    for field, column in text_fields:
        value = (payload.get(field) or '').strip()
        if not value:
            errors[field] = 'This field is required.'
        else:
            record[column] = value

    signature_value = str(payload.get('operator_signature_acknowledged') or '').strip().lower()
    if signature_value not in {'true', '1', 'yes', 'on'}:
        errors['operator_signature_acknowledged'] = 'Confirm the operator signature before submitting.'

    numeric_fields = (
        ('quantity_inspected', 'Quantity Inspected'),
        ('quantity_rejected', 'Quantity Rejected'),
    )
    for field, column in numeric_fields:
        value = (payload.get(field) or '').strip()
        if not value:
            errors[field] = 'This field is required.'
            continue
        try:
            number = int(value)
            if number < 0:
                raise ValueError
        except ValueError:
            errors[field] = 'Enter a whole number of 0 or greater.'
            continue
        record[column] = str(number)

    rev_value = (payload.get('rev') or '').strip()
    if rev_value:
        record['Rev'] = rev_value

    sheet_key = (payload.get('inspection_type') or '').strip().upper()
    sheet_label = None
    if sheet_key:
        sheet_label = EMPLOYEE_SHEET_LABELS.get(sheet_key)
        if not sheet_label:
            errors['inspection_type'] = 'Select a valid inspection data sheet.'
    else:
        errors['inspection_type'] = 'Select a valid inspection data sheet.'

    notes_value = (payload.get('notes') or '').strip()

    rejection_details_raw = payload.get('rejection_details')
    rejection_entries: list[dict[str, object]] = []
    if isinstance(rejection_details_raw, str) and rejection_details_raw.strip():
        try:
            parsed = json.loads(rejection_details_raw)
        except json.JSONDecodeError:
            errors['rejection_details'] = 'Enter valid rejection detail rows.'
        else:
            if isinstance(parsed, list):
                rejection_entries = [entry for entry in parsed if isinstance(entry, dict)]
                if len(rejection_entries) != len(parsed):
                    errors['rejection_details'] = 'Enter valid rejection detail rows.'
            elif parsed:
                errors['rejection_details'] = 'Enter valid rejection detail rows.'
    elif isinstance(rejection_details_raw, list):
        rejection_entries = [entry for entry in rejection_details_raw if isinstance(entry, dict)]
        if len(rejection_entries) != len(rejection_details_raw):
            errors['rejection_details'] = 'Enter valid rejection detail rows.'
    elif rejection_details_raw not in (None, ''):
        errors['rejection_details'] = 'Enter valid rejection detail rows.'

    formatted_rejections: list[str] = []
    if rejection_entries:
        detail_errors = False
        for entry in rejection_entries:
            ref = str(entry.get('ref', '') or '').strip()
            reason = str(entry.get('reason', '') or '').strip()
            quantity_raw = entry.get('quantity')
            try:
                quantity = int(quantity_raw)
            except (TypeError, ValueError):
                quantity = None
            if not ref or not reason or not quantity or quantity <= 0:
                detail_errors = True
                break
            reason_normalized = ' '.join(reason.split())
            formatted_rejections.append(f'{ref} - {reason_normalized} ({quantity})')
        if detail_errors:
            errors['rejection_details'] = (
                'Rejection detail entries must include a reference, reason, '
                'and quantity of 1 or more.'
            )

    additional_parts = []
    if sheet_label:
        additional_parts.append(f'{sheet_label} submission')
    if formatted_rejections:
        additional_parts.append(', '.join(formatted_rejections))
    if notes_value:
        additional_parts.append(notes_value)
    if additional_parts:
        record['Additional Information'] = ' | '.join(additional_parts)

    return record, errors, sheet_label


def _build_defect_response():
    defects, error = fetch_defect_catalog()
    payload = {'defects': defects or []}
    if error:
        payload['error'] = error
    status = 200 if not error else 503
    return payload, status


@main_bp.route('/employee/defects', methods=['GET'])
@employee_portal_required
def employee_list_defects():
    payload, status = _build_defect_response()
    return jsonify(payload), status


@main_bp.route('/employee/aoi_reports', methods=['POST'])
@employee_portal_required
def employee_add_aoi_report():
    payload = request.get_json() or {}
    record, errors, sheet_label = _prepare_employee_aoi_record(payload)
    if errors:
        return jsonify({'errors': errors}), 400
    _, error = insert_aoi_report(record)
    if error:
        return jsonify({'errors': {'base': error}}), 500
    message = 'AOI report submitted successfully.'
    if sheet_label:
        message = f'{sheet_label} submission saved successfully.'
    return jsonify({'message': message}), 201


@main_bp.route('/aoi_reports/upload', methods=['POST'])
@admin_required
def upload_aoi_reports():
    """Upload a CSV file of AOI reports."""
    uploaded = request.files.get('file')
    if not uploaded or uploaded.filename == '':
        abort(400, description='No file provided')

    stream = io.StringIO(uploaded.stream.read().decode('utf-8'))
    reader = csv.DictReader(stream)
    ordered_columns = [
        'Date',
        'Shift',
        'Operator',
        'Customer',
        'Program',
        'Assembly',
        'Rev',
        'Job Number',
        'Quantity Inspected',
        'Quantity Rejected',
        'Additional Information',
    ]
    optional_columns = ['Rev', 'Additional Information']
    required_columns = [
        column for column in ordered_columns if column not in optional_columns
    ]
    missing, unexpected, out_of_order, header_map = _compare_headers(
        reader.fieldnames, ordered_columns, optional_columns
    )
    if missing or unexpected or out_of_order:
        expected_order_display = [
            column
            for column in ordered_columns
            if column not in optional_columns or header_map.get(column)
        ]
        message_parts = [
            f"Missing columns: {', '.join(missing) if missing else 'none'}",
            f"Unexpected columns: {', '.join(unexpected) if unexpected else 'none'}",
            f"Columns out of order: {', '.join(out_of_order) if out_of_order else 'none'}",
            f"Column order should be: {', '.join(expected_order_display)}",
        ]
        abort(400, description='; '.join(message_parts))
    rows = []
    rows_with_missing = []
    for idx, row in enumerate(reader, start=2):
        if not any((value or '').strip() for value in row.values()):
            continue
        # Copy required columns (including 'Program') for each record
        current = {}
        missing_cols = []
        for col in required_columns:
            source = header_map.get(col) or col
            value = row.get(source, '')
            if value is None:
                value = ''
            value = value.strip()
            if not value:
                missing_cols.append(col)
            current[col] = value
        for col in optional_columns:
            source = header_map.get(col)
            if not source:
                continue
            value = row.get(source, '')
            if value is None:
                value = ''
            current[col] = value.strip()
        if missing_cols:
            rows_with_missing.append((idx, missing_cols))
            continue
        date_str = current.get('Date')
        if date_str:
            try:
                dt = datetime.strptime(date_str, '%m/%d/%Y')
                current['Date'] = dt.date().isoformat()
            except ValueError:
                pass
        rows.append(current)
    if rows_with_missing:
        details = '; '.join(
            f"Row {row_num}: {', '.join(columns)}" for row_num, columns in rows_with_missing
        )
        abort(400, description=f"Missing required data in rows - {details}")
    if not rows:
        return jsonify({'inserted': 0}), 200

    data, error = insert_aoi_reports_bulk(rows)
    if error:
        abort(500, description=error)
    return jsonify({'inserted': len(rows)}), 201


@main_bp.route('/fi_reports', methods=['GET'])
def get_fi_reports():
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    data, error = fetch_fi_reports()
    if error:
        abort(500, description=error)
    return jsonify(data)


@main_bp.route('/fi_reports', methods=['POST'])
@admin_required
def add_fi_report():
    payload = request.get_json() or {}
    data, error = insert_fi_report(payload)
    if error:
        abort(500, description=error)
    return jsonify(data), 201


@main_bp.route('/fi_reports/upload', methods=['POST'])
@admin_required
def upload_fi_reports():
    """Upload a CSV file of FI reports."""
    uploaded = request.files.get('file')
    if not uploaded or uploaded.filename == '':
        abort(400, description='No file provided')

    stream = io.StringIO(uploaded.stream.read().decode('utf-8'))
    reader = csv.DictReader(stream)
    ordered_columns = [
        'Date',
        'Shift',
        'Operator',
        'Customer',
        'Assembly',
        'Rev',
        'Job Number',
        'Quantity Inspected',
        'Quantity Rejected',
        'Additional Information',
    ]
    optional_columns = ['Rev', 'Additional Information']
    required_columns = [
        column for column in ordered_columns if column not in optional_columns
    ]
    missing, unexpected, out_of_order, header_map = _compare_headers(
        reader.fieldnames, ordered_columns, optional_columns
    )
    if missing or unexpected or out_of_order:
        expected_order_display = [
            column
            for column in ordered_columns
            if column not in optional_columns or header_map.get(column)
        ]
        message_parts = [
            f"Missing columns: {', '.join(missing) if missing else 'none'}",
            f"Unexpected columns: {', '.join(unexpected) if unexpected else 'none'}",
            f"Columns out of order: {', '.join(out_of_order) if out_of_order else 'none'}",
            f"Column order should be: {', '.join(expected_order_display)}",
        ]
        abort(400, description='; '.join(message_parts))
    rows = []
    rows_with_missing = []
    for idx, row in enumerate(reader, start=2):
        if not any((value or '').strip() for value in row.values()):
            continue
        current = {}
        missing_cols = []
        for col in required_columns:
            source = header_map.get(col) or col
            value = row.get(source, '')
            if value is None:
                value = ''
            value = value.strip()
            if not value:
                missing_cols.append(col)
            current[col] = value
        for col in optional_columns:
            source = header_map.get(col)
            if not source:
                continue
            value = row.get(source, '')
            if value is None:
                value = ''
            current[col] = value.strip()
        if missing_cols:
            rows_with_missing.append((idx, missing_cols))
            continue
        rows.append(current)
    if rows_with_missing:
        details = '; '.join(
            f"Row {row_num}: {', '.join(columns)}" for row_num, columns in rows_with_missing
        )
        abort(400, description=f"Missing required data in rows - {details}")
    if not rows:
        return jsonify({'inserted': 0}), 200

    inserted = 0
    for r in rows:
        _, err = insert_fi_report(r)
        if err:
            abort(500, description=err)
        inserted += 1
    return jsonify({'inserted': inserted}), 201


@main_bp.route('/dpm_reports/upload', methods=['POST'])
@admin_required
@feature_required('analysis_dpm')
def upload_dpm_reports():
    """Upload an XLS or XLSX DPM report and store rows in the MOAT DPM table."""

    uploaded = request.files.get('file')
    if not uploaded or uploaded.filename == '':
        abort(400, description='No file provided')

    base = os.path.splitext(os.path.basename(uploaded.filename))[0]
    m = re.match(
        r"^DPMReportControl\s+(\d{4}-\d{2}-\d{2})(?:\s+to\s+(\d{4}-\d{2}-\d{2}))?\s+(L\w+)$",
        base,
        re.IGNORECASE,
    )
    if not m:
        abort(
            400,
            description=(
                'Filename must be "DPMReportControl YYYY-MM-DD LX" or '
                '"DPMReportControl YYYY-MM-DD to YYYY-MM-DD LX"'
            ),
        )

    start_date, end_date, line = m.groups()
    report_date = None

    try:
        uploaded.stream.seek(0)
        if uploaded.filename.lower().endswith('.xls'):
            book = xlrd.open_workbook(file_contents=uploaded.stream.read())
            sheet = book.sheet_by_index(0)

            def cell(r, c):
                try:
                    return sheet.cell_value(r - 1, c - 1)
                except IndexError:
                    return None

            raw_date = cell(2, 1)
        else:
            wb = load_workbook(uploaded.stream, data_only=True)
            sheet = wb.active

            def cell(r, c):
                return sheet.cell(row=r, column=c).value

            raw_date = cell(2, 1)

        if raw_date:
            if isinstance(raw_date, datetime):
                report_date = raw_date.date().isoformat()
            elif isinstance(raw_date, date):
                report_date = raw_date.isoformat()
            elif isinstance(raw_date, str):
                try:
                    report_date = (
                        datetime.strptime(raw_date.strip(), "%m/%d/%Y").date().isoformat()
                    )
                except ValueError:
                    pass
    except Exception as exc:
        abort(400, description=f'Failed to read Excel file: {exc}')

    if not report_date:
        report_date = start_date

    rows = []
    row_idx = 7
    while True:
        model = cell(row_idx, 2)
        if model in (None, ''):
            row_idx += 1
            continue
        if str(model).strip().lower() == 'total':
            break
        rows.append({
            'Model Name': model,
            'Total Boards': _coerce_number(cell(row_idx, 3)),
            'Windows per board': _coerce_number(cell(row_idx, 4)),
            'Total Windows': _coerce_number(cell(row_idx, 5)),
            'NG Windows': _coerce_number(cell(row_idx, 6)),
            'DPM': _coerce_number(cell(row_idx, 7)),
            'FalseCall Windows': _coerce_number(cell(row_idx, 8)),
            'FC DPM': _coerce_number(cell(row_idx, 9)),
            'Report Date': report_date,
            'Line': line,
        })
        row_idx += 1

    if not rows:
        return jsonify({'inserted': 0}), 200

    _, error = insert_moat_dpm_bulk(rows)
    if error:
        abort(500, description=error)

    return jsonify({'inserted': len(rows)}), 201


@main_bp.route('/ppm_reports/upload', methods=['POST'])
@admin_required
@feature_required('analysis_ppm')
def upload_ppm_reports():
    """Upload an XLS or XLSX PPM report and store rows in the MOAT table."""
    uploaded = request.files.get('file')
    if not uploaded or uploaded.filename == '':
        abort(400, description='No file provided')

    base = os.path.splitext(os.path.basename(uploaded.filename))[0]
    m = re.match(
        r"^PPMReportControl\s+(\d{4}-\d{2}-\d{2})(?:\s+to\s+(\d{4}-\d{2}-\d{2}))?\s+(L\w+)$",
        base,
        re.IGNORECASE,
    )
    if not m:
        abort(
            400,
            description=(
                'Filename must be "PPMReportControl YYYY-MM-DD LX" or '
                '"PPMReportControl YYYY-MM-DD to YYYY-MM-DD LX"'
            ),
        )
    start_date, end_date, line = m.groups()
    report_date = None

    try:
        uploaded.stream.seek(0)
        if uploaded.filename.lower().endswith('.xls'):
            book = xlrd.open_workbook(file_contents=uploaded.stream.read())
            sheet = book.sheet_by_index(0)

            def cell(r, c):
                try:
                    return sheet.cell_value(r - 1, c - 1)
                except IndexError:
                    return None

            raw_date = cell(2, 1)
        else:
            wb = load_workbook(uploaded.stream, data_only=True)
            sheet = wb.active

            def cell(r, c):
                return sheet.cell(row=r, column=c).value

            raw_date = cell(2, 1)

        if raw_date:
            if isinstance(raw_date, datetime):
                report_date = raw_date.date().isoformat()
            elif isinstance(raw_date, date):
                report_date = raw_date.isoformat()
            elif isinstance(raw_date, str):
                try:
                    report_date = datetime.strptime(raw_date.strip(), "%m/%d/%Y").date().isoformat()
                except ValueError:
                    pass
    except Exception as exc:
        abort(400, description=f'Failed to read Excel file: {exc}')

    if not report_date:
        report_date = start_date

    rows = []
    row_idx = 7
    while True:
        model = cell(row_idx, 2)
        if model in (None, ''):
            row_idx += 1
            continue
        if str(model).strip().lower() == 'total':
            break
        rows.append({
            'Model Name': model,
            'Total Boards': cell(row_idx, 3) or 0,
            'Total Parts/Board': cell(row_idx, 4) or 0,
            'Total Parts': cell(row_idx, 5) or 0,
            'NG Parts': cell(row_idx, 6) or 0,
            'NG PPM': cell(row_idx, 7) or 0,
            'FalseCall Parts': cell(row_idx, 8) or 0,
            'FalseCall PPM': cell(row_idx, 9) or 0,
            'Report Date': report_date,
            'Line': line,
        })
        row_idx += 1

    if not rows:
        return jsonify({'inserted': 0}), 200

    _, error = insert_moat_bulk(rows)
    if error:
        abort(500, description=error)

    return jsonify({'inserted': len(rows)}), 201


@main_bp.route('/moat_dpm', methods=['GET'])
@feature_required('analysis_dpm')
def get_moat_dpm_data():
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    data, error = fetch_moat_dpm()
    if error:
        abort(500, description=error)
    return jsonify(data)


@main_bp.route('/moat', methods=['GET'])
def get_moat_data():
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    data, error = fetch_moat()
    if error:
        abort(500, description=error)
    return jsonify(data)


@main_bp.route('/moat', methods=['POST'])
@admin_required
def add_moat_data():
    payload = request.get_json() or {}
    data, error = insert_moat(payload)
    if error:
        abort(500, description=error)
    return jsonify(data), 201


@main_bp.route('/moat_preview', methods=['GET'])
def moat_preview():
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    source = (request.args.get('source') or '').strip().lower()
    fetcher = fetch_recent_moat_dpm if source == 'dpm' else fetch_recent_moat
    data, error = fetcher()
    if error:
        abort(500, description=error)
    if not data:
        return jsonify({
            "models": [],
            "avg_false_calls": [],
            "overall_avg": 0,
            "labels": [],
            "values": [],
            "start_date": None,
            "end_date": None,
        })
    from collections import defaultdict

    grouped = defaultdict(lambda: {"falsecall": 0, "boards": 0})
    date_values: list[date] = []
    for row in data:
        fc = (
            row.get('FalseCall Windows')
            or row.get('falsecall_windows')
            or row.get('FalseCall Parts')
            or row.get('falsecall_parts')
            or 0
        )
        boards = row.get('Total Boards') or row.get('total_boards') or 0
        model = row.get('Model Name') or row.get('model_name') or 'Unknown'
        report_date = _parse_date(row.get('Report Date') or row.get('report_date'))
        if report_date:
            date_values.append(report_date)
        grouped[model]["falsecall"] += float(fc)
        grouped[model]["boards"] += float(boards)

    models, averages = [], []
    total_avg = 0.0
    for model, vals in grouped.items():
        boards = vals["boards"]
        avg = (vals["falsecall"] / boards) if boards else 0.0
        models.append(model)
        averages.append(avg)
        total_avg += avg

    overall_avg = total_avg / len(averages) if averages else 0.0
    start_date = min(date_values).isoformat() if date_values else None
    end_date = max(date_values).isoformat() if date_values else None
    return jsonify({
        "models": models,
        "avg_false_calls": averages,
        "overall_avg": overall_avg,
        "labels": models,
        "values": averages,
        "start_date": start_date,
        "end_date": end_date,
    })


def _yield_preview(fetch_func):
    """Return yield percentages for the last 7 recorded days."""
    if 'username' not in session:
        return redirect(url_for('auth.login'))

    data, error = fetch_func()
    if error:
        abort(500, description=error)

    from datetime import datetime, timedelta
    from collections import defaultdict

    today = datetime.utcnow().date()
    start = today - timedelta(days=6)
    agg = defaultdict(lambda: {'accepted': 0, 'rejected': 0})

    for row in data:
        d = _parse_date(row.get('Date') or row.get('date'))
        if not d or d < start or d > today:
            continue
        inspected = int(row.get('Quantity Inspected') or row.get('quantity_inspected') or 0)
        rejected = int(row.get('Quantity Rejected') or row.get('quantity_rejected') or 0)
        accepted = inspected - rejected
        if accepted < 0:
            accepted = 0
        agg[d]['accepted'] += accepted
        agg[d]['rejected'] += rejected

    dates = sorted(agg.keys())
    yields = []
    for d in dates:
        a = agg[d]['accepted']
        r = agg[d]['rejected']
        tot = a + r
        y = (a / tot * 100) if tot else 0
        yields.append(y)

    avg_yield = sum(yields) / len(yields) if yields else 0
    labels = [d.isoformat() for d in dates]
    start_date = labels[0] if labels else None
    end_date = labels[-1] if labels else None

    return jsonify({
        'labels': labels,
        'values': yields,
        'yields': yields,
        'avg_yield': avg_yield,
        'start_date': start_date,
        'end_date': end_date,
    })


@main_bp.route('/aoi_preview', methods=['GET'])
def aoi_preview():
    return _yield_preview(fetch_aoi_reports)


@main_bp.route('/fi_preview', methods=['GET'])
def fi_preview():
    return _yield_preview(fetch_fi_reports)


@main_bp.route('/daily_reports_preview', methods=['GET'])
def daily_reports_preview():
    """Return recent AOI/FI daily yields for the dashboard preview."""
    if 'username' not in session:
        return redirect(url_for('auth.login'))

    rows, error = fetch_combined_reports()
    if error:
        abort(500, description=error)

    from datetime import datetime, timedelta
    from collections import defaultdict

    today = datetime.utcnow().date()
    start = today - timedelta(days=6)
    daily = defaultdict(lambda: {"inspected": 0.0, "rejected": 0.0})

    for row in rows or []:
        day = (
            _parse_date(row.get('aoi_Date'))
            or _parse_date(row.get('Date'))
            or _parse_date(row.get('fi_Date'))
        )
        if not day or day < start or day > today:
            continue
        inspected = float(
            row.get('aoi_Quantity Inspected')
            or row.get('Quantity Inspected')
            or row.get('fi_Quantity Inspected')
            or 0
        )
        rejected = float(
            row.get('aoi_Quantity Rejected')
            or row.get('Quantity Rejected')
            or row.get('fi_Quantity Rejected')
            or 0
        )
        daily[day]['inspected'] += inspected
        daily[day]['rejected'] += rejected

    dates = sorted(daily.keys())
    labels = [d.isoformat() for d in dates]
    values = []
    for d in dates:
        inspected = daily[d]['inspected']
        rejected = daily[d]['rejected']
        accepted = inspected - rejected
        yield_pct = (accepted / inspected * 100.0) if inspected else 0.0
        values.append(yield_pct)

    avg_yield = sum(values) / len(values) if values else 0.0
    start_date = labels[0] if labels else None
    end_date = labels[-1] if labels else None

    return jsonify({
        'labels': labels,
        'values': values,
        'avg_yield': avg_yield,
        'start_date': start_date,
        'end_date': end_date,
    })


@main_bp.route('/bug_reports_preview', methods=['GET'])
def bug_reports_preview():
    """Return recent bug report counts grouped by status."""

    if 'username' not in session:
        return redirect(url_for('auth.login'))

    rows, error = fetch_bug_reports()
    if error:
        abort(500, description=error)

    today = datetime.now(timezone.utc).date()
    window_days = 7
    start = today - timedelta(days=window_days - 1)

    normalized_counts: Counter[str] = Counter()
    display_labels: dict[str, str] = {}
    observed_dates: list[date] = []

    for row in rows or []:
        created = _parse_date(row.get('created_at'))
        if not created or created < start or created > today:
            continue

        observed_dates.append(created)

        raw_status = str(row.get('status') or '').strip() or 'Unknown'
        normalized = raw_status.lower()
        display = raw_status if raw_status.lower() != raw_status else raw_status.title()

        normalized_counts[normalized] += 1
        display_labels.setdefault(normalized, display)

    total_reports = sum(normalized_counts.values())

    resolved_statuses = {'resolved', 'closed', 'done', 'completed', 'fixed'}
    resolved_reports = sum(
        count for status, count in normalized_counts.items() if status in resolved_statuses
    )
    active_reports = total_reports - resolved_reports

    ordered_statuses = sorted(
        normalized_counts.items(), key=lambda item: (-item[1], display_labels.get(item[0], ''))
    )

    labels = [display_labels.get(status, status.title()) for status, _ in ordered_statuses]
    values = [count for _, count in ordered_statuses]

    status_counts = {
        display_labels.get(status, status.title()): count for status, count in normalized_counts.items()
    }

    start_date = min(observed_dates).isoformat() if observed_dates else None
    end_date = max(observed_dates).isoformat() if observed_dates else None

    summary = {
        'total_reports': total_reports,
        'resolved_reports': resolved_reports,
        'active_reports': active_reports,
        'status_counts': status_counts,
        'window_days': window_days,
    }
    if start_date:
        summary['start_date'] = start_date
    if end_date:
        summary['end_date'] = end_date

    return jsonify(
        {
            'labels': labels,
            'values': values,
            'start_date': start_date,
            'end_date': end_date,
            'summary': summary,
        }
    )


@main_bp.route('/tracker_preview', methods=['GET'])
def tracker_preview():
    if 'username' not in session:
        return redirect(url_for('auth.login'))

    tracker = _get_tracker()
    local_zone = _tracker_local_zone()

    limit = 25
    session_rows = []
    event_rows = []

    with tracker._connect() as conn:
        session_query = (
            'SELECT session_token, start_time, end_time, duration_seconds '
            'FROM sessions ORDER BY datetime(start_time) DESC LIMIT ?'
        )
        session_rows = conn.execute(session_query, (limit,)).fetchall()

        tokens = [row['session_token'] for row in session_rows if row['session_token']]
        if tokens:
            placeholders = ','.join('?' for _ in tokens)
            event_query = (
                'SELECT session_token, event_name, context, metadata, occurred_at '
                f'FROM click_events WHERE session_token IN ({placeholders}) '
                'ORDER BY datetime(occurred_at) ASC, id ASC'
            )
            event_rows = conn.execute(event_query, tokens).fetchall()

    def _loads(payload):
        if not payload:
            return None
        try:
            return json.loads(payload)
        except Exception:
            return None

    events_by_session: dict[str, list[dict]] = {token: [] for token in {row['session_token'] for row in session_rows}}
    for row in event_rows:
        token = row['session_token']
        occurred_dt = _tracker_parse_timestamp(row['occurred_at'])
        context_payload = _loads(row['context'])
        metadata_payload = _loads(row['metadata'])
        label = None
        href = None
        if isinstance(context_payload, dict):
            label = (context_payload.get('text') or '').strip() or None
            href = context_payload.get('href')
        events_by_session.setdefault(token, []).append(
            {
                'name': row['event_name'],
                'occurred': occurred_dt,
                'occurred_display': _tracker_format_timestamp(occurred_dt, local_zone)
                or row['occurred_at'],
                'href': href,
                'label': label,
                'context': context_payload,
                'metadata': metadata_payload,
            }
        )

    total_sessions = len(session_rows)
    total_events = len(event_rows)
    total_navigation = 0
    total_backtracking = 0
    durations: list[float] = []
    start_times: list[datetime] = []
    end_times: list[datetime] = []

    for row in session_rows:
        token = row['session_token']
        events = events_by_session.get(token, [])
        start_ts = _tracker_parse_timestamp(row['start_time'])
        end_ts = _tracker_parse_timestamp(row['end_time'])
        derived_end_ts, _ = _derive_session_end(end_ts, events, local_zone=local_zone)

        navigation_events = [
            event
            for event in events
            if (event.get('name') or '').lower() == 'navigate'
        ]
        total_navigation += len(navigation_events)

        seen_hrefs: set[str] = set()
        session_backtracking = 0
        for event in navigation_events:
            href = event.get('href')
            if href:
                if href in seen_hrefs:
                    event['is_backtrack'] = True
                    session_backtracking += 1
                else:
                    seen_hrefs.add(href)
        total_backtracking += session_backtracking

        duration_seconds = _calculate_session_duration(
            start_ts,
            derived_end_ts,
            events,
            row['duration_seconds'],
        )

        if duration_seconds is not None:
            durations.append(duration_seconds)
        if start_ts:
            start_times.append(start_ts)
        if derived_end_ts:
            end_times.append(derived_end_ts)

    average_duration = (
        sum(durations) / len(durations)
        if durations
        else None
    )
    average_duration_label = _tracker_format_duration(average_duration)

    first_start = min(start_times) if start_times else None
    last_end = max(end_times) if end_times else None

    def _to_iso(value):
        if not value:
            return None
        try:
            return value.astimezone(timezone.utc).isoformat()
        except Exception:
            return value.isoformat() if hasattr(value, 'isoformat') else None

    start_iso = _to_iso(first_start)
    end_iso = _to_iso(last_end)
    start_display = _tracker_format_timestamp(first_start, local_zone)
    end_display = _tracker_format_timestamp(last_end, local_zone)

    if total_sessions:
        summary_text = f"Average Time: {average_duration_label}"
    else:
        summary_text = "No recent tracking sessions recorded."

    labels = [
        'Sessions',
        'Events',
        'Navigation',
        'Backtracking',
    ]
    values = [
        total_sessions,
        total_events,
        total_navigation,
        total_backtracking,
    ]

    payload = {
        'labels': labels,
        'values': values,
        'total_sessions': total_sessions,
        'total_events': total_events,
        'total_navigation_events': total_navigation,
        'total_backtracking_events': total_backtracking,
        'average_duration_seconds': average_duration,
        'average_duration_label': average_duration_label,
        'start_time': start_iso,
        'end_time': end_iso,
        'start_date': first_start.date().isoformat() if first_start else None,
        'end_date': last_end.date().isoformat() if last_end else None,
        'start_display': start_display,
        'end_display': end_display,
        'summary_text': summary_text,
    }

    return jsonify(payload)


@main_bp.route('/forecast_preview', methods=['GET'])
def forecast_preview():
    """Return forecast summary metrics for the dashboard preview."""
    if 'username' not in session:
        return redirect(url_for('auth.login'))

    moat_rows, moat_error = fetch_recent_moat()
    if moat_error:
        abort(500, description=moat_error)
    aoi_rows, aoi_error = fetch_aoi_reports()
    if aoi_error:
        abort(500, description=aoi_error)

    from datetime import datetime, timedelta

    today = datetime.utcnow().date()
    start = today - timedelta(days=6)

    recent_aoi: list[dict] = []
    assemblies: list[str] = []
    seen: set[str] = set()

    def _add_assembly(name: str | None):
        if not name:
            return
        if name not in seen:
            assemblies.append(name)
            seen.add(name)

    for row in moat_rows or []:
        asm, _ = _split_model_name(row.get('Model Name'))
        _add_assembly(asm)

    for row in aoi_rows or []:
        day = _parse_date(row.get('Date') or row.get('aoi_Date'))
        if not day or day < start or day > today:
            continue
        recent_aoi.append(row)
        _add_assembly(row.get('Assembly') or row.get('aoi_Assembly'))

    metrics = _aggregate_forecast(assemblies, moat_rows or [], recent_aoi)
    metrics = [m for m in metrics if (m.get('boards') or m.get('inspected'))]
    metrics.sort(key=lambda m: m.get('boards', 0.0), reverse=True)
    top_metrics = metrics[:5]

    labels = [m.get('assembly') or 'Unknown' for m in top_metrics]
    values = [m.get('predictedYield', 0.0) for m in top_metrics]

    date_values: list[date] = []
    for row in moat_rows or []:
        report_date = _parse_date(row.get('Report Date') or row.get('report_date'))
        if report_date:
            date_values.append(report_date)
    for row in recent_aoi:
        day = _parse_date(row.get('Date') or row.get('aoi_Date'))
        if day:
            date_values.append(day)

    start_date = min(date_values).isoformat() if date_values else None
    end_date = max(date_values).isoformat() if date_values else None

    return jsonify({
        'labels': labels,
        'values': values,
        'start_date': start_date,
        'end_date': end_date,
    })


@main_bp.route('/analysis/dpm', methods=['GET'])
@feature_required('analysis_dpm')
def dpm_analysis():
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    return render_template(
        'dpm_analysis.html',
        username=session.get('username'),
        user_role=(session.get('role') or '').upper(),
    )


@main_bp.route('/analysis/dpm/data', methods=['GET'])
@feature_required('analysis_dpm')
def dpm_data():
    if 'username' not in session:
        return redirect(url_for('auth.login'))

    chart_type = request.args.get('type', 'avg_false_calls_per_assembly')
    start = request.args.get('start_date')
    end = request.args.get('end_date')

    data, error = fetch_moat_dpm()
    if error:
        abort(500, description=error)
    if not data:
        return jsonify({"labels": [], "values": []})

    from collections import defaultdict
    from datetime import datetime

    def parse_date(d):
        if not d:
            return None
        try:
            return datetime.fromisoformat(str(d)).date()
        except Exception:
            return None

    grouped = defaultdict(lambda: {"falsecall": 0, "boards": 0})
    for row in data:
        date = row.get('Report Date') or row.get('report_date')
        dt = parse_date(date)
        if not dt:
            continue
        if start:
            sdt = parse_date(start)
            if sdt and dt < sdt:
                continue
        if end:
            edt = parse_date(end)
            if edt and dt > edt:
                continue
        fc = (
            row.get('FalseCall Windows')
            or row.get('falsecall_windows')
            or row.get('FalseCall Parts')
            or row.get('falsecall_parts')
            or 0
        )
        boards = row.get('Total Boards') or row.get('total_boards') or 0
        try:
            grouped[dt]["falsecall"] += float(fc)
        except (TypeError, ValueError):
            pass
        try:
            grouped[dt]["boards"] += float(boards)
        except (TypeError, ValueError):
            pass

    ordered_dates = sorted(list(grouped.keys()))
    labels = [d.isoformat() for d in ordered_dates]
    values = []
    for d in ordered_dates:
        g = grouped[d]
        values.append((g["falsecall"] / g["boards"]) if g["boards"] else 0)

    return jsonify({"labels": labels, "values": values, "type": chart_type})


@main_bp.route('/analysis/dpm/saved', methods=['GET', 'POST', 'PUT'])
@feature_required('analysis_dpm')
def dpm_saved_queries():
    if 'username' not in session:
        return redirect(url_for('auth.login'))

    if request.method == 'GET':
        data, error = fetch_dpm_saved_queries()
        if error:
            fallback = _load_local_dpm_saved_charts()
            if fallback:
                return jsonify(fallback)
            abort(500, description=error)
        if not data:
            fallback = _load_local_dpm_saved_charts()
            if fallback:
                return jsonify(fallback)
        return jsonify(data or [])

    payload = request.get_json() or {}
    keys = [
        "name",
        "type",
        "params",
        "description",
        "start_date",
        "end_date",
        "value_source",
        "x_column",
        "y_agg",
        "chart_type",
        "line_color",
    ]
    payload = {k: payload.get(k) for k in keys if k in payload}
    overwrite = request.method == 'PUT' or request.args.get('overwrite')
    if overwrite:
        name = payload.get('name')
        data, error = update_dpm_saved_query(name, payload)
        status = 200
    else:
        data, error = insert_dpm_saved_query(payload)
        status = 201
    if error:
        abort(500, description=error)
    return jsonify(data), status


@main_bp.route('/analysis/dpm/fallback_chart', methods=['GET'])
@feature_required('analysis_dpm')
def dpm_run_fallback_chart():
    if 'username' not in session:
        return redirect(url_for('auth.login'))

    chart_id = request.args.get('id')
    if not chart_id:
        abort(400, description="Chart id is required")

    definitions = _load_local_dpm_saved_charts()
    definition = next((item for item in definitions if item.get('id') == chart_id), None)
    if not definition:
        abort(404, description=f"Unknown fallback chart '{chart_id}'")

    data, error = fetch_moat_dpm()
    if error:
        abort(500, description=error)
    rows = data or []

    sql = definition.get('sql') or ''
    result_rows: list[dict] = []
    params: dict[str, str] = {}
    if sql and rows:
        placeholders = set(re.findall(r':([A-Za-z_][\w]*)', sql))
        min_date: datetime | None = None
        max_date: datetime | None = None
        for row in rows:
            for key in ('Report Date', 'report_date'):
                value = row.get(key)
                if not value:
                    continue
                try:
                    dt = datetime.fromisoformat(str(value))
                except ValueError:
                    continue
                if min_date is None or dt < min_date:
                    min_date = dt
                if max_date is None or dt > max_date:
                    max_date = dt
        default_from = (min_date.date().isoformat() if min_date else '1970-01-01')
        default_to = ((max_date + timedelta(days=1)).date().isoformat() if max_date else '2100-01-01')
        for name in placeholders:
            if name == 'from':
                params['from'] = (
                    request.args.get('from')
                    or request.args.get('start')
                    or default_from
                )
            elif name == 'to':
                params['to'] = (
                    request.args.get('to')
                    or request.args.get('end')
                    or default_to
                )
            else:
                value = request.args.get(name) or request.args.get(name.lower())
                if value in (None, ''):
                    abort(400, description=f"Parameter '{name}' is required for chart '{chart_id}'")
                params[name] = value
        conn: sqlite3.Connection | None = None
        try:
            conn = sqlite3.connect(':memory:')
            conn.row_factory = sqlite3.Row
            columns = sorted({key for row in rows for key in row.keys()})
            if columns:
                def _quote(col: str) -> str:
                    return '"' + str(col).replace('"', '""') + '"'

                type_map: dict[str, str] = {col: 'REAL' for col in columns}
                for row in rows:
                    for col in columns:
                        val = row.get(col)
                        if val is None:
                            continue
                        if isinstance(val, (int, float)) and not isinstance(val, bool):
                            continue
                        try:
                            float(val)
                        except (TypeError, ValueError):
                            type_map[col] = 'TEXT'
                quoted_cols = [_quote(col) for col in columns]
                create_sql = ", ".join(f"{name} {type_map[col]}" for name, col in zip(quoted_cols, columns))
                conn.execute(f"CREATE TABLE moat_dpm ({create_sql})")
                insert_sql = (
                    f"INSERT INTO moat_dpm ({', '.join(quoted_cols)}) VALUES ({', '.join(['?'] * len(columns))})"
                )
                for row in rows:
                    values = [row.get(col) for col in columns]
                    conn.execute(insert_sql, values)
                conn.commit()
                cursor = conn.execute(sql, params)
                result_rows = [dict(item) for item in cursor.fetchall()]
            else:
                result_rows = []
        except sqlite3.Error as exc:  # pragma: no cover - fallback execution errors
            abort(500, description=f'Failed to execute chart SQL: {exc}')
        finally:
            if conn is not None:
                try:
                    conn.close()
                except Exception:  # pragma: no cover - close failures
                    pass
    else:
        result_rows = []

    payload = {
        'id': chart_id,
        'name': definition.get('name'),
        'description': definition.get('description'),
        'chart_type': definition.get('chart_type'),
        'mappings': definition.get('mappings') or {},
        'rows': result_rows,
        'notes': definition.get('notes'),
    }
    return jsonify(payload)


@main_bp.route('/analysis/ppm', methods=['GET'])
@feature_required('analysis_ppm')
def ppm_analysis():
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    return render_template(
        'ppm_analysis.html',
        username=session.get('username'),
        user_role=(session.get('role') or '').upper(),
    )


@main_bp.route('/analysis/ppm/data', methods=['GET'])
@feature_required('analysis_ppm')
def ppm_data():
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    chart_type = request.args.get('type', 'avg_false_calls_per_assembly')
    start = request.args.get('start_date')
    end = request.args.get('end_date')

    # Currently only supports avg_false_calls_per_assembly; ordered by date
    data, error = fetch_moat()
    if error:
        abort(500, description=error)
    if not data:
        return jsonify({"labels": [], "values": []})

    from collections import defaultdict
    from datetime import datetime

    def parse_date(d):
        if not d:
            return None
        # Accept date or datetime strings
        try:
            return datetime.fromisoformat(str(d)).date()
        except Exception:
            return None

    grouped = defaultdict(lambda: {"falsecall": 0, "boards": 0})
    for row in data:
        date = row.get('Report Date') or row.get('report_date')
        dt = parse_date(date)
        if not dt:
            continue
        if start:
            sdt = parse_date(start)
            if sdt and dt < sdt:
                continue
        if end:
            edt = parse_date(end)
            if edt and dt > edt:
                continue
        fc = row.get('FalseCall Parts') or row.get('falsecall_parts') or 0
        boards = row.get('Total Boards') or row.get('total_boards') or 0
        grouped[dt]["falsecall"] += fc
        grouped[dt]["boards"] += boards

    ordered_dates = sorted(list(grouped.keys()))
    labels = [d.isoformat() for d in ordered_dates]
    values = []
    for d in ordered_dates:
        g = grouped[d]
        values.append((g["falsecall"] / g["boards"]) if g["boards"] else 0)

    return jsonify({"labels": labels, "values": values, "type": chart_type})


@main_bp.route('/analysis/ppm/saved', methods=['GET', 'POST', 'PUT'])
@feature_required('analysis_ppm')
def ppm_saved_queries():
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    if request.method == 'GET':
        data, error = fetch_saved_queries()
        if error:
            abort(500, description=error)
        return jsonify(data)

    payload = request.get_json() or {}
    keys = [
        "name",
        "type",
        "params",
        "description",
        "start_date",
        "end_date",
        "value_source",
        "x_column",
        "y_agg",
        "chart_type",
        "line_color",
    ]
    payload = {k: payload.get(k) for k in keys if k in payload}
    overwrite = request.method == 'PUT' or request.args.get('overwrite')
    if overwrite:
        name = payload.get('name')
        data, error = update_saved_query(name, payload)
        status = 200
    else:
        data, error = insert_saved_query(payload)
        status = 201
    if error:
        abort(500, description=error)
    return jsonify(data), status


@main_bp.route('/tools/assembly-forecast')
@feature_required('tools_assembly_forecast')
def assembly_forecast():
    """Render the Assembly Forecast tool page."""
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    return render_template('assembly_forecast.html', username=session.get('username'))


@main_bp.route('/api/assemblies/search')
@feature_required('tools_assembly_forecast')
def api_assemblies_search():
    """Search distinct assembly names across MOAT and AOI data."""
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    q = _norm(request.args.get('q') or '')
    assemblies: set[str] = set()
    moat_rows, moat_error = fetch_moat()
    if moat_error:
        abort(500, description=moat_error)
    for row in moat_rows or []:
        asm, _ = _split_model_name(row.get("Model Name"))
        if not asm:
            continue
        asm_norm = _norm(asm)
        if q and q not in asm_norm:
            continue
        assemblies.add(asm)
    aoi_rows, aoi_error = fetch_aoi_reports()
    if aoi_error:
        abort(500, description=aoi_error)
    for row in aoi_rows or []:
        asm = row.get("Assembly") or ""
        if not asm:
            continue
        asm_norm = _norm(asm)
        if q and q not in asm_norm:
            continue
        assemblies.add(asm)
    return jsonify(sorted(assemblies))


@main_bp.route('/api/assemblies/forecast', methods=['POST'])
@feature_required('tools_assembly_forecast')
def api_assemblies_forecast():
    """Return forecast metrics for selected assemblies."""
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    payload = request.get_json(silent=True) or {}
    assemblies = payload.get('assemblies') or []
    moat_rows, moat_error = fetch_moat()
    if moat_error:
        abort(500, description=moat_error)
    aoi_rows, aoi_error = fetch_aoi_reports()
    if aoi_error:
        abort(500, description=aoi_error)
    metrics = _aggregate_forecast(assemblies, moat_rows, aoi_rows)
    return jsonify({'assemblies': metrics})


def _fig_to_data_uri(fig):
    if plt is None:
        return ''
    buf = io.BytesIO()
    fig.savefig(buf, format='png', bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    b64 = base64.b64encode(buf.read()).decode('utf-8')
    return f"data:image/png;base64,{b64}"


def _build_metrics_chart(info: dict) -> str:
    """Return a bar chart image for key assembly metrics as a data URI."""
    if plt is None:
        return ""
    labels = [
        "Yield",
        "Hist Yield",
        "AOI Rejects",
        "Past Rejects",
        "FI Rejects",
    ]
    values = [
        info.get("yield") or 0,
        info.get("pastAvg") if isinstance(info.get("pastAvg"), (int, float)) else 0,
        info.get("currentRejects") or 0,
        info.get("pastRejectsAvg") or 0,
        info.get("fiTypicalRejects") or 0,
    ]
    fig, ax = plt.subplots(figsize=(3, 2))
    ax.bar(range(len(values)), values, color="steelblue")
    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=45, ha="right", fontsize=6)
    ax.set_ylabel("Value")
    fig.tight_layout()
    return _fig_to_data_uri(fig)


def _compute_control_limits(values: list[float]) -> tuple[float, float, float]:
    """Return mean, UCL and LCL for ``values`` using ±3σ limits."""
    if not values:
        return 0.0, 0.0, 0.0
    mean = sum(values) / len(values)
    variance = sum((v - mean) ** 2 for v in values) / len(values)
    stdev = math.sqrt(variance)
    ucl = mean + 3 * stdev
    lcl = max(0.0, mean - 3 * stdev)
    return mean, ucl, lcl


def _build_assembly_moat_charts(assembly: str, moat_rows: list[dict]) -> dict[str, str]:
    """Build an overlay control chart for SMT and TH false-call data.

    Returns a dictionary with a single key ``overlayChart`` containing a
    data URI for the generated chart (or an empty string if unavailable).
    """
    if plt is None:
        return {"overlayChart": ""}

    records: list[dict[str, str | float]] = []
    asm_lower = assembly.lower()
    for row in moat_rows or []:
        model = (
            row.get("Model Name")
            or row.get("model_name")
            or row.get("Model")
            or ""
        )
        model_lower = str(model).lower()
        if asm_lower not in model_lower:
            continue
        try:
            fc = float(row.get("FalseCall Parts") or row.get("falsecall_parts") or 0)
            boards = float(row.get("Total Boards") or row.get("total_boards") or 0)
        except (TypeError, ValueError):
            continue
        if boards == 0:
            continue
        group = "th" if "th" in model_lower else "smt" if "smt" in model_lower else None
        if group is None:
            continue
        records.append(
            {
                "group": group,
                "date": row.get("Report Date") or row.get("report_date") or "",
                "val": fc / boards,
            }
        )

    # average values per day for each group
    grouped: dict[tuple[str, str], dict[str, float]] = defaultdict(lambda: {"sum": 0.0, "count": 0})
    for rec in records:
        key = (rec["group"], rec["date"])
        grouped[key]["sum"] += rec["val"]
        grouped[key]["count"] += 1
    averaged_records = [
        {
            "group": g,
            "date": d,
            "val": info["sum"] / info["count"] if info["count"] else 0.0,
        }
        for (g, d), info in grouped.items()
    ]

    smt_data = [r for r in averaged_records if r["group"] == "smt"]
    th_data = [r for r in averaged_records if r["group"] == "th"]
    if not (smt_data or th_data):
        return {"overlayChart": ""}

    fig, ax = plt.subplots(figsize=(6, 3))
    groups = {
        "smt": {
            "data": smt_data,
            "color": "tab:blue",
            "label": "SMT",
        },
        "th": {
            "data": th_data,
            "color": "tab:orange",
            "label": "TH",
        },
    }
    for key, info in groups.items():
        data = info["data"]
        if not data:
            continue
        data.sort(key=lambda d: d.get("date", ""))
        dates = [d["date"] for d in data]
        vals = [d["val"] for d in data]
        mean, ucl, lcl = _compute_control_limits(vals)
        color = info["color"]
        label = info["label"]
        ax.plot(dates, vals, marker="o", color=color, label=f"{label} False Calls/Board")
        ax.axhline(mean, linestyle="--", color=color, label=f"{label} Mean")
        ax.axhline(ucl, linestyle="--", color=color, label=f"{label} +3σ")
        ax.axhline(lcl, linestyle="--", color=color, label=f"{label} -3σ")

    ax.set_ylabel("False Calls/Board")
    ax.set_title("SMT vs TH False Calls Control Chart")
    ax.tick_params(axis="x", rotation=45)
    ax.legend(loc="center left", bbox_to_anchor=(1, 0.5))

    return {"overlayChart": _fig_to_data_uri(fig)}


def _generate_report_charts(payload):
    if plt is None:
        return {
            'yieldTrendImg': '',
            'operatorRejectImg': '',
            'modelFalseCallsImg': '',
            'fcVsNgRateImg': '',
            'fcNgRatioImg': '',
        }
    charts: dict[str, str] = {}

    # Yield trend chart
    fig, ax = plt.subplots(figsize=(8, 4))
    dates = payload.get('yieldData', {}).get('dates', [])
    yields = payload.get('yieldData', {}).get('yields', [])
    if dates and yields:
        ax.plot(dates, yields, marker='o')
        ax.set_xlabel('Date')
        ax.set_ylabel('Yield %')
        ax.set_title('Yield Trend')
        ax.tick_params(axis='x', rotation=45)
    charts['yieldTrendImg'] = _fig_to_data_uri(fig)

    # Operator reject chart (stacked bar)
    fig, ax = plt.subplots(figsize=(8, 4))
    ops = payload.get('operators', [])
    if ops:
        names = [o['name'] for o in ops]
        accepted = [o['inspected'] - o['rejected'] for o in ops]
        rejected = [o['rejected'] for o in ops]
        ax.bar(names, accepted, label='Accepted')
        ax.bar(names, rejected, bottom=accepted, label='Rejected')
        ax.set_ylabel('Boards')
        ax.set_title('Operator Rejects')
        ax.tick_params(axis='x', rotation=45)
        ax.legend()
    charts['operatorRejectImg'] = _fig_to_data_uri(fig)

    # Model false calls chart with control limits
    fig, ax = plt.subplots(figsize=(8, 4))
    models = payload.get('models', [])
    if models:
        labels = [m['name'] for m in models]
        vals = [m['falseCalls'] for m in models]
        ax.plot(labels, vals, marker='o', color='orange', label='False Calls')
        mean = sum(vals) / len(vals)
        std = math.sqrt(sum((v - mean) ** 2 for v in vals) / len(vals))
        upper = mean + 3 * std
        lower = max(mean - 3 * std, 0)
        ax.plot(labels, [mean] * len(labels), linestyle='--', color='blue', label='Mean')
        ax.plot(labels, [upper] * len(labels), linestyle='--', color='green', label='+3σ')
        ax.plot(labels, [lower] * len(labels), linestyle='--', color='red', label='-3σ')
        ax.set_ylabel('False Calls/Board')
        ax.set_title('False Calls by Model')
        ax.tick_params(axis='x', labelbottom=False)
        ax.legend()
    charts['modelFalseCallsImg'] = _fig_to_data_uri(fig)

    # FC vs NG rate chart
    fig, ax = plt.subplots(figsize=(8, 4))
    fc_vs_ng = payload.get('fcVsNgRate', {})
    dates = fc_vs_ng.get('dates', [])
    ng_ppm = fc_vs_ng.get('ngPpm', [])
    fc_ppm = fc_vs_ng.get('fcPpm', [])
    if dates:
        ax.plot(dates, ng_ppm, color='red', label='NG PPM')
        ax.plot(dates, fc_ppm, color='blue', label='FalseCall PPM')
        ax.set_ylabel('PPM')
        ax.set_title('FC vs NG Rate')
        ax.tick_params(axis='x', rotation=45)
        ax.legend()
    charts['fcVsNgRateImg'] = _fig_to_data_uri(fig)

    # FC/NG ratio chart
    fig, ax = plt.subplots(figsize=(8, 4))
    fc_ng = payload.get('fcNgRatio', {})
    models = fc_ng.get('models', [])
    ratios = fc_ng.get('ratios', [])
    if models:
        ax.bar(models, ratios, color='teal')
        ax.set_ylabel('FC/NG Ratio')
        ax.set_title('FC/NG Ratio by Model')
        ax.tick_params(axis='x', labelbottom=False)
    charts['fcNgRatioImg'] = _fig_to_data_uri(fig)

    return charts


def _normalize_line_name(raw: str | None) -> str:
    text = (raw or '').strip()
    if not text:
        return 'UNKNOWN'
    return text.upper()


def _normalize_assembly_name(row: dict) -> str:
    for key in (
        'Assembly',
        'assembly',
        'Model Name',
        'model_name',
        'Model',
        'model',
    ):
        value = row.get(key)
        if value not in (None, ''):
            return str(value)
    return 'Unknown'


def _coerce_date_key(row: dict, *keys: str) -> date | None:
    for key in keys:
        dt = _parse_date(row.get(key))
        if dt:
            return dt
    return None


def _line_bucket() -> dict[str, object]:
    return {
        'total_parts': 0.0,
        'total_boards': 0.0,
        'false_call_parts': 0.0,
        'ng_parts': 0.0,
        'total_windows': 0.0,
        'ng_windows': 0.0,
        'false_call_windows': 0.0,
        'ppm_values': [],
        'fc_dpm_values': [],
        'dpm_values': [],
        'dates': set(),
        'defects': defaultdict(float),
    }


def _daily_bucket() -> dict[str, float]:
    return {
        'parts': 0.0,
        'boards': 0.0,
        'false_calls': 0.0,
        'ng_parts': 0.0,
        'windows': 0.0,
        'ng_windows': 0.0,
        'fc_windows': 0.0,
    }


def _assembly_bucket() -> dict[str, object]:
    return {
        'total_parts': 0.0,
        'false_calls': 0.0,
        'ng_parts': 0.0,
        'boards': 0.0,
        'windows': 0.0,
        'ng_windows': 0.0,
        'fc_windows': 0.0,
        'defects': defaultdict(float),
        'dates': set(),
    }


def _assembly_daily_bucket() -> dict[str, float]:
    return {
        'parts': 0.0,
        'false_calls': 0.0,
        'ng_parts': 0.0,
        'boards': 0.0,
        'windows': 0.0,
        'ng_windows': 0.0,
        'fc_windows': 0.0,
    }


def _safe_ratio(num: float, den: float) -> float:
    if den == 0:
        return 0.0
    return num / den


def build_line_report_payload(start: date | None = None, end: date | None = None) -> dict:
    line_report_sql = """
        select
            report_date,
            line,
            model_name,
            ppm_total_boards,
            ppm_total_parts,
            ppm_falsecall_parts,
            ppm_ng_parts,
            dpm_total_boards,
            dpm_total_windows,
            dpm_ng_windows,
            dpm_falsecall_windows
        from aoi_base_daily
        where (%(start_date)s is null or report_date >= %(start_date)s)
          and (%(end_date)s is null or report_date <= %(end_date)s)
    """

    params = {
        "start_date": start.isoformat() if start else None,
        "end_date": end.isoformat() if end else None,
    }

    if query_aoi_base_daily is not _ORIGINAL_AOI_QUERY:
        grouped_rows, error = query_aoi_base_daily(line_report_sql, params)
        if error:
            abort(500, description=error)
    else:
        moat_rows, moat_error = fetch_moat(start_date=start, end_date=end)
        if moat_error:
            abort(500, description=moat_error)

        dpm_rows, dpm_error = fetch_moat_dpm(start_date=start, end_date=end)
        if dpm_error:
            abort(500, description=dpm_error)

        combined: dict[tuple[date, str, str], dict] = {}

        def _coalesce_model(value: str | None) -> str:
            if value in (None, ''):
                return 'Unknown'
            return str(value)

        def _add_ppm_row(row: dict) -> None:
            dt = _parse_date(row.get('Report Date') or row.get('report_date'))
            if not dt:
                return
            line = row.get('Line') or row.get('line')
            if not line:
                return
            model = _coalesce_model(row.get('Model Name') or row.get('model_name'))
            key = (dt, str(line), model)
            entry = combined.setdefault(
                key,
                {
                    'report_date': dt.isoformat(),
                    'Report Date': dt.isoformat(),
                    'line': str(line),
                    'Line': str(line),
                    'model_name': model,
                    'Model Name': model,
                },
            )
            entry['Total Parts'] = entry['ppm_total_parts'] = _coerce_number(row.get('Total Parts'))
            entry['Total Boards'] = entry['ppm_total_boards'] = _coerce_number(row.get('Total Boards'))
            entry['FalseCall Parts'] = entry['ppm_falsecall_parts'] = _coerce_number(row.get('FalseCall Parts'))
            entry['NG Parts'] = entry['ppm_ng_parts'] = _coerce_number(row.get('NG Parts'))
            if 'FalseCall PPM' in row:
                entry['FalseCall PPM'] = _coerce_number(row.get('FalseCall PPM'))

        def _add_dpm_row(row: dict) -> None:
            dt = _parse_date(row.get('Report Date') or row.get('report_date'))
            if not dt:
                return
            line = row.get('Line') or row.get('line')
            if not line:
                return
            model = _coalesce_model(row.get('Model Name') or row.get('model_name'))
            key = (dt, str(line), model)
            entry = combined.setdefault(
                key,
                {
                    'report_date': dt.isoformat(),
                    'Report Date': dt.isoformat(),
                    'line': str(line),
                    'Line': str(line),
                    'model_name': model,
                    'Model Name': model,
                },
            )
            total_boards = _coerce_number(row.get('Total Boards'), default=None)
            if total_boards is not None:
                entry['dpm_total_boards'] = total_boards
            total_windows = _coerce_number(row.get('Total Windows'), default=None)
            if total_windows is not None:
                entry['dpm_total_windows'] = entry['Total Windows'] = total_windows
            ng_windows = _coerce_number(row.get('NG Windows'), default=None)
            if ng_windows is not None:
                entry['dpm_ng_windows'] = entry['NG Windows'] = ng_windows
            fc_windows = _coerce_number(row.get('FalseCall Windows'), default=None)
            if fc_windows is not None:
                entry['dpm_falsecall_windows'] = entry['FalseCall Windows'] = fc_windows
            windows_per_board = _coerce_number(row.get('Windows per board'), default=None)
            if windows_per_board is not None:
                entry['Windows per board'] = entry['windows_per_board'] = windows_per_board
            dpm_value = _coerce_number(row.get('DPM'), default=None)
            if dpm_value is not None:
                entry['dpm_dpm'] = dpm_value
            fc_dpm_value = _coerce_number(row.get('FC DPM'), default=None)
            if fc_dpm_value is not None:
                entry['dpm_falsecall_dpm'] = fc_dpm_value

        for item in moat_rows or []:
            if isinstance(item, dict):
                _add_ppm_row(item)

        for item in dpm_rows or []:
            if isinstance(item, dict):
                _add_dpm_row(item)

        grouped_rows: dict[str, dict[str, list[dict]]] = {}
        for (dt, line, _model), data in combined.items():
            date_key = dt.isoformat()
            line_map = grouped_rows.setdefault(date_key, {})
            line_map.setdefault(line, []).append(data)

    grouped_rows = grouped_rows or {}

    line_totals: dict[str, dict[str, object]] = defaultdict(_line_bucket)
    line_daily: dict[str, dict[date, dict[str, float]]] = defaultdict(
        lambda: defaultdict(_daily_bucket)
    )
    assembly_line: dict[str, dict[str, dict[str, object]]] = defaultdict(
        lambda: defaultdict(_assembly_bucket)
    )
    assembly_daily: dict[str, dict[str, dict[date, dict[str, float]]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(_assembly_daily_bucket))
    )

    overall = {
        'total_parts': 0.0,
        'false_calls': 0.0,
        'ng_parts': 0.0,
        'total_boards': 0.0,
        'total_windows': 0.0,
        'ng_windows': 0.0,
        'false_call_windows': 0.0,
    }

    def _within_range(dt: date | None) -> bool:
        if not dt:
            return False
        if start and dt < start:
            return False
        if end and dt > end:
            return False
        return True

    def _extract_number(
        row: dict, *keys: str, default: float | None = 0.0
    ) -> float | None:
        for key in keys:
            if key in row:
                return _coerce_number(row.get(key), default=default)
        return default

    for date_key, line_map in grouped_rows.items():
        dt = _parse_date(date_key)
        if start or end:
            if not _within_range(dt):
                continue

        for raw_line, records in line_map.items():
            if not records:
                continue

            line = _normalize_line_name(raw_line)
            bucket = line_totals[line]
            day_bucket = line_daily[line][dt] if dt else None

            for row in records:
                assembly = _normalize_assembly_name(row)
                parts = _extract_number(
                    row,
                    'Total Parts',
                    'total_parts',
                    'parts',
                    'totalParts',
                    'ppm_total_parts',
                )
                boards = _extract_number(
                    row,
                    'Total Boards',
                    'total_boards',
                    'boards',
                    'totalBoards',
                    'ppm_total_boards',
                    'dpm_total_boards',
                )
                fc_parts = _extract_number(
                    row,
                    'FalseCall Parts',
                    'falsecall_parts',
                    'false_call_parts',
                    'false_calls',
                    'falseCalls',
                    'ppm_false_calls',
                    'ppm_falsecall_parts',
                )
                ng_parts = _extract_number(
                    row,
                    'NG Parts',
                    'ng_parts',
                    'ngParts',
                    'ng_parts_total',
                    'ppm_ng_parts',
                )

                windows = _extract_number(
                    row,
                    'Total Windows',
                    'total_windows',
                    'windows',
                    'totalWindows',
                    'dpm_total_windows',
                    default=None,
                )

                if windows in (None, 0.0):
                    windows_per_board = _extract_number(
                        row,
                        'Windows per board',
                        'windows_per_board',
                        'windowsPerBoard',
                        default=None,
                    )
                    boards_value = (
                        boards
                        if boards
                        else _extract_number(
                            row,
                            'Total Boards',
                            'total_boards',
                            'boards',
                            'totalBoards',
                            'ppm_total_boards',
                            'dpm_total_boards',
                            default=None,
                        )
                    )
                    if (
                        windows_per_board is not None
                        and boards_value is not None
                    ):
                        windows = windows_per_board * boards_value

                if windows is None:
                    windows = 0.0

                ng_windows = _extract_number(
                    row,
                    'NG Windows',
                    'ng_windows',
                    'ngWindows',
                    'dpm_ng_windows',
                )
                fc_windows = _extract_number(
                    row,
                    'FalseCall Windows',
                    'falsecall_windows',
                    'false_call_windows',
                    'fc_windows',
                    'falseCallWindows',
                    'dpm_false_calls',
                    'dpm_falsecall_windows',
                )
                dpm_value = _extract_number(
                    row,
                    'DPM',
                    'dpm',
                    'defect_dpm',
                    'dpm_dpm',
                    default=None,
                )
                fc_dpm_value = _extract_number(
                    row,
                    'FC DPM',
                    'fc_dpm',
                    'dpm_falsecall_dpm',
                    default=None,
                )
                ppm_value = _extract_number(
                    row,
                    'FalseCall PPM',
                    'falsecall_ppm',
                    'false_call_ppm',
                    'ppm_falsecall_ppm',
                    default=None,
                )
                parts_value = parts or 0.0
                boards_value = boards or 0.0
                fc_parts_value = fc_parts or 0.0
                ng_parts_value = ng_parts or 0.0
                windows_value = windows or 0.0
                ng_windows_value = ng_windows or 0.0
                fc_windows_value = fc_windows or 0.0

                if ppm_value is None and parts_value:
                    ppm_value = _safe_ratio(fc_parts_value, parts_value) * 1_000_000

                if fc_dpm_value is None and windows_value:
                    fc_dpm_value = _safe_ratio(fc_windows_value, windows_value) * 1_000_000

                if dpm_value is None:
                    if windows_value:
                        dpm_value = (
                            _safe_ratio(ng_windows_value, windows_value) * 1_000_000
                        )
                    elif parts_value:
                        confirmed_value = max(0.0, ng_parts_value - fc_parts_value)
                        dpm_value = (
                            _safe_ratio(confirmed_value, parts_value) * 1_000_000
                        )

                defect_name = (
                    row.get('Defect Name')
                    or row.get('Defect')
                    or row.get('defect_name')
                )

                bucket['total_parts'] += parts_value
                bucket['total_boards'] += boards_value
                bucket['false_call_parts'] += fc_parts_value
                bucket['ng_parts'] += ng_parts_value
                bucket['total_windows'] += windows_value
                bucket['ng_windows'] += ng_windows_value
                bucket['false_call_windows'] += fc_windows_value

                if ppm_value:
                    bucket['ppm_values'].append(ppm_value)
                if dpm_value:
                    bucket['dpm_values'].append(dpm_value)
                if fc_dpm_value:
                    bucket['fc_dpm_values'].append(fc_dpm_value)

                if dt:
                    bucket['dates'].add(dt)
                    day_bucket = line_daily[line][dt]
                    day_bucket['parts'] += parts_value
                    day_bucket['boards'] += boards_value
                    day_bucket['false_calls'] += fc_parts_value
                    day_bucket['ng_parts'] += ng_parts_value
                    day_bucket['windows'] += windows_value
                    day_bucket['ng_windows'] += ng_windows_value
                    day_bucket['fc_windows'] += fc_windows_value

                asm_bucket = assembly_line[assembly][line]
                asm_bucket['total_parts'] += parts_value
                asm_bucket['false_calls'] += fc_parts_value
                asm_bucket['ng_parts'] += ng_parts_value
                asm_bucket['boards'] += boards_value
                asm_bucket['windows'] += windows_value
                asm_bucket['ng_windows'] += ng_windows_value
                asm_bucket['fc_windows'] += fc_windows_value

                if defect_name:
                    defect = str(defect_name).strip() or 'Unknown'
                    asm_bucket['defects'][defect] += ng_windows_value
                    bucket['defects'][defect] += ng_windows_value
                elif ng_windows_value:
                    asm_bucket['defects']['Unknown'] += ng_windows_value
                    bucket['defects']['Unknown'] += ng_windows_value

                if dt:
                    asm_bucket['dates'].add(dt)
                    asm_day = assembly_daily[assembly][line][dt]
                    asm_day['parts'] += parts_value
                    asm_day['false_calls'] += fc_parts_value
                    asm_day['ng_parts'] += ng_parts_value
                    asm_day['boards'] += boards_value
                    asm_day['windows'] += windows_value
                    asm_day['ng_windows'] += ng_windows_value
                    asm_day['fc_windows'] += fc_windows_value

                overall['total_parts'] += parts_value
                overall['false_calls'] += fc_parts_value
                overall['ng_parts'] += ng_parts_value
                overall['total_boards'] += boards_value
                overall['total_windows'] += windows_value
                overall['ng_windows'] += ng_windows_value
                overall['false_call_windows'] += fc_windows_value

    def _line_metrics(line: str, info: dict[str, object]) -> dict[str, object]:
        parts = info['total_parts']
        boards = info['total_boards']
        fc_parts = info['false_call_parts']
        ng_parts = info['ng_parts']
        windows = info['total_windows']
        ng_windows = info['ng_windows']
        fc_windows = info['false_call_windows']
        confirmed_parts = max(0.0, ng_parts - fc_parts)
        window_confirmed = ng_windows if windows else None

        window_yield = (
            100.0 * (windows - ng_windows) / windows
            if windows
            else None
        )
        raw_part_yield = (
            100.0 * (parts - ng_parts) / parts
            if parts
            else None
        )
        true_part_yield = (
            100.0 * (parts - confirmed_parts) / parts
            if parts
            else None
        )

        fc_per_board = _safe_ratio(fc_parts, boards)
        windows_per_board = _safe_ratio(windows, boards) if boards else None
        defects_per_board = _safe_ratio(ng_windows, boards) if boards else None
        false_call_ppm = (
            _safe_ratio(fc_parts, parts) * 1_000_000
            if parts
            else None
        )
        false_call_dpm = (
            _safe_ratio(fc_windows, windows) * 1_000_000
            if windows
            else None
        )
        if false_call_dpm is None and info['fc_dpm_values']:
            false_call_dpm = sum(info['fc_dpm_values']) / len(info['fc_dpm_values'])

        defect_dpm = (
            _safe_ratio(ng_windows, windows) * 1_000_000
            if windows
            else None
        )
        if defect_dpm is None and parts:
            defect_dpm = _safe_ratio(confirmed_parts, parts) * 1_000_000
        date_count = len(info['dates']) or 1
        boards_per_day = boards / date_count if boards else 0.0
        return {
            'line': line,
            'totalParts': parts,
            'totalBoards': boards,
            'totalWindows': windows,
            'ngParts': ng_parts,
            'ngWindows': ng_windows,
            'falseCalls': fc_parts,
            'confirmedDefects': confirmed_parts,
            'windowConfirmedDefects': window_confirmed,
            'windowYield': window_yield,
            'rawPartYield': raw_part_yield,
            'truePartYield': true_part_yield,
            'falseCallsPerBoard': fc_per_board,
            'windowsPerBoard': windows_per_board,
            'defectsPerBoard': defects_per_board,
            'falseCallPpm': false_call_ppm,
            'falseCallDpm': false_call_dpm,
            'defectDpm': defect_dpm,
            'boardsPerDay': boards_per_day,
            'datesActive': date_count,
        }

    def _line_period_summary(line: str, info: dict[str, object]) -> dict[str, object]:
        parts = info['total_parts']
        boards = info['total_boards']
        false_calls = info['false_call_parts']
        ng_parts = info['ng_parts']
        windows = info['total_windows']
        ng_windows = info['ng_windows']
        confirmed_parts = max(0.0, ng_parts - false_calls)

        raw_part_yield = (
            100.0 * (parts - ng_parts) / parts if parts else None
        )
        true_part_yield = (
            100.0 * (parts - confirmed_parts) / parts if parts else None
        )
        window_yield = (
            100.0 * (windows - ng_windows) / windows if windows else None
        )

        return {
            'line': line,
            'true_part_yield_pct': true_part_yield,
            'window_yield_pct': window_yield,
            'raw_part_yield_pct': raw_part_yield,
            'fc_per_board': _safe_ratio(false_calls, boards) if boards else 0.0,
            'defects_per_board': _safe_ratio(ng_windows, boards) if boards else 0.0,
            'total_boards': boards,
            'total_parts': parts,
            'total_windows': windows,
            'false_calls': false_calls,
            'ng_windows': ng_windows,
        }

    line_metrics = [
        _line_metrics(line, info)
        for line, info in sorted(line_totals.items())
        if info['total_parts'] or info['total_windows']
    ]

    line_period_summaries = [
        _line_period_summary(line, info)
        for line, info in sorted(line_totals.items())
        if info['total_parts'] or info['total_windows'] or info['total_boards']
    ]

    company_confirmed = max(0.0, overall['ng_parts'] - overall['false_calls'])
    company_window_yield = (
        100.0 * (overall['total_windows'] - overall['ng_windows']) / overall['total_windows']
        if overall['total_windows']
        else None
    )
    company_raw_part_yield = (
        100.0
        * (overall['total_parts'] - overall['ng_parts'])
        / overall['total_parts']
        if overall['total_parts']
        else None
    )
    company_true_part_yield = (
        100.0
        * (overall['total_parts'] - company_confirmed)
        / overall['total_parts']
        if overall['total_parts']
        else None
    )
    company_fc_rate = _safe_ratio(overall['false_calls'], overall['total_boards'])
    company_false_call_ppm = (
        _safe_ratio(overall['false_calls'], overall['total_parts']) * 1_000_000
        if overall['total_parts']
        else None
    )
    company_false_call_dpm = (
        _safe_ratio(overall['false_call_windows'], overall['total_windows']) * 1_000_000
        if overall['total_windows']
        else None
    )
    company_dpm = (
        _safe_ratio(overall['ng_windows'], overall['total_windows']) * 1_000_000
        if overall['total_windows']
        else (
            _safe_ratio(company_confirmed, overall['total_parts']) * 1_000_000
            if overall['total_parts']
            else None
        )
    )

    def _resolve_yield(window_value, true_value, raw_value) -> float:
        for value in (window_value, true_value, raw_value):
            if value is not None:
                return value
        return 0.0

    company_yield_value = _resolve_yield(
        company_window_yield, company_true_part_yield, company_raw_part_yield
    )

    line_vs_company = []
    for metrics in line_metrics:
        line_vs_company.append(
            {
                'line': metrics['line'],
                'windowYieldDelta': _resolve_yield(
                    metrics.get('windowYield'),
                    metrics.get('truePartYield'),
                    metrics.get('rawPartYield'),
                )
                - company_yield_value,
                'falseCallDelta': metrics['falseCallsPerBoard'] - company_fc_rate,
                'falseCallPpmDelta': (
                    (metrics['falseCallPpm'] or 0.0)
                    - (company_false_call_ppm or 0.0)
                ),
                'falseCallDpmDelta': (
                    (metrics['falseCallDpm'] or 0.0)
                    - (company_false_call_dpm or 0.0)
                ),
                'defectDpmDelta': (
                    (metrics['defectDpm'] or 0.0)
                    - (company_dpm or 0.0)
                ),
            }
        )

    def _summarize_line_trend(line: str, daily: dict[date, dict[str, float]]):
        entries = []
        for dt, values in sorted(daily.items()):
            parts = values['parts']
            boards = values['boards']
            fc_parts = values['false_calls']
            ng_parts = values['ng_parts']
            windows = values['windows']
            ng_windows = values['ng_windows']
            fc_windows = values['fc_windows']
            confirmed_parts = max(0.0, ng_parts - fc_parts)
            window_confirmed = ng_windows if windows else None
            window_yield = (
                100.0 * (windows - ng_windows) / windows
                if windows
                else None
            )
            raw_part_yield = (
                100.0 * (parts - ng_parts) / parts
                if parts
                else None
            )
            true_part_yield = (
                100.0 * (parts - confirmed_parts) / parts
                if parts
                else None
            )
            fc_rate = _safe_ratio(fc_parts, boards) if boards else None
            false_call_ppm = (
                _safe_ratio(fc_parts, parts) * 1_000_000
                if parts
                else None
            )
            false_call_dpm = (
                _safe_ratio(fc_windows, windows) * 1_000_000
                if windows
                else None
            )
            defect_dpm = (
                _safe_ratio(ng_windows, windows) * 1_000_000
                if windows
                else (
                    _safe_ratio(confirmed_parts, parts) * 1_000_000
                    if parts
                    else None
                )
            )
            windows_per_board = _safe_ratio(windows, boards) if boards else None
            defects_per_board = _safe_ratio(ng_windows, boards) if boards else None
            entries.append(
                {
                    'date': dt.isoformat(),
                    'windowYield': window_yield,
                    'rawPartYield': raw_part_yield,
                    'truePartYield': true_part_yield,
                    'falseCallsPerBoard': fc_rate,
                    'falseCallPpm': false_call_ppm,
                    'falseCallDpm': false_call_dpm,
                    'defectDpm': defect_dpm,
                    'confirmedDefects': confirmed_parts,
                    'windowConfirmedDefects': window_confirmed,
                    'boards': boards,
                    'parts': parts,
                    'ngParts': ng_parts,
                    'falseCalls': fc_parts,
                    'windows': windows,
                    'ngWindows': ng_windows,
                    'windowsPerBoard': windows_per_board,
                    'defectsPerBoard': defects_per_board,
                }
            )
        return {'line': line, 'entries': entries}

    line_trends = [
        _summarize_line_trend(line, daily)
        for line, daily in line_daily.items()
        if daily
    ]

    def _stddev(values: list[float]) -> float:
        if len(values) < 2:
            return 0.0
        avg = sum(values) / len(values)
        return math.sqrt(sum((v - avg) ** 2 for v in values) / len(values))

    benchmarking = {
        'bestYield': max(
            line_metrics,
            key=lambda m: _resolve_yield(
                m.get('windowYield'),
                m.get('truePartYield'),
                m.get('rawPartYield'),
            ),
            default=None,
        ),
        'lowestFalseCalls': min(
            line_metrics, key=lambda m: m['falseCallsPerBoard'], default=None
        ),
        'mostConsistent': None,
        'lineVsCompany': line_vs_company,
    }

    consistency_scores: list[tuple[str, float]] = []
    for trend in line_trends:
        yields = [
            _resolve_yield(
                entry.get('windowYield'),
                entry.get('truePartYield'),
                entry.get('rawPartYield'),
            )
            for entry in trend['entries']
            if any(
                value is not None
                for value in (
                    entry.get('windowYield'),
                    entry.get('truePartYield'),
                    entry.get('rawPartYield'),
                )
            )
        ]
        if not yields:
            continue
        consistency_scores.append((trend['line'], _stddev(yields)))
    if consistency_scores:
        line, score = min(consistency_scores, key=lambda item: item[1])
        match = next((m for m in line_metrics if m['line'] == line), None)
        if match:
            benchmarking['mostConsistent'] = {**match, 'stddev': score}

    assembly_comparisons = []
    for assembly, line_map in sorted(assembly_line.items()):
        lines_info: dict[str, dict[str, object]] = {}
        for line, info in line_map.items():
            parts = info['total_parts']
            boards = info['boards']
            fc_parts = info['false_calls']
            ng_parts = info['ng_parts']
            windows = info['windows']
            ng_windows = info['ng_windows']
            fc_windows = info['fc_windows']
            confirmed_parts = max(0.0, ng_parts - fc_parts)
            window_confirmed = ng_windows if windows else None
            window_yield = (
                100.0 * (windows - ng_windows) / windows
                if windows
                else None
            )
            raw_part_yield = (
                100.0 * (parts - ng_parts) / parts
                if parts
                else None
            )
            true_part_yield = (
                100.0 * (parts - confirmed_parts) / parts
                if parts
                else None
            )
            fc_rate = _safe_ratio(fc_parts, boards) if boards else None
            false_call_ppm = (
                _safe_ratio(fc_parts, parts) * 1_000_000 if parts else None
            )
            false_call_dpm = (
                _safe_ratio(fc_windows, windows) * 1_000_000 if windows else None
            )
            defect_dpm = (
                _safe_ratio(ng_windows, windows) * 1_000_000
                if windows
                else (
                    _safe_ratio(confirmed_parts, parts) * 1_000_000
                    if parts
                    else None
                )
            )
            defects = info['defects']
            defect_total = sum(defects.values())
            defect_mix = (
                {name: _safe_ratio(value, defect_total) for name, value in defects.items()}
                if defect_total
                else {}
            )
            lines_info[line] = {
                'windowYield': window_yield,
                'rawPartYield': raw_part_yield,
                'truePartYield': true_part_yield,
                'falseCallsPerBoard': fc_rate,
                'falseCallPpm': false_call_ppm,
                'falseCallDpm': false_call_dpm,
                'defectDpm': defect_dpm,
                'defectMix': defect_mix,
                'confirmedDefects': confirmed_parts,
                'windowConfirmedDefects': window_confirmed,
                'parts': parts,
                'boards': boards,
                'windows': windows,
                'ngParts': ng_parts,
                'ngWindows': ng_windows,
                'falseCalls': fc_parts,
                'windowsPerBoard': _safe_ratio(windows, boards) if boards else None,
                'defectsPerBoard': _safe_ratio(ng_windows, boards) if boards else None,
            }
        if len(lines_info) >= 2:
            assembly_comparisons.append({'assembly': assembly, 'lines': lines_info})

    yield_variance = []
    false_call_variance = []
    for comp in assembly_comparisons:
        yields = [
            _resolve_yield(
                v.get('windowYield'), v.get('truePartYield'), v.get('rawPartYield')
            )
            for v in comp['lines'].values()
            if any(
                value is not None
                for value in (
                    v.get('windowYield'), v.get('truePartYield'), v.get('rawPartYield')
                )
            )
        ]
        if len(yields) > 1:
            yield_variance.append(
                {
                    'assembly': comp['assembly'],
                    'stddev': _stddev(yields),
                }
            )
        fc_rates = [
            v['falseCallsPerBoard']
            for v in comp['lines'].values()
            if v['falseCallsPerBoard'] is not None
        ]
        if len(fc_rates) > 1:
            false_call_variance.append(
                {
                    'assembly': comp['assembly'],
                    'stddev': _stddev(fc_rates),
                }
            )

    defect_similarity = []
    for comp in assembly_comparisons:
        lines = list(comp['lines'].items())
        mixes = [
            (line, info['defectMix'])
            for line, info in lines
            if info['defectMix']
        ]
        if len(mixes) < 2:
            continue
        categories: set[str] = set()
        for _, mix in mixes:
            categories.update(mix.keys())
        categories = sorted(categories)
        pairwise = []
        for (line_a, mix_a), (line_b, mix_b) in combinations(mixes, 2):
            vec_a = [mix_a.get(cat, 0.0) for cat in categories]
            vec_b = [mix_b.get(cat, 0.0) for cat in categories]
            norm_a = math.sqrt(sum(v * v for v in vec_a))
            norm_b = math.sqrt(sum(v * v for v in vec_b))
            if not norm_a or not norm_b:
                similarity = 0.0
            else:
                similarity = sum(a * b for a, b in zip(vec_a, vec_b)) / (norm_a * norm_b)
            pairwise.append({
                'lines': [line_a, line_b],
                'similarity': similarity,
            })
        if pairwise:
            defect_similarity.append(
                {
                    'assembly': comp['assembly'],
                    'pairs': pairwise,
                }
            )

    line_drift = []
    for trend in line_trends:
        yields = [
            _resolve_yield(
                entry.get('windowYield'),
                entry.get('truePartYield'),
                entry.get('rawPartYield'),
            )
            for entry in trend['entries']
            if any(
                value is not None
                for value in (
                    entry.get('windowYield'),
                    entry.get('truePartYield'),
                    entry.get('rawPartYield'),
                )
            )
        ]
        if len(yields) < 2:
            continue
        line_drift.append(
            {
                'line': trend['line'],
                'start': yields[0],
                'end': yields[-1],
                'change': yields[-1] - yields[0],
            }
        )

    assembly_learning = []
    for assembly, line_map in assembly_daily.items():
        for line, entries in line_map.items():
            ordered = sorted(entries.items())
            yields = []
            for dt, info in ordered:
                confirmed = max(0.0, info['ng_parts'] - info['false_calls'])
                if info['windows']:
                    yield_pct = (
                        100.0 * (info['windows'] - info['ng_windows']) / info['windows']
                    )
                elif info['parts']:
                    yield_pct = (
                        100.0 * (info['parts'] - confirmed) / info['parts']
                    )
                else:
                    yield_pct = None
                if yield_pct is not None:
                    yields.append((dt.isoformat(), yield_pct))
            if len(yields) >= 2:
                assembly_learning.append(
                    {
                        'assembly': assembly,
                        'line': line,
                        'start': yields[0],
                        'end': yields[-1],
                        'change': yields[-1][1] - yields[0][1],
                    }
                )

    cross_line = {
        'yieldVariance': sorted(yield_variance, key=lambda x: x['stddev'], reverse=True),
        'falseCallVariance': sorted(
            false_call_variance, key=lambda x: x['stddev'], reverse=True
        ),
        'defectSimilarity': defect_similarity,
    }

    trend_insights = {
        'lineDrift': line_drift,
        'assemblyLearning': assembly_learning,
    }

    def _summary_yield_value(summary: dict[str, object]) -> float | None:
        for key in ('true_part_yield_pct', 'window_yield_pct', 'raw_part_yield_pct'):
            value = summary.get(key)
            if value is not None:
                return value
        return None

    focus_summary = max(
        (s for s in line_period_summaries if s.get('total_boards', 0.0) > 0.0),
        key=lambda item: item['total_boards'],
        default=None,
    )
    if focus_summary is None and line_period_summaries:
        focus_summary = line_period_summaries[0]

    ranked_summaries = [
        (summary, _summary_yield_value(summary))
        for summary in line_period_summaries
    ]
    ranked_summaries = [
        (summary, value)
        for summary, value in ranked_summaries
        if value is not None
    ]

    best_summary = max(ranked_summaries, key=lambda item: item[1], default=None)
    worst_summary = min(ranked_summaries, key=lambda item: item[1], default=None)

    overall_defects_per_board = (
        _safe_ratio(overall['ng_windows'], overall['total_boards'])
        if overall['total_boards']
        else None
    )

    line_period_summary = {
        'lines': line_period_summaries,
        'focus': focus_summary,
        'best': best_summary[0] if best_summary else None,
        'worst': worst_summary[0] if worst_summary else None,
        'overall': {
            'line': 'All Lines',
            'true_part_yield_pct': company_true_part_yield,
            'window_yield_pct': company_window_yield,
            'raw_part_yield_pct': company_raw_part_yield,
            'fc_per_board': company_fc_rate,
            'defects_per_board': overall_defects_per_board,
            'total_boards': overall['total_boards'],
            'total_parts': overall['total_parts'],
            'total_windows': overall['total_windows'],
            'false_calls': overall['false_calls'],
            'ng_windows': overall['ng_windows'],
        },
        'dateRange': {
            'start': start.isoformat() if start else None,
            'end': end.isoformat() if end else None,
        },
    }

    return {
        'lineMetrics': line_metrics,
        'assemblyComparisons': assembly_comparisons,
        'crossLine': cross_line,
        'lineTrends': line_trends,
        'trendInsights': trend_insights,
        'benchmarking': benchmarking,
        'linePeriodSummary': line_period_summary,
        'companyAverages': {
            'windowYield': company_window_yield,
            'rawPartYield': company_raw_part_yield,
            'truePartYield': company_true_part_yield,
            'falseCallsPerBoard': company_fc_rate,
            'falseCallPpm': company_false_call_ppm,
            'falseCallDpm': company_false_call_dpm,
            'defectDpm': company_dpm,
            'ngParts': overall['ng_parts'],
            'ngWindows': overall['ng_windows'],
            'windowsPerBoard': (
                _safe_ratio(overall['total_windows'], overall['total_boards'])
                if overall['total_boards']
                else None
            ),
            'defectsPerBoard': (
                _safe_ratio(overall['ng_windows'], overall['total_boards'])
                if overall['total_boards']
                else None
            ),
        },
    }


def _generate_line_report_charts(payload: dict) -> dict[str, str]:
    if plt is None:
        return {
            'lineYieldOverlayImg': '',
            'lineFalseCallSmallMultiplesImg': '',
            'lineDefectSmallMultiplesImg': '',
            'linePpmDpmComparisonImg': '',
        }

    import matplotlib.dates as mdates

    def _empty_chart(message: str):
        fig, ax = plt.subplots(figsize=(8, 3))
        ax.text(0.5, 0.5, message, ha='center', va='center', transform=ax.transAxes)
        ax.axis('off')
        return fig

    def _sanitize(values: list[float | None]) -> list[float]:
        cleaned: list[float] = []
        for value in values:
            if value is None:
                cleaned.append(float('nan'))
            else:
                try:
                    cleaned.append(float(value))
                except (TypeError, ValueError):
                    cleaned.append(float('nan'))
        return cleaned

    def _ratio(num, den) -> float | None:
        try:
            if den:
                return float(num) / float(den)
        except (TypeError, ValueError):
            return None
        return None

    charts: dict[str, str] = {}

    trends = payload.get('lineTrends', []) or []

    aggregated: dict[str, dict[str, float]] = {}
    for trend in trends:
        for entry in trend.get('entries', []) or []:
            date_str = entry.get('date')
            if not date_str:
                continue
            bucket = aggregated.setdefault(
                date_str,
                {
                    'boards': 0.0,
                    'parts': 0.0,
                    'ng_parts': 0.0,
                    'false_calls': 0.0,
                    'windows': 0.0,
                    'ng_windows': 0.0,
                },
            )
            bucket['boards'] += float(entry.get('boards') or 0.0)
            bucket['parts'] += float(entry.get('parts') or 0.0)
            bucket['ng_parts'] += float(entry.get('ngParts') or 0.0)
            bucket['false_calls'] += float(entry.get('falseCalls') or 0.0)
            bucket['windows'] += float(entry.get('windows') or 0.0)
            bucket['ng_windows'] += float(entry.get('ngWindows') or 0.0)

    aggregated_dates = sorted(aggregated.keys())
    parsed_dates: list[datetime] = []
    window_yield: list[float | None] = []
    true_part_yield: list[float | None] = []
    raw_part_yield: list[float | None] = []
    false_call_ppm: list[float | None] = []
    defect_dpm: list[float | None] = []

    for date_str in aggregated_dates:
        try:
            dt = datetime.fromisoformat(date_str)
        except ValueError:
            continue
        totals = aggregated[date_str]
        boards = totals['boards']
        parts = totals['parts']
        ng_parts = totals['ng_parts']
        false_calls = totals['false_calls']
        windows = totals['windows']
        ng_windows = totals['ng_windows']

        parsed_dates.append(dt)
        window_yield.append(
            (100.0 * (windows - ng_windows) / windows) if windows else None
        )
        raw_part_yield.append(
            (100.0 * (parts - ng_parts) / parts) if parts else None
        )
        confirmed_parts = max(0.0, ng_parts - false_calls)
        true_part_yield.append(
            (100.0 * (parts - confirmed_parts) / parts) if parts else None
        )
        rate = _ratio(false_calls, parts) if parts else None
        false_call_ppm.append(rate * 1_000_000 if rate is not None else None)
        if windows:
            defect_rate = _ratio(ng_windows, windows)
            defect_dpm.append(
                defect_rate * 1_000_000 if defect_rate is not None else None
            )
        elif parts:
            defect_rate = _ratio(confirmed_parts, parts)
            defect_dpm.append(
                defect_rate * 1_000_000 if defect_rate is not None else None
            )
        else:
            defect_dpm.append(None)

    # Overlay chart for company-wide yields over time
    if parsed_dates:
        fig, ax = plt.subplots(figsize=(8, 4))
        ax.plot(
            parsed_dates,
            _sanitize(window_yield),
            marker='o',
            linewidth=2,
            color='teal',
            label='Window Yield %',
        )
        ax.plot(
            parsed_dates,
            _sanitize(true_part_yield),
            marker='s',
            linewidth=2,
            color='slateblue',
            label='True Part Yield %',
        )
        ax.plot(
            parsed_dates,
            _sanitize(raw_part_yield),
            marker='^',
            linewidth=2,
            color='darkorange',
            label='Raw Part Yield %',
        )
        ax.set_ylabel('Yield %')
        ax.set_xlabel('Date')
        ax.set_title('Yield Performance Over Time')
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%b %d'))
        ax.tick_params(axis='x', rotation=45)
        ax.grid(alpha=0.3, linestyle='--')
        ax.legend(loc='best')
        fig.tight_layout()
    else:
        fig = _empty_chart('No yield trend data available')
    charts['lineYieldOverlayImg'] = _fig_to_data_uri(fig)

    def _extract_series(metric_key: str) -> list[tuple[str, list[datetime], list[float | None]]]:
        series: list[tuple[str, list[datetime], list[float | None]]] = []
        for trend in sorted(trends, key=lambda t: t.get('line') or ''):
            entries = trend.get('entries', []) or []
            dates: list[datetime] = []
            values: list[float | None] = []
            for entry in entries:
                date_str = entry.get('date')
                if not date_str:
                    continue
                try:
                    dt = datetime.fromisoformat(date_str)
                except ValueError:
                    continue
                dates.append(dt)
                values.append(entry.get(metric_key))
            if dates:
                series.append((trend.get('line') or 'Line', dates, values))
        return series

    def _build_small_multiples(
        metric_key: str,
        ylabel: str,
        title: str,
    ) -> str:
        series = _extract_series(metric_key)
        if not series:
            fig = _empty_chart(f'No {ylabel.lower()} data available')
            return _fig_to_data_uri(fig)

        cols = 2 if len(series) > 1 else 1
        rows = math.ceil(len(series) / cols)
        fig, axes = plt.subplots(
            rows,
            cols,
            figsize=(cols * 3.2, rows * 2.6),
            sharex=True,
        )
        axes_iter = axes.flat if hasattr(axes, 'flat') else [axes]
        axes_list = list(axes_iter)

        for idx, ax in enumerate(axes_list):
            if idx < len(series):
                line_name, dates, values = series[idx]
                ax.plot(
                    dates,
                    _sanitize(values),
                    marker='o',
                    linewidth=1.5,
                    color='tab:blue' if 'False' in ylabel else 'tab:green',
                )
                ax.set_title(line_name)
                ax.set_ylabel(ylabel)
                ax.grid(alpha=0.3, linestyle='--')
                ax.xaxis.set_major_formatter(mdates.DateFormatter('%b %d'))
                ax.tick_params(axis='x', rotation=45)
            else:
                ax.axis('off')

        fig.suptitle(title)
        fig.tight_layout(rect=[0, 0, 1, 0.94])
        return _fig_to_data_uri(fig)

    charts['lineFalseCallSmallMultiplesImg'] = _build_small_multiples(
        'falseCallsPerBoard',
        'False Calls / Board',
        'False Calls per Board by Line',
    )

    charts['lineDefectSmallMultiplesImg'] = _build_small_multiples(
        'defectsPerBoard',
        'Defects / Board',
        'Defects per Board by Line',
    )

    # Dual-axis comparison of false-call PPM vs. defect DPM across the date range
    if parsed_dates:
        fig, ax_left = plt.subplots(figsize=(8, 4))
        ax_right = ax_left.twinx()

        ax_left.plot(
            parsed_dates,
            _sanitize(false_call_ppm),
            marker='o',
            color='tab:blue',
            linewidth=1.8,
            label='False Call PPM',
        )
        ax_right.plot(
            parsed_dates,
            _sanitize(defect_dpm),
            marker='s',
            color='tab:red',
            linewidth=1.8,
            label='Defect DPM',
        )

        ax_left.set_xlabel('Date')
        ax_left.set_ylabel('False Call PPM')
        ax_right.set_ylabel('Defect DPM')
        ax_left.set_title('False Call PPM vs Defect DPM')
        ax_left.xaxis.set_major_formatter(mdates.DateFormatter('%b %d'))
        ax_left.tick_params(axis='x', rotation=45)
        ax_left.grid(alpha=0.3, linestyle='--')

        left_handles, left_labels = ax_left.get_legend_handles_labels()
        right_handles, right_labels = ax_right.get_legend_handles_labels()
        ax_left.legend(
            left_handles + right_handles,
            left_labels + right_labels,
            loc='best',
        )
        fig.tight_layout()
    else:
        fig = _empty_chart('No PPM/DPM trend data available')
    charts['linePpmDpmComparisonImg'] = _fig_to_data_uri(fig)

    return charts


def _generate_operator_report_charts(payload):
    """Generate charts for the operator report.

    This now produces a single chart with boards inspected as bars and
    reject rates as a line on a secondary y-axis.
    """
    if plt is None:
        return {"dailyImg": ""}

    charts: dict[str, str] = {}

    daily = payload.get("daily", {})
    dates = daily.get("dates", [])
    inspected = daily.get("inspected", [])
    rates = daily.get("rejectRates", [])

    fig, ax = plt.subplots(figsize=(8, 4))
    ax2 = ax.twinx()

    if dates and inspected:
        ax.bar(dates, inspected, color="steelblue", label="Boards Inspected")
        ax.set_ylabel("Boards Inspected")
    if dates and rates:
        ax2.plot(dates, rates, marker="o", color="crimson", label="Reject %")
        ax2.set_ylabel("Reject %")

    ax.set_xlabel("Date")
    ax.set_title("Daily Reject Rate and Boards Inspected")
    ax.tick_params(axis="x", rotation=45)

    lines, labels = [], []
    for a in (ax, ax2):
        l, lab = a.get_legend_handles_labels()
        lines.extend(l)
        labels.extend(lab)
    ax.legend(lines, labels, loc="best")

    charts["dailyImg"] = _fig_to_data_uri(fig)

    return charts


def _generate_aoi_daily_report_charts(payload):
    """Generate charts for the AOI daily report.

    Creates a simple bar chart showing the total quantity inspected for
    1st and 2nd shift.
    """
    s1 = payload.get("shift1_total")
    s2 = payload.get("shift2_total")
    s1 = s1 if s1 is not None else payload.get("shiftTotals", {}).get("shift1", {}).get("inspected", 0)
    s2 = s2 if s2 is not None else payload.get("shiftTotals", {}).get("shift2", {}).get("inspected", 0)
    s1_pct = payload.get("shift1_reject_pct")
    s2_pct = payload.get("shift2_reject_pct")
    s1_pct = s1_pct if s1_pct is not None else payload.get("shiftTotals", {}).get("shift1", {}).get("rejectRate", 0)
    s2_pct = s2_pct if s2_pct is not None else payload.get("shiftTotals", {}).get("shift2", {}).get("rejectRate", 0)

    if s1 > s2:
        diff = s1 - s2
        desc = f"1st shift inspected {diff} more boards than 2nd shift"
        pct_diff = s1_pct - s2_pct
        if pct_diff > 0:
            desc += f" and had a reject rate {pct_diff:.2f} percentage points higher."
        elif pct_diff < 0:
            desc += f" and had a reject rate {abs(pct_diff):.2f} percentage points lower."
        else:
            desc += " and had the same reject rate."
    elif s2 > s1:
        diff = s2 - s1
        desc = f"2nd shift inspected {diff} more boards than 1st shift"
        pct_diff = s2_pct - s1_pct
        if pct_diff > 0:
            desc += f" and had a reject rate {pct_diff:.2f} percentage points higher."
        elif pct_diff < 0:
            desc += f" and had a reject rate {abs(pct_diff):.2f} percentage points lower."
        else:
            desc += " and had the same reject rate."
    else:
        desc = "Both shifts inspected the same number of boards"
        pct_diff = s1_pct - s2_pct
        if pct_diff > 0:
            desc += f", but 1st shift's reject rate was {pct_diff:.2f} percentage points higher."
        elif pct_diff < 0:
            desc += f", but 2nd shift's reject rate was {abs(pct_diff):.2f} percentage points higher."
        else:
            desc += " and had the same reject rate."

    if plt is None:
        return {"shiftImg": "", "shiftImgDesc": desc}

    fig, ax = plt.subplots(figsize=(6, 4))
    ax.bar(["1st", "2nd"], [s1, s2], color="steelblue")
    ax.set_ylabel("Boards Inspected")
    ax.set_title("Boards Inspected by Shift")

    charts: dict[str, str] = {}
    charts["shiftImg"] = _fig_to_data_uri(fig)
    charts["shiftImgDesc"] = desc
    return charts


def build_report_payload(start=None, end=None):
    """Aggregate yield, operator and false-call stats for the AOI integrated report.

    This function now merges AOI report rows that have not yet been
    incorporated into ``combined_reports`` so the AOI Integrated Report can
    surface newly uploaded AOI data without waiting for the combined table
    to refresh.
    """
    combined, error = fetch_combined_reports()
    if error:
        current_app.logger.error("Combined report fetch failed: %s", error)

    phrases = current_app.config.get('NON_AOI_PHRASES', [])
    from collections import defaultdict

    by_date = defaultdict(lambda: {'inspected': 0.0, 'aoi_rej': 0.0, 'fi_rej': 0.0})
    by_assembly = defaultdict(lambda: {'inspected': 0.0, 'aoi_rej': 0.0, 'fi_rej': 0.0})
    # Track job numbers found in combined data so we can avoid double counting
    combined_jobs: set[str | None] = set()

    for row in combined or []:
        dt = _parse_date(row.get('aoi_Date') or row.get('Date') or row.get('date'))
        if start and (not dt or dt < start):
            continue
        if end and (not dt or dt > end):
            continue

        inspected = float(row.get('aoi_Quantity Inspected') or row.get('Quantity Inspected') or 0)
        aoi_rej = float(row.get('aoi_Quantity Rejected') or row.get('Quantity Rejected') or 0)

        fi_rej_val = row.get('fi_Quantity Rejected')
        try:
            fi_rej = float(fi_rej_val)
        except (TypeError, ValueError):
            fi_rej = 0.0
        if fi_rej == 0.0:
            info = row.get('fi_Additional Information') or row.get('fi_Add Info') or ''
            fi_rej = parse_fi_rejections(info, phrases)

        assembly = row.get('aoi_Assembly') or row.get('Assembly') or 'Unknown'
        job_number = row.get('aoi_Job Number') or row.get('Job Number')

        by_date[dt]['inspected'] += inspected
        by_date[dt]['aoi_rej'] += aoi_rej
        by_date[dt]['fi_rej'] += fi_rej

        by_assembly[assembly]['inspected'] += inspected
        by_assembly[assembly]['aoi_rej'] += aoi_rej
        by_assembly[assembly]['fi_rej'] += fi_rej

        combined_jobs.add(job_number)

    # Operator statistics now come from AOI reports exclusively, and we also
    # merge any AOI-only rows into the by_date/by_assembly aggregations above.
    aoi_reports, error = fetch_aoi_reports()
    if error:
        abort(500, description=error)

    by_operator = defaultdict(lambda: {'inspected': 0.0, 'rejected': 0.0})
    for row in aoi_reports or []:
        dt = _parse_date(row.get('aoi_Date') or row.get('Date'))
        if start and (not dt or dt < start):
            continue
        if end and (not dt or dt > end):
            continue

        inspected = float(row.get('aoi_Quantity Inspected') or row.get('Quantity Inspected') or 0)
        rejected = float(row.get('aoi_Quantity Rejected') or row.get('Quantity Rejected') or 0)
        operator = row.get('aoi_Operator') or row.get('Operator') or 'Unknown'
        by_operator[operator]['inspected'] += inspected
        by_operator[operator]['rejected'] += rejected

        job_number = row.get('aoi_Job Number') or row.get('Job Number')
        # Only augment yield statistics if this job was absent from combined_reports
        if job_number not in combined_jobs:
            assembly = row.get('aoi_Assembly') or row.get('Assembly') or 'Unknown'
            by_date[dt]['inspected'] += inspected
            by_date[dt]['aoi_rej'] += rejected
            by_assembly[assembly]['inspected'] += inspected
            by_assembly[assembly]['aoi_rej'] += rejected
            combined_jobs.add(job_number)

    dates = sorted(d for d in by_date.keys() if d)
    yields = []
    for d in dates:
        vals = by_date[d]
        rej = vals['aoi_rej'] + vals['fi_rej']
        y = ((vals['inspected'] - rej) / vals['inspected'] * 100.0) if vals['inspected'] else 0.0
        yields.append(y)

    assembly_yields = {}
    for asm, vals in by_assembly.items():
        rej = vals['aoi_rej'] + vals['fi_rej']
        assembly_yields[asm] = ((vals['inspected'] - rej) / vals['inspected'] * 100.0) if vals['inspected'] else 0.0

    operator_rows = [
        {
            'name': op,
            'inspected': vals['inspected'],
            'rejected': vals['rejected'],
        }
        for op, vals in by_operator.items()
    ]

    moat, error = fetch_moat()
    if error:
        abort(500, description=error)
    model_group = defaultdict(lambda: {'fc': 0.0, 'boards': 0.0})
    fc_vs_ng = defaultdict(lambda: {'ng': 0.0, 'fc': 0.0, 'parts': 0.0})
    fc_ng_ratio = defaultdict(lambda: {'fc': 0.0, 'ng': 0.0})
    for row in moat or []:
        dt = _parse_date(row.get('Report Date') or row.get('report_date'))
        if start and (not dt or dt < start):
            continue
        if end and (not dt or dt > end):
            continue
        model = (
            row.get('Model')
            or row.get('model')
            or row.get('Model Name')
            or row.get('model_name')
            or 'Unknown'
        )
        fc = float(row.get('FalseCall Parts') or row.get('falsecall_parts') or 0)
        boards = float(row.get('Total Boards') or row.get('total_boards') or 0)
        model_group[model]['fc'] += fc
        model_group[model]['boards'] += boards

        parts = float(row.get('Total Parts') or row.get('total_parts') or 0)
        ng_parts_val = row.get('NG Parts') or row.get('ng_parts')
        if ng_parts_val is not None:
            try:
                ng_parts = float(ng_parts_val)
            except (TypeError, ValueError):
                ng_parts = 0.0
        else:
            ng_ppm_val = (
                row.get('NG PPM')
                or row.get('ng_ppm')
                or row.get('NG_PPM')
                or 0
            )
            try:
                ng_ppm = float(ng_ppm_val)
            except (TypeError, ValueError):
                ng_ppm = 0.0
            ng_parts = (parts * ng_ppm) / 1_000_000 if parts and ng_ppm else 0.0
        ag = fc_vs_ng[dt]
        ag['ng'] += ng_parts
        ag['fc'] += fc
        ag['parts'] += parts

        fc_ng_ratio[model]['fc'] += fc
        fc_ng_ratio[model]['ng'] += ng_parts

    model_rows = []
    for model, vals in model_group.items():
        fc_per_board = (vals['fc'] / vals['boards']) if vals['boards'] else 0.0
        model_rows.append({'name': model, 'falseCalls': fc_per_board})

    combined_ratios = []
    for model, vals in fc_ng_ratio.items():
        ng_val = vals['ng']
        if ng_val <= 2:
            continue
        ratio = (vals['fc'] / ng_val) if ng_val else 0.0
        combined_ratios.append(
            {'model': model, 'fc': vals['fc'], 'ng': ng_val, 'ratio': ratio}
        )
    combined_ratios.sort(key=lambda x: x['ratio'], reverse=True)
    top_ratios = combined_ratios[:10]
    fc_ng_ratio_data = {
        'models': [t['model'] for t in top_ratios],
        'fcParts': [t['fc'] for t in top_ratios],
        'ngParts': [t['ng'] for t in top_ratios],
        'ratios': [t['ratio'] for t in top_ratios],
    }
    fc_ng_ratio_summary = {
        'top': [{'name': t['model'], 'ratio': t['ratio']} for t in combined_ratios[:3]]
    }

    fc_vs_ng_dates = sorted(d for d in fc_vs_ng.keys() if d)
    ng_ppm_series: list[float] = []
    fc_ppm_series: list[float] = []
    for d in fc_vs_ng_dates:
        vals = fc_vs_ng[d]
        parts = vals['parts']
        ng_ppm_series.append((vals['ng'] / parts * 1_000_000) if parts else 0.0)
        fc_ppm_series.append((vals['fc'] / parts * 1_000_000) if parts else 0.0)

    # Correlation and trend for FC vs NG
    n = min(len(ng_ppm_series), len(fc_ppm_series))
    if n > 1:
        avg_ng = sum(ng_ppm_series) / n
        avg_fc = sum(fc_ppm_series) / n
        num = sum((ng_ppm_series[i] - avg_ng) * (fc_ppm_series[i] - avg_fc) for i in range(n))
        den_ng = sum((ng_ppm_series[i] - avg_ng) ** 2 for i in range(n))
        den_fc = sum((fc_ppm_series[i] - avg_fc) ** 2 for i in range(n))
        corr = num / math.sqrt(den_ng * den_fc) if den_ng and den_fc else 0.0
        fc_trend = (
            'increased'
            if fc_ppm_series[0] < fc_ppm_series[-1]
            else 'decreased' if fc_ppm_series[0] > fc_ppm_series[-1] else 'stable'
        )
    else:
        corr = 0.0
        fc_trend = 'stable'
    fc_vs_ng_summary = {'correlation': corr, 'fcTrend': fc_trend}

    # --- Precompute summaries -------------------------------------------------
    if yields:
        avg_yield = sum(yields) / len(yields)
        worst_idx = min(range(len(yields)), key=lambda i: yields[i])
        worst_day = {
            'date': dates[worst_idx].isoformat(),
            'yield': yields[worst_idx],
        }
    else:
        avg_yield = 0.0
        worst_day = {'date': None, 'yield': 0.0}

    if assembly_yields:
        worst_asm = min(assembly_yields.items(), key=lambda item: item[1])
        worst_assembly = {'assembly': worst_asm[0], 'yield': worst_asm[1]}
    else:
        worst_assembly = {'assembly': None, 'yield': 0.0}

    yield_summary = {
        'avg': avg_yield,
        'worstDay': worst_day,
        'worstAssembly': worst_assembly,
    }

    ops = []
    for op in operator_rows:
        inspected = op['inspected']
        rate = (op['rejected'] / inspected * 100.0) if inspected else 0.0
        ops.append({**op, 'rate': rate})

    total_boards = sum(o['inspected'] for o in operator_rows)
    avg_rate = sum(o['rate'] for o in ops) / len(ops) if ops else 0.0
    num_ops = len(ops)
    avg_boards = total_boards / num_ops if num_ops else 0.0
    if ops:
        min_op = min(ops, key=lambda o: o['rate'])
        max_op = max(ops, key=lambda o: o['rate'])
    else:
        min_op = max_op = {'name': None, 'rate': 0.0}

    operator_summary = {
        'totalBoards': total_boards,
        'avgRate': avg_rate,
        'min': {'name': min_op['name'], 'rate': min_op['rate']},
        'max': {'name': max_op['name'], 'rate': max_op['rate']},
        'avgBoards': avg_boards,
    }

    avg_fc = sum(m['falseCalls'] for m in model_rows) / len(model_rows) if model_rows else 0.0
    problem_assemblies = [m for m in model_rows if m['falseCalls'] > 20]
    model_summary = {
        'avgFalseCalls': avg_fc,
        'over20': [m['name'] for m in problem_assemblies],
    }
    dates_iso = [d.isoformat() for d in dates]
    yield_pairs = list(zip(dates_iso, yields))

    fc_vs_ng_dates_iso = [d.isoformat() for d in fc_vs_ng_dates]
    fc_vs_ng_pairs = list(zip(fc_vs_ng_dates_iso, ng_ppm_series, fc_ppm_series))

    fc_ng_ratio_pairs = list(
        zip(
            fc_ng_ratio_data['models'],
            fc_ng_ratio_data['fcParts'],
            fc_ng_ratio_data['ngParts'],
            fc_ng_ratio_data['ratios'],
        )
    )

    # Centralized targets for key metrics so deltas can be computed uniformly
    targets = {
        'avg_yield': 98.0,
        'operator_rate': 5.0,
        'false_calls': 10.0,
    }

    def _kpi(label, value, target_key):
        item = {'label': label, 'value': value}
        target = targets.get(target_key)
        if target is not None:
            item['target'] = target
            item['delta'] = value - target
        return item

    summary_kpis = [
        _kpi('Average Yield', yield_summary['avg'], 'avg_yield'),
        _kpi('Operator Defect Rate', operator_summary['avgRate'], 'operator_rate'),
        _kpi('False Calls per Board', model_summary['avgFalseCalls'], 'false_calls'),
    ]

    summary_actions = [
        {'label': m['name'], 'value': m['falseCalls']}
        for m in problem_assemblies
    ]

    top_risks = [
        _kpi(asm, assembly_yields[asm], 'avg_yield')
        for asm, _ in sorted(assembly_yields.items(), key=lambda x: x[1])[:3]
    ]

    summary_charts = [
        {'label': 'Yield Trend', 'data': yield_pairs},
        {'label': 'FC vs NG', 'data': fc_vs_ng_pairs},
    ]

    executive_summary = {
        'kpis': summary_kpis,
        'actions': summary_actions,
        'topRisks': top_risks,
        'charts': summary_charts,
    }

    highlights = summary_actions
    kpis = summary_kpis

    charts = {
        'yield': yield_pairs,
        'fcVsNg': fc_vs_ng_pairs,
        'fcNgRatio': fc_ng_ratio_pairs,
    }

    top_tables = {
        'operators': ops,
        'models': model_rows,
    }

    jobs = [
        {'label': op['name'], 'value': op['inspected']} for op in ops
    ]

    appendix = {
        'yield': yield_pairs,
        'fcVsNg': fc_vs_ng_pairs,
        'fcNgRatio': fc_ng_ratio_pairs,
    }

    return {
        'yieldData': {
            'dates': dates_iso,
            'yields': yields,
            'assemblyYields': assembly_yields,
        },
        'yield_pairs': yield_pairs,
        'operators': ops,
        'models': model_rows,
        'fcVsNgRate': {
            'dates': fc_vs_ng_dates_iso,
            'ngPpm': ng_ppm_series,
            'fcPpm': fc_ppm_series,
        },
        'fc_vs_ng_pairs': fc_vs_ng_pairs,
        'fcVsNgSummary': fc_vs_ng_summary,
        'fcNgRatio': fc_ng_ratio_data,
        'fc_ng_ratio_pairs': fc_ng_ratio_pairs,
        'fcNgRatioSummary': fc_ng_ratio_summary,
        'yieldSummary': yield_summary,
        'operatorSummary': operator_summary,
        'modelSummary': model_summary,
        'problemAssemblies': problem_assemblies,
        'summary_kpis': summary_kpis,
        'summary_actions': summary_actions,
        'top_risks': top_risks,
        'summary_charts': summary_charts,
        'executive_summary': executive_summary,
        'highlights': highlights,
        'kpis': kpis,
        'charts': charts,
        'top_tables': top_tables,
        'jobs': jobs,
        'avgBoards': avg_boards,
        'appendix': appendix,
    }


def build_part_report_payload(
    start: date | None = None,
    end: date | None = None,
    *,
    page_size: int = 1000,
    fetcher=None,
) -> dict:
    """Aggregate part-level AOI analytics for the Part Report."""

    if fetcher is None:
        fetcher = fetch_part_results

    rows, error = fetcher(start_date=start, end_date=end, page_size=page_size)
    if error:
        abort(500, description=error)

    rows = [row for row in (rows or []) if isinstance(row, dict)]

    def _label(value: object, fallback: str = "Unknown") -> str:
        if value in (None, ""):
            return fallback
        text = str(value).strip()
        return text or fallback

    defect_code_counter: Counter[str] = Counter()
    component_family_counter: Counter[str] = Counter()
    assembly_counter: Counter[str] = Counter()
    line_counter: Counter[str] = Counter()
    program_counter: Counter[str] = Counter()
    part_total_counter: Counter[str] = Counter()
    part_false_counter: Counter[str] = Counter()
    defect_type_false_counter: Counter[str] = Counter()
    program_false_counter: Counter[str] = Counter()
    family_false_counter: Counter[str] = Counter()

    offset_x: list[float] = []
    offset_y: list[float] = []
    rotation_values: list[float] = []
    height_values: list[float] = []
    density_values: list[float] = []

    part_numbers: set[str] = set()
    assemblies: set[str] = set()
    lines: set[str] = set()
    programs: set[str] = set()
    operators: set[str] = set()
    boards: set[str] = set()

    false_calls: list[dict] = []

    daily_counts: dict[date, dict[str, float]] = defaultdict(
        lambda: {"defects": 0.0, "false_calls": 0.0}
    )

    operator_summary: dict[str, dict[str, object]] = defaultdict(
        lambda: {
            "total": 0.0,
            "false_calls": 0.0,
            "confirmed": 0.0,
            "dispositions": Counter(),
            "assemblies": Counter(),
            "lines": Counter(),
        }
    )
    process_summary: dict[str, dict[str, object]] = defaultdict(
        lambda: {
            "defects": 0.0,
            "false_calls": 0.0,
            "operators": Counter(),
        }
    )

    min_date: date | None = None
    max_date: date | None = None

    for row in rows:
        part_number = _label(row.get("part_number"))
        assembly = _label(row.get("assembly"))
        line_name = _label(row.get("line"))
        program_name = _label(row.get("program"))
        component_family = _label(row.get("component_family"))
        defect_code = _label(row.get("defect_code"))
        defect_type = _label(row.get("defect_type"))
        operator = _label(row.get("operator"), "Unassigned")
        disposition = _label(row.get("operator_disposition"), "Unreviewed")
        confirmation = _label(row.get("operator_confirmation"), "").lower()

        part_numbers.add(part_number)
        assemblies.add(assembly)
        lines.add(line_name)
        programs.add(program_name)
        operators.add(operator)

        board_serial = row.get("board_serial")
        if board_serial not in (None, ""):
            boards.add(str(board_serial))

        part_total_counter[part_number] += 1
        defect_code_counter[defect_code] += 1
        component_family_counter[component_family] += 1
        assembly_counter[assembly] += 1
        line_counter[line_name] += 1
        program_counter[program_name] += 1

        dt = _parse_date(row.get("inspection_date") or row.get("report_date"))
        if dt:
            stats = daily_counts[dt]
            stats["defects"] += 1
            min_date = min(min_date, dt) if min_date else dt
            max_date = max(max_date, dt) if max_date else dt

        offset_value = _coerce_number(row.get("offset_x"), default=None)
        if offset_value is not None:
            offset_x.append(offset_value)
        offset_value = _coerce_number(row.get("offset_y"), default=None)
        if offset_value is not None:
            offset_y.append(offset_value)
        rotation = _coerce_number(row.get("offset_theta") or row.get("rotation"), default=None)
        if rotation is not None:
            rotation_values.append(rotation)
        height = _coerce_number(row.get("height"), default=None)
        if height is not None:
            height_values.append(height)
        density = _coerce_number(row.get("defect_density"), default=None)
        if density is not None:
            density_values.append(density)

        is_false_call = bool(row.get("false_call"))
        if confirmation in {"confirmed", "reject", "true", "ng"}:
            is_false_call = False
        if is_false_call:
            false_calls.append(row)
            part_false_counter[part_number] += 1
            defect_type_false_counter[defect_type] += 1
            program_false_counter[program_name] += 1
            family_false_counter[component_family] += 1
            if dt:
                daily_counts[dt]["false_calls"] += 1

        bucket = operator_summary[operator]
        bucket["total"] += 1
        if is_false_call:
            bucket["false_calls"] += 1
        if confirmation in {"confirmed", "reject", "true", "ng"} or (
            disposition.lower() in {"confirmed", "reject", "true", "ng"}
        ):
            bucket["confirmed"] += 1
        bucket["dispositions"][disposition] += 1
        bucket["assemblies"][assembly] += 1
        bucket["lines"][line_name] += 1

        process_key = f"{line_name} • {program_name}"
        process_bucket = process_summary[process_key]
        process_bucket["defects"] += 1
        if is_false_call:
            process_bucket["false_calls"] += 1
        process_bucket["operators"][operator] += 1

    def _distribution(counter: Counter[str]) -> list[dict[str, object]]:
        total = sum(counter.values())
        results: list[dict[str, object]] = []
        for label, count in counter.most_common():
            share = _safe_ratio(count, total)
            results.append({"label": label, "count": count, "share": share})
        return results

    def _mean(values: list[float]) -> float:
        return float(mean(values)) if values else 0.0

    def _stdev(values: list[float]) -> float:
        return float(pstdev(values)) if len(values) > 1 else 0.0

    total_records = len(rows)
    total_false_calls = len(false_calls)
    total_boards = len(boards)
    defects_per_board = _safe_ratio(total_records, total_boards)
    false_calls_per_board = _safe_ratio(total_false_calls, total_boards)

    time_series = [
        {
            "date": dt.isoformat(),
            "defects": stats["defects"],
            "falseCalls": stats["false_calls"],
        }
        for dt, stats in sorted(daily_counts.items())
    ]

    operator_rows = []
    for name, bucket in sorted(
        operator_summary.items(), key=lambda item: item[1]["total"], reverse=True
    ):
        total = bucket["total"]
        false_count = bucket["false_calls"]
        confirmed = bucket["confirmed"]
        operator_rows.append(
            {
                "operator": name,
                "total": total,
                "falseCalls": false_count,
                "confirmed": confirmed,
                "falseCallRate": _safe_ratio(false_count, total),
                "confirmationRate": _safe_ratio(confirmed, total),
                "topDispositions": _distribution(bucket["dispositions"])[:5],
                "topAssemblies": _distribution(bucket["assemblies"])[:5],
                "topLines": _distribution(bucket["lines"])[:5],
            }
        )

    process_rows = []
    for key, bucket in sorted(
        process_summary.items(), key=lambda item: item[1]["defects"], reverse=True
    ):
        process_rows.append(
            {
                "process": key,
                "defects": bucket["defects"],
                "falseCalls": bucket["false_calls"],
                "falseCallRate": _safe_ratio(bucket["false_calls"], bucket["defects"]),
                "topOperators": _distribution(bucket["operators"])[:5],
            }
        )

    defect_by_code = _distribution(defect_code_counter)
    family_distribution = _distribution(component_family_counter)
    assembly_distribution = _distribution(assembly_counter)
    line_distribution = _distribution(line_counter)
    program_distribution = _distribution(program_counter)

    false_call_patterns = {
        "total": total_false_calls,
        "share": _safe_ratio(total_false_calls, total_records),
        "byPartNumber": _distribution(part_false_counter),
        "byDefectType": _distribution(defect_type_false_counter),
        "byProgram": _distribution(program_false_counter),
        "byFamily": _distribution(family_false_counter),
    }

    yield_reliability = {
        "defectsPerBoard": defects_per_board,
        "falseCallsPerBoard": false_calls_per_board,
        "criticalPartsPareto": _distribution(part_total_counter)[:10],
        "dailyTrend": time_series,
        "familyShare": family_distribution,
        "densityMean": _mean(density_values) if density_values else 0.0,
        "densityStdDev": _stdev(density_values) if density_values else 0.0,
    }

    spatial_metrics = {
        "offsets": {
            "meanX": _mean(offset_x),
            "meanY": _mean(offset_y),
            "absMeanX": _mean([abs(value) for value in offset_x]) if offset_x else 0.0,
            "absMeanY": _mean([abs(value) for value in offset_y]) if offset_y else 0.0,
            "stdevX": _stdev(offset_x),
            "stdevY": _stdev(offset_y),
            "samples": len(offset_x),
        },
        "rotation": {
            "mean": _mean(rotation_values),
            "stdev": _stdev(rotation_values),
            "samples": len(rotation_values),
        },
        "height": {
            "mean": _mean(height_values),
            "stdev": _stdev(height_values),
            "min": float(min(height_values)) if height_values else 0.0,
            "max": float(max(height_values)) if height_values else 0.0,
            "samples": len(height_values),
        },
    }

    highlights: list[str] = []
    opportunities: list[str] = []
    business_value: list[str] = []

    if defect_by_code:
        top = defect_by_code[0]
        highlights.append(
            f"Defect code {top['label']} represents {top['share'] * 100:.1f}% of all detections."
        )
    if false_call_patterns["share"] > 0.2:
        opportunities.append(
            "False-call share exceeds 20%, indicating additional review automation opportunities."
        )
    if yield_reliability["defectsPerBoard"] > 0.0:
        business_value.append(
            "Reducing critical part escapes directly improves first-pass yield and lowers rework cost."
        )
    if not highlights:
        highlights.append("Part-level performance remains stable across the selected period.")

    meta = {
        "totalRecords": total_records,
        "totalFalseCalls": total_false_calls,
        "dateRange": [
            min_date.isoformat() if min_date else None,
            max_date.isoformat() if max_date else None,
        ],
        "uniquePartNumbers": len(part_numbers),
        "uniqueAssemblies": len(assemblies),
        "uniquePrograms": len(programs),
        "uniqueLines": len(lines),
        "uniqueOperators": len(operators),
        "totalBoards": total_boards,
    }

    return {
        "meta": meta,
        "defectDistributions": {
            "byDefectCode": defect_by_code,
            "byComponentFamily": family_distribution,
            "byAssembly": assembly_distribution,
            "byLine": line_distribution,
            "byProgram": program_distribution,
        },
        "spatialMetrics": spatial_metrics,
        "falseCallPatterns": false_call_patterns,
        "yieldReliability": yield_reliability,
        "operatorLinkages": {
            "byOperator": operator_rows,
            "byProcess": process_rows,
        },
        "timeSeries": {
            "daily": time_series,
        },
        "insights": {
            "highlights": highlights,
            "opportunities": opportunities,
            "businessValue": business_value,
        },
    }


@main_bp.route('/api/reports/part', methods=['GET'])
@feature_required('reports_part')
def api_part_report():
    """Return aggregated part-level analytics."""

    if 'username' not in session:
        return redirect(url_for('auth.login'))

    start = _parse_date(request.args.get('start_date'))
    end = _parse_date(request.args.get('end_date'))

    payload = build_part_report_payload(start, end)
    payload['start'] = start.isoformat() if start else ''
    payload['end'] = end.isoformat() if end else ''
    return jsonify(payload)


@main_bp.route('/reports/part', methods=['GET'])
@feature_required('reports_part')
def part_report():
    """Render the Part Report landing page."""

    if 'username' not in session:
        return redirect(url_for('auth.login'))

    return render_template('part_report.html', username=session.get('username'))


@main_bp.route('/reports/part/export')
@feature_required('reports_part')
def export_part_report():
    if 'username' not in session:
        return redirect(url_for('auth.login'))

    start = _parse_date(request.args.get('start_date'))
    end = _parse_date(request.args.get('end_date'))
    start_str = start.strftime('%y%m%d') if start else ''
    end_str = end.strftime('%y%m%d') if end else ''

    payload = build_part_report_payload(start, end)
    payload['start'] = start.isoformat() if start else ''
    payload['end'] = end.isoformat() if end else ''

    body = request.get_json(silent=True) or {}

    def _get(name, default=''):
        return request.args.get(name, body.get(name, default))

    def _get_bool(name, default=True):
        value = request.args.get(name)
        if value is None:
            value = body.get(name)
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        return str(value).lower() not in {'0', 'false', 'no'}

    show_cover = _get_bool('show_cover')
    show_overview = _get_bool('show_overview', default=True)
    show_insights = _get_bool('show_insights', default=True)
    show_advanced = _get_bool('show_advanced', default=True)
    show_calculations = _get_bool('show_calculations', default=True)
    show_business = _get_bool('show_business', default=True)

    title = _get('title') or 'Part Report'
    subtitle = _get('subtitle')
    report_date = _get('report_date')
    period = _get('period')
    author = _get('author')
    logo_url = _get('logo_url') or url_for(
        'static', filename='images/company-logo.png', _external=True
    )
    footer_left = _get('footer_left')
    report_id = _get('report_id')
    contact = _get('contact', 'tschwartz@4spectra.com')
    confidentiality = _get('confidentiality', 'Spectra-Tech • Confidential')
    generated_at = datetime.now(_report_timezone()).strftime('%Y-%m-%d %H:%M:%S %Z')

    report_css = _load_report_css()

    context = {
        'show_cover': show_cover,
        'show_overview': show_overview,
        'show_insights': show_insights,
        'show_advanced': show_advanced,
        'show_calculations': show_calculations,
        'show_business': show_business,
        'title': title,
        'subtitle': subtitle,
        'report_date': report_date,
        'period': period,
        'author': author,
        'logo_url': logo_url,
        'footer_left': footer_left,
        'report_id': report_id,
        'contact': contact,
        'confidentiality': confidentiality,
        'generated_at': generated_at,
        'report_css': report_css,
        **payload,
    }

    html = render_template('report/part/index.html', **context)

    fmt = request.args.get('format') or 'html'
    filename_stem = f"{start_str}_{end_str}_part_report"
    if fmt == 'pdf':
        try:
            pdf = render_html_to_pdf(html, base_url=request.url_root)
        except PdfGenerationError as exc:
            return jsonify({'message': str(exc)}), 503
        return send_file(
            io.BytesIO(pdf),
            mimetype='application/pdf',
            download_name=f"{filename_stem}.pdf",
            as_attachment=True,
        )
    if fmt == 'html':
        return send_file(
            io.BytesIO(html.encode('utf-8')),
            mimetype='text/html',
            download_name=f"{filename_stem}.html",
            as_attachment=True,
        )
    return jsonify({'message': 'Unsupported format. Choose pdf or html.'}), 400


@main_bp.route('/api/reports/line', methods=['GET'])
@feature_required('reports_line')
def api_line_report():
    """Return aggregated line-level AOI metrics."""
    if 'username' not in session:
        return redirect(url_for('auth.login'))

    start = _parse_date(request.args.get('start_date'))
    end = _parse_date(request.args.get('end_date'))

    payload = build_line_report_payload(start, end)
    payload['start'] = start.isoformat() if start else ''
    payload['end'] = end.isoformat() if end else ''
    return jsonify(payload)


@main_bp.route('/reports/line', methods=['GET'])
@feature_required('reports_line')
def line_report():
    """Render the Line Report page."""
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    return render_template('line_report.html', username=session.get('username'))


@main_bp.route('/reports/line/export')
@feature_required('reports_line')
def export_line_report():
    if 'username' not in session:
        return redirect(url_for('auth.login'))

    start = _parse_date(request.args.get('start_date'))
    end = _parse_date(request.args.get('end_date'))
    start_str = start.strftime('%y%m%d') if start else ''
    end_str = end.strftime('%y%m%d') if end else ''

    payload = build_line_report_payload(start, end)
    payload['start'] = start.isoformat() if start else ''
    payload['end'] = end.isoformat() if end else ''
    charts = _generate_line_report_charts(payload)

    body = request.get_json(silent=True) or {}

    def _get(name, default=''):
        return request.args.get(name, body.get(name, default))

    def _get_bool(name, default=True):
        value = request.args.get(name)
        if value is None:
            value = body.get(name)
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        return str(value).lower() not in {'0', 'false', 'no'}

    show_cover = _get_bool('show_cover')
    show_summary = _get_bool('show_summary')
    title = _get('title') or 'Line Report'
    subtitle = _get('subtitle')
    report_date = _get('report_date')
    period = _get('period')
    author = _get('author')
    logo_url = _get('logo_url') or url_for(
        'static', filename='images/company-logo.png', _external=True
    )
    footer_left = _get('footer_left')
    report_id = _get('report_id')
    contact = _get('contact', 'tschwartz@4spectra.com')
    confidentiality = _get('confidentiality', 'Spectra-Tech • Confidential')
    generated_at = datetime.now(_report_timezone()).strftime('%Y-%m-%d %H:%M:%S %Z')

    report_css = _load_report_css()

    context = {
        'show_cover': show_cover,
        'show_summary': show_summary,
        'title': title,
        'subtitle': subtitle,
        'report_date': report_date,
        'period': period,
        'author': author,
        'logo_url': logo_url,
        'footer_left': footer_left,
        'report_id': report_id,
        'contact': contact,
        'confidentiality': confidentiality,
        'generated_at': generated_at,
        'report_css': report_css,
        **payload,
        **(charts or {}),
    }

    html = render_template('report/line/index.html', **context)

    fmt = request.args.get('format') or 'html'
    filename_stem = f"{start_str}_{end_str}_line_report"
    if fmt == 'pdf':
        try:
            pdf = render_html_to_pdf(html, base_url=request.url_root)
        except PdfGenerationError as exc:
            return jsonify({'message': str(exc)}), 503
        return send_file(
            io.BytesIO(pdf),
            mimetype='application/pdf',
            download_name=f"{filename_stem}.pdf",
            as_attachment=True,
        )
    if fmt == 'html':
        return send_file(
            io.BytesIO(html.encode('utf-8')),
            mimetype='text/html',
            download_name=f"{filename_stem}.html",
            as_attachment=True,
        )
    return jsonify({'message': 'Unsupported format. Choose pdf or html.'}), 400


@main_bp.route('/api/reports/integrated', methods=['GET'])
@feature_required('reports_integrated')
def api_integrated_report():
    """Aggregate yield, operator and false-call stats for the AOI integrated report."""
    if 'username' not in session:
        return redirect(url_for('auth.login'))

    start = _parse_date(request.args.get('start_date'))
    end = _parse_date(request.args.get('end_date'))

    payload = build_report_payload(start, end)
    return jsonify(payload)


@main_bp.route('/reports/integrated', methods=['GET'])
@feature_required('reports_integrated')
def integrated_report():
    """Render the AOI Integrated Report page."""
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    return render_template('integrated_report.html', username=session.get('username'))


@main_bp.route('/reports/integrated/export')
@feature_required('reports_integrated')
def export_integrated_report():
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    start = _parse_date(request.args.get('start_date'))
    end = _parse_date(request.args.get('end_date'))
    start_str = start.strftime('%y%m%d') if start else ''
    end_str = end.strftime('%y%m%d') if end else ''
    payload = build_report_payload(start, end)
    payload['start'] = start.isoformat() if start else ''
    payload['end'] = end.isoformat() if end else ''
    charts = _generate_report_charts(payload)
    body = request.get_json(silent=True) or {}

    def _get(name, default=''):
        return request.args.get(name, body.get(name, default))

    def _get_bool(name, default=True):
        value = request.args.get(name)
        if value is None:
            value = body.get(name)
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        return str(value).lower() not in {'0', 'false', 'no'}

    show_cover = _get_bool('show_cover')
    show_summary = _get_bool('show_summary')
    title = _get('title')
    subtitle = _get('subtitle')
    report_date = _get('report_date')
    period = _get('period')
    author = _get('author')
    logo_url = _get('logo_url') or url_for(
        'static', filename='images/company-logo.png', _external=True
    )
    footer_left = _get('footer_left')
    report_id = _get('report_id')
    contact = _get('contact', 'tschwartz@4spectra.com')
    confidentiality = _get('confidentiality', 'Spectra-Tech • Confidential')
    generated_at = datetime.now(_report_timezone()).strftime('%Y-%m-%d %H:%M:%S %Z')

    if show_summary:
        payload.setdefault(
            'yieldSummary',
            {
                'avg': 0.0,
                'worstDay': {'date': None, 'yield': 0.0},
                'worstAssembly': {'assembly': None, 'yield': 0.0},
            },
        )
        payload.setdefault('operatorSummary', payload.get('summary', {}))
        payload.setdefault('modelSummary', {'avgFalseCalls': 0.0})

    report_css = _load_report_css()

    html = render_template(
        'report/integrated/index.html',
        show_cover=show_cover,
        show_summary=show_summary,
        title=title,
        subtitle=subtitle,
        report_date=report_date,
        period=period,
        author=author,
        logo_url=logo_url,
        footer_left=footer_left,
        report_id=report_id,
        contact=contact,
        confidentiality=confidentiality,
        generated_at=generated_at,
        report_css=report_css,
        **payload,
        **charts,
    )
    fmt = request.args.get('format')
    filename_stem = f"{start_str}_{end_str}_aoiIR"
    if fmt == 'pdf':
        try:
            pdf = render_html_to_pdf(html, base_url=request.url_root)
        except PdfGenerationError as exc:
            return jsonify({'message': str(exc)}), 503
        return send_file(
            io.BytesIO(pdf),
            mimetype='application/pdf',
            download_name=f"{filename_stem}.pdf",
            as_attachment=True,
        )
    if fmt == 'html':
        return send_file(
            io.BytesIO(html.encode('utf-8')),
            mimetype='text/html',
            download_name=f"{filename_stem}.html",
            as_attachment=True,
        )
    return html


@main_bp.route('/reports/aoi_daily', methods=['GET'])
@feature_required('reports_aoi_daily')
def aoi_daily_report_page():
    """Render the AOI Daily Report page."""
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    return render_template('aoi_daily_report.html', username=session.get('username'))


@main_bp.route('/reports/aoi_daily/export')
@feature_required('reports_aoi_daily')
def export_aoi_daily_report():
    if 'username' not in session:
        return redirect(url_for('auth.login'))

    day = _parse_date(request.args.get('date'))
    if not day:
        abort(400, description='Invalid date')

    operator = request.args.get('operator') or None
    assembly = request.args.get('assembly') or None

    start = end = day.isoformat()
    generated_at = datetime.now(_report_timezone()).strftime('%Y-%m-%d %H:%M:%S %Z')
    contact = request.args.get('contact', 'tschawtz@4spectra.com')

    payload = build_aoi_daily_report_payload(day, operator, assembly)
    charts = _generate_aoi_daily_report_charts(payload)
    payload.update(charts)

    show_cover = (
        str(request.args.get('show_cover', 'true')).lower() not in {'0', 'false', 'no'}
    )
    report_css = _load_report_css()

    logo_url = request.args.get('logo_url') or url_for(
        'static', filename='images/company-logo.png', _external=True
    )

    html = render_template(
        'report/aoi_daily/index.html',
        day=day.isoformat(),
        show_cover=show_cover,
        start=start,
        end=end,
        generated_at=generated_at,
        contact=contact,
        logo_url=logo_url,
        report_css=report_css,
        **payload,
    )

    fmt = request.args.get('format')
    filename_stem = f"{day.strftime('%y%m%d')}_aoi_daily_report"
    if fmt == 'pdf':
        try:
            pdf = render_html_to_pdf(html, base_url=request.url_root)
        except PdfGenerationError as exc:
            return jsonify({'message': str(exc)}), 503
        return send_file(
            io.BytesIO(pdf),
            mimetype='application/pdf',
            download_name=f"{filename_stem}.pdf",
            as_attachment=True,
        )
    if fmt == 'html':
        return send_file(
            io.BytesIO(html.encode('utf-8')),
            mimetype='text/html',
            download_name=f"{filename_stem}.html",
            as_attachment=True,
        )
    return html


def _aggregate_operator_report(start=None, end=None, operator: str | None = None):
    """Aggregate AOI report rows for the operator report."""
    from collections import defaultdict

    # Normalize operator filter to a set of lowercase names
    operators = {
        o.strip().lower() for o in (operator or '').split(',') if o.strip()
    }

    rows, error = fetch_aoi_reports()
    if error:
        abort(500, description=error)

    daily = defaultdict(lambda: {'inspected': 0.0, 'rejected': 0.0})
    assemblies = defaultdict(lambda: {'inspected': 0.0, 'rejected': 0.0})
    total_inspected = 0.0
    total_rejected = 0.0
    unique_ops: set[str] = set()

    for row in rows or []:
        date_val = _parse_date(row.get('Date') or row.get('aoi_Date'))
        if start and date_val and date_val < start:
            continue
        if end and date_val and date_val > end:
            continue

        op_name = (row.get('Operator') or row.get('aoi_Operator') or '').strip()
        if operators and op_name.lower() not in operators:
            continue
        if op_name:
            unique_ops.add(op_name)

        inspected = float(
            row.get('Quantity Inspected')
            or row.get('quantity_inspected')
            or row.get('aoi_Quantity Inspected')
            or 0
        )
        rejected = float(
            row.get('Quantity Rejected')
            or row.get('quantity_rejected')
            or row.get('aoi_Quantity Rejected')
            or 0
        )

        if date_val:
            daily[date_val]['inspected'] += inspected
            daily[date_val]['rejected'] += rejected

        asm = row.get('Assembly') or row.get('aoi_Assembly') or 'Unknown'
        assemblies[asm]['inspected'] += inspected
        assemblies[asm]['rejected'] += rejected

        total_inspected += inspected
        total_rejected += rejected

    dates_sorted = sorted(daily.keys())
    daily_dates = [d.isoformat() for d in dates_sorted]
    daily_inspected = [daily[d]['inspected'] for d in dates_sorted]
    daily_reject_rates = [
        (daily[d]['rejected'] / daily[d]['inspected'] * 100)
        if daily[d]['inspected']
        else 0
        for d in dates_sorted
    ]

    num_days = len(dates_sorted)
    avg_per_shift = total_inspected / num_days if num_days else 0
    avg_reject_rate = (
        (total_rejected / total_inspected) * 100 if total_inspected else 0
    )
    num_ops = len(unique_ops)
    avg_boards = total_inspected / num_ops if num_ops else 0

    combined, error = fetch_combined_reports()
    if error:
        abort(500, description=error)

    fi_data = defaultdict(lambda: {'fi_rejected': 0.0, 'aoi_inspected': 0.0})
    for row in combined or []:
        date_val = _parse_date(row.get('aoi_Date') or row.get('Date'))
        if start and date_val and date_val < start:
            continue
        if end and date_val and date_val > end:
            continue

        op_name = (row.get('aoi_Operator') or row.get('Operator') or '').strip()
        if operators and op_name.lower() not in operators:
            continue

        asm = row.get('aoi_Assembly') or row.get('Assembly') or 'Unknown'
        fi_data[asm]['fi_rejected'] += float(row.get('fi_Quantity Rejected') or 0)
        fi_data[asm]['aoi_inspected'] += float(
            row.get('aoi_Quantity Inspected')
            or row.get('Quantity Inspected')
            or 0
        )

    assemblies_list = []
    for asm, counts in sorted(
        assemblies.items(), key=lambda x: x[1]['inspected'], reverse=True
    ):
        fi_info = fi_data.get(asm)
        if fi_info and fi_info['aoi_inspected']:
            fi_rate = fi_info['fi_rejected'] / fi_info['aoi_inspected'] * 100
        else:
            fi_rate = None
        assemblies_list.append(
            {
                'assembly': asm,
                'inspected': counts['inspected'],
                'rejected': counts['rejected'],
                'fiRejectRate': fi_rate,
            }
        )

    return {
        'daily': {
            'dates': daily_dates,
            'inspected': daily_inspected,
            'rejectRates': daily_reject_rates,
        },
        'summary': {
            'totalBoards': total_inspected,
            'avgPerShift': avg_per_shift,
            'avgRejectRate': avg_reject_rate,
            'avgBoards': avg_boards,
        },
        'assemblies': assemblies_list,
    }


@main_bp.route('/api/reports/operator', methods=['GET'])
@feature_required('reports_operator')
def api_operator_report():
    """Return operator report data filtered by date range and operator."""
    if 'username' not in session:
        return redirect(url_for('auth.login'))

    start = _parse_date(request.args.get('start_date'))
    end = _parse_date(request.args.get('end_date'))
    operator = request.args.get('operator') or None

    payload = _aggregate_operator_report(start, end, operator)
    return jsonify(payload)


def build_operator_report_payload(start=None, end=None, operator: str | None = None):
    """Build the operator report payload for export."""
    return _aggregate_operator_report(start, end, operator)


@main_bp.route('/api/reports/aoi_daily', methods=['GET'])
@feature_required('reports_aoi_daily')
def api_aoi_daily_report():
    """Return AOI daily report data for preview."""
    if 'username' not in session:
        return redirect(url_for('auth.login'))

    day = _parse_date(request.args.get('date'))
    if not day:
        abort(400, description='Invalid date')

    operator = request.args.get('operator') or None
    assembly = request.args.get('assembly') or None

    payload = build_aoi_daily_report_payload(day, operator, assembly)
    return jsonify(payload)


def build_aoi_daily_report_payload(
    day: date, operator: str | None = None, assembly: str | None = None
):
    """Build the AOI daily report payload for a specific day."""
    rows, error = fetch_aoi_reports()
    if error:
        abort(500, description=error)
    fi_rows, fi_error = fetch_fi_reports()
    if fi_error:
        abort(500, description=fi_error)

    op_filter = (
        {o.strip().lower() for o in operator.split(',') if o.strip()}
        if operator
        else None
    )
    asm_filter = (
        {a.strip().lower() for a in assembly.split(',') if a.strip()}
        if assembly
        else None
    )

    shift_rows = {"shift1": [], "shift2": []}
    shift_totals = {
        "shift1": {"inspected": 0, "rejected": 0},
        "shift2": {"inspected": 0, "rejected": 0},
    }
    assemblies: dict[str, dict[str, int | set]] = {}

    for row in rows or []:
        dt = _parse_date(row.get("Date") or row.get("date"))
        if not dt or dt != day:
            continue

        raw_shift = str(row.get("Shift") or "").lower()
        if raw_shift in {"1", "1st", "first", "shift 1", "shift1", "1st shift"}:
            shift_key = "shift1"
        elif raw_shift in {"2", "2nd", "second", "shift 2", "shift2", "2nd shift"}:
            shift_key = "shift2"
        else:
            continue

        op_name = row.get("Operator") or "Unknown"
        asm_name = row.get("Assembly") or "Unknown"
        program = row.get("Program") or "Unknown"
        if op_filter and op_name.lower() not in op_filter:
            continue
        if asm_filter and asm_name.lower() not in asm_filter:
            continue

        inspected = int(row.get("Quantity Inspected") or 0)
        rejected = int(row.get("Quantity Rejected") or 0)

        entry = {
            "operator": op_name,
            "program": program,
            "assembly": asm_name,
            "job": row.get("Job Number") or "",
            "inspected": inspected,
            "rejected": rejected,
        }
        shift_rows[shift_key].append(entry)
        shift_totals[shift_key]["inspected"] += inspected
        shift_totals[shift_key]["rejected"] += rejected

        assemblies.setdefault(
            asm_name,
            {"inspected": 0, "rejected": 0, "operators": set()},
        )
        assemblies[asm_name]["inspected"] += inspected
        assemblies[asm_name]["rejected"] += rejected
        assemblies[asm_name]["operators"].add(op_name)

    for info in shift_totals.values():
        ins = info["inspected"]
        rej = info["rejected"]
        info["rejectRate"] = (rej / ins * 100) if ins else 0

    # Aggregate FI typical rejects per assembly from historical FI data
    phrases = current_app.config.get("NON_AOI_PHRASES", [])
    fi_assembly: dict[str, list[int]] = defaultdict(list)
    for row in fi_rows or []:
        dt = _parse_date(row.get("Date") or row.get("date"))
        if not dt or dt >= day:
            continue
        asm = row.get("Assembly") or "Unknown"
        try:
            fi_rej = int(row.get("Quantity Rejected") or 0)
        except (TypeError, ValueError):
            fi_rej = 0
        if fi_rej == 0:
            info = row.get("Additional Information") or row.get("Add Info") or ""
            fi_rej = parse_fi_rejections(info, phrases)
        fi_assembly[asm].append(fi_rej)

    assembly_info = []
    for asm, vals in assemblies.items():
        ins = vals["inspected"]
        rej = vals["rejected"]
        today_yield = ((ins - rej) / ins * 100) if ins else 0

        # Past 4 job average yield and reject count
        past_rows = [
            r
            for r in rows or []
            if (r.get("Assembly") or "Unknown") == asm
            and (d := _parse_date(r.get("Date") or r.get("date")))
            and d < day
        ]
        job_groups: dict[str, dict[str, int | date | None]] = {}
        for r in past_rows:
            job = r.get("Job Number") or ""
            d = _parse_date(r.get("Date") or r.get("date"))
            g = job_groups.setdefault(
                job, {"inspected": 0, "rejected": 0, "date": d}
            )
            g["inspected"] += int(r.get("Quantity Inspected") or 0)
            g["rejected"] += int(r.get("Quantity Rejected") or 0)
            if d and (g["date"] is None or d > g["date"]):
                g["date"] = d

        jobs = sorted(
            job_groups.values(), key=lambda g: g.get("date") or date.min, reverse=True
        )
        yields: list[float] = []
        rejects: list[int] = []
        for g in jobs:
            i = g["inspected"]
            rj = g["rejected"]
            rejects.append(rj)
            if i:
                yields.append((i - rj) / i * 100)
        past_avg: float | str
        if yields:
            past_avg = sum(yields) / len(yields)
        else:
            past_avg = "first run"
        past_rej_avg = sum(rejects) / len(rejects) if rejects else 0

        fi_vals = fi_assembly.get(asm, [])
        fi_typical = sum(fi_vals) / len(fi_vals) if fi_vals else 0

        info = {
            "assembly": asm,
            "operators": sorted(vals.get("operators", set())),
            "boards": ins,
            "yield": today_yield,
            "pastAvg": past_avg,
            "currentRejects": rej,
            "pastRejectsAvg": past_rej_avg,
            "fiTypicalRejects": fi_typical,
        }
        info["metricsChart"] = _build_metrics_chart(info)
        assembly_info.append(info)
    moat_rows, moat_error = fetch_moat()
    if moat_error:
        current_app.logger.error("Failed to fetch MOAT data: %s", moat_error)
        moat_rows = []
    for asm in assembly_info:
        asm["overlayChart"] = _build_assembly_moat_charts(
            asm["assembly"], moat_rows
        ).get("overlayChart", "")

    # Compute overall shift summary statistics for template consumption
    s1_total = shift_totals["shift1"].get("inspected", 0)
    s2_total = shift_totals["shift2"].get("inspected", 0)
    s1_reject_pct = round(shift_totals["shift1"].get("rejectRate", 0), 2)
    s2_reject_pct = round(shift_totals["shift2"].get("rejectRate", 0), 2)
    shift_total_diff = abs(s1_total - s2_total)
    shift_reject_pct_diff = round(abs(s1_reject_pct - s2_reject_pct), 2)

    return {
        "date": day.isoformat(),
        "shift1": shift_rows["shift1"],
        "shift2": shift_rows["shift2"],
        "shiftTotals": shift_totals,
        "shift1_total": s1_total,
        "shift2_total": s2_total,
        "shift1_reject_pct": s1_reject_pct,
        "shift2_reject_pct": s2_reject_pct,
        "shift_total_diff": shift_total_diff,
        "shift_reject_pct_diff": shift_reject_pct_diff,
        "assemblies": assembly_info,
    }


@main_bp.route('/reports/operator', methods=['GET'])
@feature_required('reports_operator')
def operator_report():
    """Render the Operator Report page."""
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    return render_template('operator_report.html', username=session.get('username'))


@main_bp.route('/reports/operator/export')
@feature_required('reports_operator')
def export_operator_report():
    if 'username' not in session:
        return redirect(url_for('auth.login'))

    start = _parse_date(request.args.get('start_date'))
    end = _parse_date(request.args.get('end_date'))
    operator = request.args.get('operator') or None

    start_str = start.strftime('%y%m%d') if start else ''
    end_str = end.strftime('%y%m%d') if end else ''

    payload = build_operator_report_payload(start, end, operator)
    payload['start'] = start.isoformat() if start else ''
    payload['end'] = end.isoformat() if end else ''

    combined, error = fetch_combined_reports()
    if error:
        current_app.logger.error("Combined report fetch failed: %s", error)

    charts = _generate_operator_report_charts(payload)
    payload.update(charts)

    body = request.get_json(silent=True) or {}

    def _get(name, default=''):
        return request.args.get(name, body.get(name, default))

    def _get_bool(name, default=True):
        value = request.args.get(name)
        if value is None:
            value = body.get(name)
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        return str(value).lower() not in {'0', 'false', 'no'}

    show_cover = _get_bool('show_cover')
    show_summary = _get_bool('show_summary')
    title = _get('title') or (operator or 'Operator Report')
    subtitle = _get('subtitle')
    report_date = _get('report_date')
    period = _get('period')
    author = _get('author')
    logo_url = _get('logo_url') or url_for(
        'static', filename='images/company-logo.png', _external=True
    )
    footer_left = _get('footer_left')
    report_id = _get('report_id')
    contact = _get('contact', 'tschwartz@4spectra.com')
    confidentiality = _get('confidentiality', 'Spectra-Tech • Confidential')
    generated_at = datetime.now(_report_timezone()).strftime('%Y-%m-%d %H:%M:%S %Z')

    if show_summary:
        payload.setdefault(
            'yieldSummary',
            {
                'avg': 0.0,
                'worstDay': {'date': None, 'yield': 0.0},
                'worstAssembly': {'assembly': None, 'yield': 0.0},
            },
        )
        payload.setdefault('operatorSummary', payload.get('summary', {}))
        payload.setdefault('modelSummary', {'avgFalseCalls': 0.0})

    report_css = _load_report_css()

    html = render_template(
        'report/operator/index.html',
        show_cover=show_cover,
        show_summary=show_summary,
        title=title,
        subtitle=subtitle,
        report_date=report_date,
        period=period,
        author=author,
        logo_url=logo_url,
        footer_left=footer_left,
        report_id=report_id,
        contact=contact,
        confidentiality=confidentiality,
        generated_at=generated_at,
        operator=operator,
        report_css=report_css,
        **payload,
    )

    fmt = request.args.get('format')
    filename_stem = f"{start_str}_{end_str}_operator_report"
    if fmt == 'pdf':
        try:
            pdf = render_html_to_pdf(html, base_url=request.url_root)
        except PdfGenerationError as exc:
            return jsonify({'message': str(exc)}), 503
        return send_file(
            io.BytesIO(pdf),
            mimetype='application/pdf',
            download_name=f"{filename_stem}.pdf",
            as_attachment=True,
        )
    if fmt == 'html':
        return send_file(
            io.BytesIO(html.encode('utf-8')),
            mimetype='text/html',
            download_name=f"{filename_stem}.html",
            as_attachment=True,
        )
    return html


@main_bp.route('/analysis/aoi/grades', methods=['GET'])
@feature_required('analysis_aoi_grades')
def aoi_grades():
    """Return AOI grades computed from combined reports.

    Optional query parameters:
        - start_date
        - end_date
        - operators (comma-separated)
        - job_numbers (comma-separated)
    """
    if 'username' not in session:
        return redirect(url_for('auth.login'))

    start = request.args.get('start_date')
    end = request.args.get('end_date')
    operators = request.args.get('operators', '')
    job_numbers = request.args.get('job_numbers', '')

    data, error = fetch_combined_reports()
    if error:
        abort(500, description=error)

    def to_set(values):
        return {v.strip() for v in values.split(',') if v.strip()}

    start_dt = _parse_date(start)
    end_dt = _parse_date(end)
    operator_set = to_set(operators)
    job_set = to_set(job_numbers)

    filtered = []
    for row in data:
        date_val = row.get('aoi_Date') or row.get('Date') or row.get('date')
        dt = _parse_date(date_val)
        if start_dt and (not dt or dt < start_dt):
            continue
        if end_dt and (not dt or dt > end_dt):
            continue
        operator = row.get('aoi_Operator') or row.get('Operator')
        if operator_set and (operator not in operator_set):
            continue
        job_number = row.get('aoi_Job Number') or row.get('Job Number')
        if job_set and (job_number not in job_set):
            continue
        info = row.get('fi_Additional Information') or ""
        phrases = current_app.config.get("NON_AOI_PHRASES", [])
        row['fi_Quantity Rejected'] = parse_fi_rejections(info, phrases)
        filtered.append(row)

    grades = calculate_aoi_grades(filtered)
    return jsonify(grades)


@main_bp.route('/analysis/aoi/grades/escape_pareto', methods=['GET'])
@feature_required('analysis_aoi_grades')
def aoi_grades_escape_pareto():
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    group = request.args.get('group', 'model')  # 'model'|'operator'|'shift'
    start = _parse_date(request.args.get('start_date'))
    end = _parse_date(request.args.get('end_date'))
    data, error = fetch_combined_reports()
    if error:
        abort(500, description=error)

    key_map = {
        'model': lambda r: r.get('aoi_Assembly') or r.get('Model') or r.get('Assembly') or 'Unknown',
        'operator': lambda r: r.get('aoi_Operator') or r.get('Operator') or 'Unknown',
        'shift': lambda r: r.get('aoi_Shift') or r.get('Shift') or 'Unknown',
    }
    key_fn = key_map.get(group, key_map['model'])

    agg = defaultdict(lambda: {'fi_rej': 0.0, 'aoi_passed': 0.0})
    total_rej = 0.0

    for row in data:
        dt = _parse_date(row.get('aoi_Date') or row.get('Date') or row.get('date'))
        if start and (not dt or dt < start):
            continue
        if end and (not dt or dt > end):
            continue
        key = key_fn(row)
        passed = _aoi_passed(row)
        fi_rej = _fi_rejected(row)
        agg[key]['fi_rej'] += fi_rej
        agg[key]['aoi_passed'] += passed
        total_rej += fi_rej

    items = []
    for k, v in agg.items():
        denom = v['aoi_passed']
        rate = (1000.0 * v['fi_rej'] / denom) if denom else 0.0
        items.append({'key': k, 'fi_rej': v['fi_rej'], 'escape_rate_per_1k': rate})
    items.sort(key=lambda x: x['fi_rej'], reverse=True)

    cumulative = 0.0
    out = []
    for it in items:
        share = (it['fi_rej'] / total_rej) if total_rej else 0.0
        cumulative += share
        out.append({**it, 'cum_share': cumulative})
    return jsonify({'group': group, 'items': out, 'total_fi_rejects': total_rej})


@main_bp.route('/analysis/aoi/grades/gap_risk', methods=['GET'])
@feature_required('analysis_aoi_grades')
def aoi_grades_gap_risk():
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    start = _parse_date(request.args.get('start_date'))
    end = _parse_date(request.args.get('end_date'))
    data, error = fetch_combined_reports()
    if error:
        abort(500, description=error)

    buckets = [
        (lambda d: d is not None and d <= 1, '≤1d'),
        (lambda d: d is not None and 2 <= d <= 3, '2–3d'),
        (lambda d: d is not None and 4 <= d <= 7, '4–7d'),
        (lambda d: d is None or d > 7, '>7d'),
    ]
    hist = Counter()
    fi_by_bucket = Counter()
    total_fi = 0.0
    phrases = current_app.config.get("NON_AOI_PHRASES", [])

    for row in data:
        dt = _parse_date(row.get('aoi_Date') or row.get('Date') or row.get('date'))
        if start and (not dt or dt < start):
            continue
        if end and (not dt or dt > end):
            continue
        gd = _gap_days(row)
        label = None
        for pred, name in buckets:
            if pred(gd):
                label = name
                break
        hist[label] += 1
        info = row.get('fi_Additional Information') or ""
        rej = parse_fi_rejections(info, phrases)
        fi_by_bucket[label] += rej
        total_fi += rej

    labels = ['≤1d', '2–3d', '4–7d', '>7d']
    return jsonify({
        'labels': labels,
        'histogram': [hist.get(l, 0) for l in labels],
        'fi_share': [ (fi_by_bucket.get(l, 0.0) / total_fi) if total_fi else 0.0 for l in labels ],
    })


@main_bp.route('/analysis/aoi/grades/learning_curves', methods=['GET'])
@feature_required('analysis_aoi_grades')
def aoi_grades_learning_curves():
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    window = int(request.args.get('window', 10))
    op_filter = request.args.get('operators')
    start = _parse_date(request.args.get('start_date'))
    end = _parse_date(request.args.get('end_date'))
    data, error = fetch_combined_reports()
    if error:
        abort(500, description=error)

    per_op = defaultdict(list)
    job_totals = defaultdict(float)
    for row in data:
        dt = _parse_date(row.get('aoi_Date') or row.get('Date') or row.get('date'))
        if start and (not dt or dt < start):
            continue
        if end and (not dt or dt > end):
            continue
        job = row.get('aoi_Job Number') or row.get('Job Number') or 'Unknown'
        job_totals[job] += _aoi_passed(row)

    job_fi_rej = defaultdict(float)
    for row in data:
        dt = _parse_date(row.get('aoi_Date') or row.get('Date') or row.get('date'))
        if start and (not dt or dt < start):
            continue
        if end and (not dt or dt > end):
            continue
        job = row.get('aoi_Job Number') or row.get('Job Number') or 'Unknown'
        info = row.get('fi_Additional Information') or ""
        phrases = current_app.config.get("NON_AOI_PHRASES", [])
        rej = parse_fi_rejections(info, phrases)
        job_fi_rej[job] = max(job_fi_rej[job], rej)

    allowed = None
    if op_filter:
        allowed = {s.strip() for s in op_filter.split(',') if s.strip()}

    for row in data:
        dt = _parse_date(row.get('aoi_Date') or row.get('Date') or row.get('date'))
        if start and (not dt or dt < start):
            continue
        if end and (not dt or dt > end):
            continue
        op = row.get('aoi_Operator') or row.get('Operator') or 'Unknown'
        if allowed and op not in allowed:
            continue
        job = row.get('aoi_Job Number') or row.get('Job Number') or 'Unknown'
        passed = _aoi_passed(row)
        total = job_totals.get(job, 0.0)
        share = (passed / total) if total else 0.0
        attr_missed = share * job_fi_rej.get(job, 0.0)
        rate = (1000.0 * attr_missed / passed) if passed else 0.0
        if dt:
            per_op[op].append((dt, rate))

    out = {}
    for op, seq in per_op.items():
        seq.sort(key=lambda x: x[0])
        vals = [r for _, r in seq]
        dates = [d.isoformat() for d, _ in seq]
        roll = []
        for i in range(len(vals)):
            start_i = max(0, i - window + 1)
            sub = sorted(vals[start_i : i + 1])
            m = sub[len(sub)//2] if sub else 0.0
            roll.append(m)
        out[op] = { 'dates': dates, 'rates': vals, 'rolling_median': roll }

    return jsonify(out)


@main_bp.route('/analysis/aoi/grades/smt_th_heatmap', methods=['GET'])
@feature_required('analysis_aoi_grades')
def aoi_grades_smt_th_heatmap():
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    start = _parse_date(request.args.get('start_date'))
    end = _parse_date(request.args.get('end_date'))
    data, error = fetch_combined_reports()
    if error:
        abort(500, description=error)

    # Buckets
    stations = set()
    parts = set()
    agg = defaultdict(lambda: {'fi_rej': 0.0, 'passed': 0.0})
    for row in data:
        dt = _parse_date(row.get('aoi_Date') or row.get('Date') or row.get('date'))
        if start and (not dt or dt < start):
            continue
        if end and (not dt or dt > end):
            continue
        station = row.get('aoi_Station') or row.get('Station') or 'Unknown'
        part = row.get('fi_Part Type') or row.get('fi_part_type') or 'Unknown'
        stations.add(station)
        parts.add(part)
        key = (station, part)
        agg[key]['fi_rej'] += _fi_rejected(row)
        agg[key]['passed'] += _aoi_passed(row)

    stations = sorted(stations)
    parts = sorted(parts)
    matrix = []
    for s in stations:
        row_vals = []
        for p in parts:
            v = agg.get((s, p), {'fi_rej': 0.0, 'passed': 0.0})
            rate = (1000.0 * v['fi_rej'] / v['passed']) if v['passed'] else 0.0
            row_vals.append(rate)
        matrix.append(row_vals)
    return jsonify({'stations': stations, 'part_types': parts, 'matrix': matrix})


@main_bp.route('/analysis/aoi/grades/shift_effect', methods=['GET'])
@feature_required('analysis_aoi_grades')
def aoi_grades_shift_effect():
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    start = _parse_date(request.args.get('start_date'))
    end = _parse_date(request.args.get('end_date'))
    data, error = fetch_combined_reports()
    if error:
        abort(500, description=error)

    # Build per-record escape rate per 1k
    per_shift = defaultdict(list)
    per_weekday_shift = defaultdict(lambda: defaultdict(list))
    phrases = current_app.config.get("NON_AOI_PHRASES", [])
    for row in data:
        dt = _parse_date(row.get('aoi_Date') or row.get('Date') or row.get('date'))
        if start and (not dt or dt < start):
            continue
        if end and (not dt or dt > end):
            continue
        shift = row.get('aoi_Shift') or row.get('Shift') or 'Unknown'
        passed = _aoi_passed(row)
        info = row.get('fi_Additional Information') or ""
        rej = parse_fi_rejections(info, phrases)
        rate = (1000.0 * rej / passed) if passed else 0.0
        per_shift[shift].append(rate)
        if dt:
            wd = dt.weekday()  # 0=Mon
            per_weekday_shift[wd][shift].append(rate)

    # Summaries
    def _summary(xs):
        xs = sorted(xs)
        n = len(xs)
        if n == 0:
            return { 'n': 0, 'mean': 0, 'q1': 0, 'median': 0, 'q3': 0 }
        def q(p):
            k = int(p*(n-1))
            return xs[k]
        return {
            'n': n,
            'mean': sum(xs)/n,
            'q1': q(0.25),
            'median': q(0.5),
            'q3': q(0.75),
        }

    shift_labels = sorted(per_shift.keys())
    shift_stats = { s: _summary(per_shift[s]) for s in shift_labels }

    weekdays = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun']
    heat = []
    all_shifts = sorted({s for d in per_weekday_shift.values() for s in d.keys()})
    for i in range(7):
        row_vals = []
        for s in all_shifts:
            xs = per_weekday_shift[i][s]
            v = (sum(xs)/len(xs)) if xs else 0.0
            row_vals.append(v)
        heat.append(row_vals)

    return jsonify({
        'shifts': shift_labels,
        'shift_stats': shift_stats,
        'weekday_labels': weekdays,
        'weekday_shifts': all_shifts,
        'weekday_heat': heat,
    })


@main_bp.route('/analysis/aoi/grades/customer_yield', methods=['GET'])
@feature_required('analysis_aoi_grades')
def aoi_grades_customer_yield():
    """Return per-customer true yield using AOI and FI rejects.

    True yield per customer = (AOI inspected - AOI rejected - FI rejected) / AOI inspected.
    """
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    start = _parse_date(request.args.get('start_date'))
    end = _parse_date(request.args.get('end_date'))
    data, error = fetch_combined_reports()
    if error:
        abort(500, description=error)

    from collections import defaultdict

    agg = defaultdict(lambda: {'inspected': 0.0, 'aoi_rej': 0.0, 'fi_rej': 0.0})
    label_map = {}
    for row in data:
        dt = _parse_date(row.get('aoi_Date') or row.get('Date') or row.get('date'))
        if start and (not dt or dt < start):
            continue
        if end and (not dt or dt > end):
            continue
        raw = (row.get('aoi_Customer') or row.get('Customer') or 'Unknown').strip()
        norm = raw.lower()
        label_map.setdefault(norm, raw)
        aoi_ins = float(row.get('aoi_Quantity Inspected') or row.get('Quantity Inspected') or 0)
        aoi_rej = float(row.get('aoi_Quantity Rejected') or row.get('Quantity Rejected') or 0)
        fi_rej = float(row.get('fi_Quantity Rejected') or 0)
        agg[norm]['inspected'] += aoi_ins
        agg[norm]['aoi_rej'] += aoi_rej
        agg[norm]['fi_rej'] += fi_rej

    items = []
    for norm, vals in agg.items():
        ins = vals['inspected']
        true_accepted = max(0.0, ins - vals['aoi_rej'] - vals['fi_rej'])
        yld = (true_accepted / ins * 100.0) if ins else 0.0
        items.append((label_map[norm], yld))

    # Sort by yield descending for readability
    items.sort(key=lambda x: x[1], reverse=True)
    return jsonify({
        'labels': [i[0] for i in items],
        'yields': [i[1] for i in items],
    })


@main_bp.route('/analysis/aoi/grades/program_trend', methods=['GET'])
@feature_required('analysis_aoi_grades')
def aoi_grades_program_trend():
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    start = _parse_date(request.args.get('start_date'))
    end = _parse_date(request.args.get('end_date'))
    data, error = fetch_combined_reports()
    if error:
        abort(500, description=error)

    # Aggregate by model/rev and calendar month
    from collections import defaultdict
    agg = defaultdict(lambda: defaultdict(lambda: {'fi': 0.0, 'passed': 0.0}))
    phrases = current_app.config.get("NON_AOI_PHRASES", [])
    for row in data:
        dt = _parse_date(row.get('aoi_Date') or row.get('Date') or row.get('date'))
        if start and (not dt or dt < start):
            continue
        if end and (not dt or dt > end):
            continue
        model = row.get('aoi_Assembly') or row.get('Assembly') or 'Unknown'
        rev = row.get('aoi_Rev') or row.get('Rev') or ''
        key = f"{model} {rev}".strip()
        month = dt.replace(day=1).isoformat() if dt else 'Unknown'
        info = row.get('fi_Additional Information') or ""
        rej = parse_fi_rejections(info, phrases)
        agg[key][month]['fi'] += rej
        agg[key][month]['passed'] += _aoi_passed(row)

    # Build aligned series per key
    months = sorted({m for d in agg.values() for m in d.keys() if m != 'Unknown'})
    datasets = []
    for key, m in agg.items():
        data_points = []
        for mon in months:
            v = m.get(mon, {'fi': 0.0, 'passed': 0.0})
            rate = (1000.0 * v['fi'] / v['passed']) if v['passed'] else 0.0
            data_points.append(rate)
        datasets.append({'label': key, 'data': data_points})
    return jsonify({'months': months, 'datasets': datasets})


@main_bp.route('/analysis/aoi/grades/adjusted_operator_ranking', methods=['GET'])
@feature_required('analysis_aoi_grades')
def aoi_grades_adjusted_operator_ranking():
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    import numpy as np
    start = _parse_date(request.args.get('start_date'))
    end = _parse_date(request.args.get('end_date'))
    data, error = fetch_combined_reports()
    if error:
        abort(500, description=error)

    rows = []
    phrases = current_app.config.get("NON_AOI_PHRASES", [])
    for row in data:
        dt = _parse_date(row.get('aoi_Date') or row.get('Date') or row.get('date'))
        if start and (not dt or dt < start):
            continue
        if end and (not dt or dt > end):
            continue
        op = row.get('aoi_Operator') or row.get('Operator') or 'Unknown'
        model = row.get('aoi_Assembly') or row.get('Assembly') or 'Unknown'
        shift = row.get('aoi_Shift') or row.get('Shift') or 'Unknown'
        passed = _aoi_passed(row)
        info = row.get('fi_Additional Information') or ""
        rej = parse_fi_rejections(info, phrases)
        y = (1000.0 * rej / passed) if passed else 0.0
        rows.append((op, model, shift, passed, y))

    if not rows:
        return jsonify({'operators': [], 'effects': []})

    # Build design matrix: intercept + model dummies + shift dummies + log(volume)
    ops = sorted({r[0] for r in rows})
    models = sorted({r[1] for r in rows})
    shifts = sorted({r[2] for r in rows})
    op_index = {k: i for i, k in enumerate(ops)}
    model_index = {k: i for i, k in enumerate(models)}
    shift_index = {k: i for i, k in enumerate(shifts)}

    n = len(rows)
    p = 1 + (len(models)-1) + (len(shifts)-1) + 1 + len(ops)  # intercept + effects + log(vol) + operator effects
    X = np.zeros((n, p))
    y = np.zeros(n)

    # Column layout
    col = 0
    intercept_col = col; col += 1
    model_cols = {m: intercept_col + 1 + i for i, m in enumerate(models[1:])}
    col = intercept_col + 1 + max(0, len(models)-1)
    shift_cols = {s: col + i for i, s in enumerate(shifts[1:])}
    col += max(0, len(shifts)-1)
    logv_col = col; col += 1
    op_cols = {o: col + i for i, o in enumerate(ops)}

    for i, (op, model, shift, passed, yi) in enumerate(rows):
        X[i, intercept_col] = 1.0
        if model in model_cols:
            X[i, model_cols[model]] = 1.0
        if shift in shift_cols:
            X[i, shift_cols[shift]] = 1.0
        X[i, logv_col] = np.log(max(passed, 1.0))
        X[i, op_cols[op]] = 1.0
        y[i] = yi

    # Ridge regularization for stability
    lam = 1.0
    XtX = X.T @ X + lam * np.eye(X.shape[1])
    Xty = X.T @ y
    beta = np.linalg.solve(XtX, Xty)

    effects = []
    for op in ops:
        eff = float(beta[op_cols[op]])
        # naive CI based on residual variance and count per operator
        idx = [i for i, r in enumerate(rows) if r[0] == op]
        resid = y[idx] - X[idx] @ beta
        var = float((resid @ resid) / max(1, len(idx)-1))
        se = (var ** 0.5) / (len(idx) ** 0.5)
        effects.append({'operator': op, 'effect': eff, 'lower': eff - 1.96*se, 'upper': eff + 1.96*se, 'n': len(idx)})

    # Sort best (lower effect is better) ascending
    effects.sort(key=lambda d: d['effect'])
    return jsonify({'operators': ops, 'effects': effects})


@main_bp.route('/analysis/aoi/grades/view', methods=['GET'])
@feature_required('analysis_aoi_grades')
def aoi_grades_page():
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    return render_template('aoi_grades.html', username=session.get('username'))


@main_bp.route('/analysis/tracker-logs', methods=['GET'])
@admin_required
def analysis_tracker_logs():
    tracker = _get_tracker()
    args = request.args

    if args.get('reset'):
        return redirect(url_for('main.analysis_tracker_logs'))

    requested_tab = (args.get('tab') or '').strip().lower()
    active_tab = 'bug-reports' if requested_tab == 'bug-reports' else 'analytics'

    local_zone = _tracker_local_zone()

    def _parse_date_arg(value, *, clamp_end: bool = False):
        if not value:
            return None
        cleaned = str(value).strip()
        try:
            if len(cleaned) == 10:
                dt = datetime.fromisoformat(cleaned)
                if clamp_end:
                    dt = dt.replace(hour=23, minute=59, second=59)
            else:
                dt = datetime.fromisoformat(cleaned)
        except ValueError:
            return None
        return dt.replace(tzinfo=timezone.utc) if dt.tzinfo is None else dt

    def _loads(payload):
        if not payload:
            return None
        try:
            return json.loads(payload)
        except Exception:
            return None

    limit = args.get('limit', type=int) or 50
    limit = max(10, min(limit, 250))

    role_filter = (args.get('role') or '').strip().upper() or None
    event_filter = (args.get('event') or '').strip() or None
    event_filter_lower = event_filter.lower() if event_filter else None
    search_filter = (args.get('session') or '').strip()
    backtracking_filter = (args.get('backtracking') or '').strip().lower() or 'any'
    start_filter_input = (args.get('start') or '').strip()
    end_filter_input = (args.get('end') or '').strip()

    start_dt = _parse_date_arg(start_filter_input, clamp_end=False)
    end_dt = _parse_date_arg(end_filter_input, clamp_end=True)

    where_clauses: list[str] = []
    params: list[str] = []
    if role_filter:
        where_clauses.append('user_role = ?')
        params.append(role_filter)
    if search_filter:
        like = f"%{search_filter.lower()}%"
        where_clauses.append(
            '('
            'LOWER(COALESCE(username, "")) LIKE ? OR '
            'LOWER(COALESCE(user_id, "")) LIKE ? OR '
            'LOWER(session_token) LIKE ?'
            ')'
        )
        params.extend([like, like, like])
    if start_dt:
        where_clauses.append('datetime(start_time) >= datetime(?)')
        params.append(start_dt.astimezone(timezone.utc).isoformat())
    if end_dt:
        where_clauses.append('datetime(COALESCE(end_time, start_time)) <= datetime(?)')
        params.append(end_dt.astimezone(timezone.utc).isoformat())

    query = (
        'SELECT id, session_token, user_id, username, user_role, '
        'start_time, end_time, duration_seconds '
        'FROM sessions'
    )
    if where_clauses:
        query += ' WHERE ' + ' AND '.join(where_clauses)
    query += ' ORDER BY datetime(start_time) DESC LIMIT ?'
    params.append(limit)

    with tracker._connect() as conn:
        session_rows = conn.execute(query, params).fetchall()

        tokens = [row['session_token'] for row in session_rows]
        event_rows = []
        if tokens:
            placeholders = ','.join('?' for _ in tokens)
            event_query = (
                'SELECT id, session_token, user_id, user_role, event_name, '
                'context, metadata, occurred_at '
                f'FROM click_events WHERE session_token IN ({placeholders}) '
                'ORDER BY datetime(occurred_at) ASC, id ASC'
            )
            event_rows = conn.execute(event_query, tokens).fetchall()

    role_options = sorted(
        {row['user_role'] for row in session_rows if row['user_role']}
    )
    event_options = sorted({row['event_name'] for row in event_rows})

    events_by_session: dict[str, list[dict]] = {token: [] for token in (tokens or [])}
    for row in event_rows:
        occurred_dt = _tracker_parse_timestamp(row['occurred_at'])
        context_payload = _loads(row['context'])
        metadata_payload = _loads(row['metadata'])
        label = None
        href = None
        if isinstance(context_payload, dict):
            label = (context_payload.get('text') or '').strip() or None
            href = context_payload.get('href')
        events_by_session.setdefault(row['session_token'], []).append(
            {
                'id': row['id'],
                'name': row['event_name'],
                'occurred': occurred_dt,
                'occurred_display': _tracker_format_timestamp(occurred_dt, local_zone)
                or row['occurred_at'],
                'context': context_payload,
                'metadata': metadata_payload,
                'label': label,
                'href': href,
                'is_backtrack': False,
            }
        )

    session_details: list[dict] = []
    all_events: list[dict] = []
    for row in session_rows:
        token = row['session_token']
        start_ts = _tracker_parse_timestamp(row['start_time'])
        end_ts = _tracker_parse_timestamp(row['end_time'])
        events = events_by_session.get(token, [])

        derived_end_ts, derived_end_display = _derive_session_end(
            end_ts, events, local_zone=local_zone
        )

        navigation_events = [ev for ev in events if ev['name'].lower() == 'navigate']
        seen_hrefs: set[str] = set()
        backtracking_events: list[dict] = []
        for event in navigation_events:
            href = event.get('href')
            if href:
                if href in seen_hrefs:
                    event['is_backtrack'] = True
                    backtracking_events.append(event)
                else:
                    seen_hrefs.add(href)

        duration_seconds = _calculate_session_duration(
            start_ts,
            derived_end_ts,
            events,
            row['duration_seconds'],
        )

        duration_label = _tracker_format_duration(duration_seconds)
        path = [
            {
                'href': event.get('href'),
                'label': event.get('label'),
                'occurred': event['occurred_display'],
                'is_backtrack': event['is_backtrack'],
            }
            for event in navigation_events
        ]

        last_event_display = events[-1]['occurred_display'] if events else None
        detail = {
            'token': token,
            'user_id': row['user_id'],
            'username': row['username'],
            'role': row['user_role'],
            'start_display': _tracker_format_timestamp(start_ts, local_zone),
            'end_display': derived_end_display,
            'duration_seconds': duration_seconds,
            'duration_label': duration_label,
            'event_count': len(events),
            'navigation_count': len(navigation_events),
            'backtracking_count': len(backtracking_events),
            'has_backtracking': bool(backtracking_events),
            'backtracking_events': [
                {
                    'href': event.get('href'),
                    'label': event.get('label'),
                    'occurred': event['occurred_display'],
                }
                for event in backtracking_events
            ],
            'path': path,
            'last_event_display': last_event_display,
            'events': events,
        }
        session_details.append(detail)

        for event in events:
            if event_filter_lower and event['name'].lower() != event_filter_lower:
                continue
            flattened = {
                'session_token': token,
                'user': row['username'] or row['user_id'] or 'Unknown',
                'role': row['user_role'],
                'event': event['name'],
                'occurred': event['occurred_display'],
                'href': event.get('href'),
                'label': event.get('label'),
                'is_backtrack': event['is_backtrack'],
                'context': event['context'],
                'metadata': event['metadata'],
            }
            all_events.append(flattened)

    if backtracking_filter == 'only':
        session_details = [s for s in session_details if s['has_backtracking']]
    elif backtracking_filter in {'none', 'exclude'}:
        session_details = [s for s in session_details if not s['has_backtracking']]

    included_tokens = {detail['token'] for detail in session_details}
    if included_tokens:
        all_events = [
            event for event in all_events if event['session_token'] in included_tokens
        ]
    else:
        all_events = []

    max_events = 400
    if len(all_events) > max_events:
        all_events = all_events[-max_events:]

    total_events = sum(detail['event_count'] for detail in session_details)
    total_navigation = sum(detail['navigation_count'] for detail in session_details)
    total_backtracking = sum(detail['backtracking_count'] for detail in session_details)
    duration_values = [
        detail['duration_seconds']
        for detail in session_details
        if detail['duration_seconds'] is not None
    ]
    avg_duration = mean(duration_values) if duration_values else None

    stats = {
        'total_sessions': len(session_details),
        'total_events': total_events,
        'total_navigation': total_navigation,
        'total_backtracking': total_backtracking,
        'avg_duration': _tracker_format_duration(avg_duration)
        if avg_duration is not None
        else '--',
    }

    chart_data = {
        'labels': [
            (detail['username'] or detail['user_id'] or detail['token'])
            for detail in session_details
        ],
        'durations': [
            round(detail['duration_seconds'], 2) if detail['duration_seconds'] else 0
            for detail in session_details
        ],
        'backtracking': [detail['backtracking_count'] for detail in session_details],
        'navigation': [detail['navigation_count'] for detail in session_details],
    }

    event_counter = Counter(event['event'] for event in all_events) if all_events else Counter()
    event_summary_data = {
        'labels': [item[0] for item in event_counter.most_common()],
        'counts': [item[1] for item in event_counter.most_common()],
    }

    role_counter = Counter(detail['role'] or 'Unknown' for detail in session_details)
    role_breakdown_data = {
        'labels': [item[0] for item in role_counter.most_common()],
        'counts': [item[1] for item in role_counter.most_common()],
    }

    filters = {
        'role': role_filter,
        'event': event_filter,
        'session': search_filter,
        'backtracking': backtracking_filter,
        'start': start_filter_input,
        'end': end_filter_input,
        'limit': limit,
    }

    backtracking_sessions = [s for s in session_details if s['has_backtracking']]

    bug_reports_raw: list[dict] = []
    bug_reports_error: str | None = None
    try:
        bug_reports_raw, bug_reports_error = fetch_bug_reports()
    except Exception as exc:  # pragma: no cover - defensive safeguard
        bug_reports_raw = []
        bug_reports_error = str(exc)

    assignable_users: list[dict] = []
    assignee_error: str | None = None
    try:
        fetched_users, assignee_error = _fetch_configured_users()
    except Exception as exc:  # pragma: no cover - defensive safeguard
        fetched_users = []
        assignee_error = str(exc)

    for user in fetched_users or []:
        user_id = user.get('id')
        if user_id is None:
            continue
        assignable_users.append(
            {
                'id': str(user_id),
                'label': user.get('display_name') or user.get('username') or str(user_id),
                'role': user.get('role'),
            }
        )

    assignable_lookup: dict[str, dict] = {}
    for user in assignable_users:
        user_id = user.get('id')
        if user_id is None:
            continue
        assignable_lookup[str(user_id)] = user

    status_counter: Counter[str] = Counter()
    formatted_bug_reports: list[dict] = []
    for record in bug_reports_raw or []:
        status_value = (record.get('status') or 'open').strip().lower()
        status_counter[status_value] += 1

        created_at = _tracker_parse_timestamp(record.get('created_at'))
        updated_at = _tracker_parse_timestamp(record.get('updated_at'))

        assignee_id = record.get('assignee_id')
        assignee_id_str = str(assignee_id) if assignee_id not in (None, '') else ''
        assignee_label = record.get('assignee_name') or assignable_lookup.get(assignee_id_str, {}).get('label')

        reporter_id = record.get('reporter_id')
        reporter_label = _resolve_user_display_name(
            reporter_id,
            assignable_lookup,
            record.get('reporter_name'),
        )

        formatted_bug_reports.append(
            {
                'id': record.get('id'),
                'title': record.get('title'),
                'description': record.get('description'),
                'priority': record.get('priority') or 'Unspecified',
                'status': status_value,
                'status_label': status_value.replace('_', ' ').title(),
                'reporter': reporter_label,
                'reporter_display_name': reporter_label,
                'reporter_id': reporter_id,
                'assignee_id': assignee_id_str,
                'assignee_label': assignee_label or 'Unassigned',
                'assignee_token': assignee_id_str or 'unassigned',
                'notes': record.get('notes') or '',
                'created_at': created_at,
                'created_display': _tracker_format_timestamp(created_at, local_zone)
                or '—',
                'updated_at': updated_at,
                'updated_display': _tracker_format_timestamp(updated_at, local_zone)
                or '—',
                'raw': record,
            }
        )

    formatted_bug_reports.sort(key=lambda item: item['created_at'] or datetime.min.replace(tzinfo=timezone.utc), reverse=True)

    recent_bug_reports = formatted_bug_reports[:5]

    known_statuses = ['open', 'in_progress', 'resolved', 'on_hold']
    status_counts = [
        {
            'value': value,
            'label': value.replace('_', ' ').title(),
            'count': status_counter.get(value, 0),
        }
        for value in known_statuses
    ]

    for value, count in status_counter.items():
        if value in known_statuses:
            continue
        status_counts.append(
            {
                'value': value,
                'label': value.replace('_', ' ').title(),
                'count': count,
            }
        )

    total_bug_reports = sum(entry['count'] for entry in status_counts)

    bug_update_base = url_for('main.update_bug_report', report_id=0).rsplit('/', 1)[0]

    return render_template(
        'analysis_tracker_logs.html',
        sessions=session_details,
        events=all_events,
        stats=stats,
        filters=filters,
        role_options=role_options,
        event_options=event_options,
        chart_data=chart_data,
        event_summary_data=event_summary_data,
        role_breakdown_data=role_breakdown_data,
        backtracking_sessions=backtracking_sessions,
        bug_reports=formatted_bug_reports,
        bug_status_counts=status_counts,
        bug_total=total_bug_reports,
        recent_bug_reports=recent_bug_reports,
        bug_reports_error=bug_reports_error,
        bug_assignee_error=assignee_error,
        bug_assignees=assignable_users,
        bug_update_base=bug_update_base,
        bug_status_options=[
            {'value': value, 'label': value.replace('_', ' ').title()}
            for value in known_statuses
        ],
        active_tab=active_tab,
    )


@main_bp.route('/analysis/aoi', methods=['GET'])
@feature_required('analysis_aoi_daily')
def aoi_daily_reports():
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    return render_template(
        'aoi_daily_reports.html',
        username=session.get('username'),
        user_role=(session.get('role') or '').upper(),
    )


def _daily_data(fetch_func):
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    start = request.args.get('start_date')
    end = request.args.get('end_date')
    job_numbers = request.args.get('job_numbers', '')
    rev_numbers = request.args.get('rev_numbers', '')
    assemblies = request.args.get('assemblies', '')
    customers = request.args.get('customers', '')
    operators = request.args.get('operators', '')

    data, error = fetch_func()
    if error:
        abort(500, description=error)

    from datetime import datetime
    from collections import defaultdict

    def parse_date(d):
        if not d:
            return None
        try:
            return datetime.fromisoformat(str(d)).date()
        except Exception:
            return None

    def to_list(s):
        return [x.strip() for x in s.split(',') if x.strip()]

    start_dt = parse_date(start)
    end_dt = parse_date(end)
    job_numbers = set(to_list(job_numbers))
    rev_numbers = set(to_list(rev_numbers))
    assemblies = set(to_list(assemblies))
    customers = set(x.lower() for x in to_list(customers))
    operators = set(to_list(operators))

    filtered = []
    for row in data:
        date = parse_date(row.get('Date') or row.get('date'))
        if start_dt and (not date or date < start_dt):
            continue
        if end_dt and (not date or date > end_dt):
            continue
        if job_numbers and (row.get('Job Number') not in job_numbers):
            continue
        if rev_numbers and (row.get('Rev') not in rev_numbers):
            continue
        if assemblies and (row.get('Assembly') not in assemblies):
            continue
        if customers and ((row.get('Customer') or '').lower() not in customers):
            continue
        if operators and (row.get('Operator') not in operators):
            continue
        filtered.append(row)

    view = request.args.get('view')

    if view == 'shift':
        agg = defaultdict(lambda: {'1st': {'accepted': 0, 'rejected': 0}, '2nd': {'accepted': 0, 'rejected': 0}})
        totals = {'1st': {'accepted': 0, 'rejected': 0}, '2nd': {'accepted': 0, 'rejected': 0}}
        for row in filtered:
            date = parse_date(row.get('Date') or row.get('date'))
            shift_raw = str(row.get('Shift') or row.get('shift') or '').lower()
            if shift_raw in ('1', '1st', 'first', 'shift 1', 'shift1', '1st shift'):
                shift = '1st'
            elif shift_raw in ('2', '2nd', 'second', 'shift 2', 'shift2', '2nd shift'):
                shift = '2nd'
            else:
                continue
            inspected = int(row.get('Quantity Inspected') or row.get('quantity_inspected') or 0)
            rejected = int(row.get('Quantity Rejected') or row.get('quantity_rejected') or 0)
            accepted = inspected - rejected
            if accepted < 0:
                accepted = 0
            agg[date][shift]['accepted'] += accepted
            agg[date][shift]['rejected'] += rejected
            totals[shift]['accepted'] += accepted
            totals[shift]['rejected'] += rejected

        dates = sorted(agg.keys())
        s1_acc = [agg[d]['1st']['accepted'] for d in dates]
        s1_rej = [agg[d]['1st']['rejected'] for d in dates]
        s2_acc = [agg[d]['2nd']['accepted'] for d in dates]
        s2_rej = [agg[d]['2nd']['rejected'] for d in dates]

        def avg_rate(tot):
            total = tot['accepted'] + tot['rejected']
            return (tot['rejected'] / total * 100) if total else 0

        return jsonify({
            'labels': [d.isoformat() for d in dates],
            'shift1': {'accepted': s1_acc, 'rejected': s1_rej, 'avg_reject_rate': avg_rate(totals['1st'])},
            'shift2': {'accepted': s2_acc, 'rejected': s2_rej, 'avg_reject_rate': avg_rate(totals['2nd'])},
        })

    if view == 'yield':
        agg = defaultdict(lambda: {'accepted': 0, 'rejected': 0})
        for row in filtered:
            date = parse_date(row.get('Date') or row.get('date'))
            inspected = int(row.get('Quantity Inspected') or row.get('quantity_inspected') or 0)
            rejected = int(row.get('Quantity Rejected') or row.get('quantity_rejected') or 0)
            accepted = inspected - rejected
            if accepted < 0:
                accepted = 0
            agg[date]['accepted'] += accepted
            agg[date]['rejected'] += rejected

        dates = sorted(agg.keys())
        yields = []
        for d in dates:
            a = agg[d]['accepted']
            r = agg[d]['rejected']
            tot = a + r
            y = (a / tot * 100) if tot else 0
            yields.append(y)
        avg_yield = sum(yields) / len(yields) if yields else 0
        min_yield = min(yields) if yields else 0
        max_yield = max(yields) if yields else 0
        return jsonify({
            'labels': [d.isoformat() for d in dates],
            'yields': yields,
            'avg_yield': avg_yield,
            'min_yield': min_yield,
            'max_yield': max_yield,
        })

    if view == 'customer_rate':
        agg = defaultdict(lambda: {'accepted': 0, 'rejected': 0})
        label_map = {}
        for row in filtered:
            raw = (row.get('Customer') or 'Unknown').strip()
            norm = raw.lower()
            label_map.setdefault(norm, raw)
            inspected = int(row.get('Quantity Inspected') or row.get('quantity_inspected') or 0)
            rejected = int(row.get('Quantity Rejected') or row.get('quantity_rejected') or 0)
            accepted = inspected - rejected
            if accepted < 0:
                accepted = 0
            agg[norm]['accepted'] += accepted
            agg[norm]['rejected'] += rejected

        items = []
        for norm, vals in agg.items():
            tot = vals['accepted'] + vals['rejected']
            rate = (vals['rejected'] / tot * 100) if tot else 0
            items.append((label_map[norm], rate))
        items.sort(key=lambda x: x[1], reverse=True)
        labels = [i[0] for i in items]
        rates = [i[1] for i in items]
        avg_rate = sum(rates) / len(rates) if rates else 0
        max_rate = max(rates) if rates else 0
        min_rate = min(rates) if rates else 0
        max_customer = labels[rates.index(max_rate)] if rates else ''
        min_customer = labels[rates.index(min_rate)] if rates else ''
        return jsonify({
            'labels': labels,
            'rates': rates,
            'avg_rate': avg_rate,
            'max_rate': max_rate,
            'max_customer': max_customer,
            'min_rate': min_rate,
            'min_customer': min_customer,
        })

    if view == 'assembly':
        agg = defaultdict(lambda: {'inspected': 0, 'rejected': 0})
        for row in filtered:
            asm = row.get('Assembly') or 'Unknown'
            inspected = int(row.get('Quantity Inspected') or row.get('quantity_inspected') or 0)
            rejected = int(row.get('Quantity Rejected') or row.get('quantity_rejected') or 0)
            agg[asm]['inspected'] += inspected
            agg[asm]['rejected'] += rejected

        items = []
        for asm, vals in agg.items():
            ins = vals['inspected']
            rej = vals['rejected']
            yld = ((ins - rej) / ins * 100) if ins else 0
            items.append((asm, ins, rej, yld))
        items.sort(key=lambda x: x[0])
        return jsonify({
            'assemblies': [i[0] for i in items],
            'inspected': [i[1] for i in items],
            'rejected': [i[2] for i in items],
            'yields': [i[3] for i in items],
        })

    agg = defaultdict(lambda: {'accepted': 0, 'rejected': 0})
    for row in filtered:
        op = row.get('Operator') or 'Unknown'
        inspected = int(row.get('Quantity Inspected') or row.get('quantity_inspected') or 0)
        rejected = int(row.get('Quantity Rejected') or row.get('quantity_rejected') or 0)
        accepted = inspected - rejected
        if accepted < 0:
            accepted = 0
        agg[op]['accepted'] += accepted
        agg[op]['rejected'] += rejected

    items = sorted(agg.items(), key=lambda kv: kv[1]['accepted'] + kv[1]['rejected'], reverse=True)
    labels = [k for k, _ in items]
    accepted_vals = [v['accepted'] for _, v in items]
    rejected_vals = [v['rejected'] for _, v in items]

    return jsonify({'labels': labels, 'accepted': accepted_vals, 'rejected': rejected_vals})


@main_bp.route('/analysis/aoi/data', methods=['GET'])
@feature_required('analysis_aoi_daily')
def aoi_daily_data():
    return _daily_data(fetch_aoi_reports)


@main_bp.route('/analysis/fi', methods=['GET'])
@feature_required('analysis_fi_daily')
def fi_daily_reports():
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    return render_template(
        'fi_daily_reports.html',
        username=session.get('username'),
        user_role=(session.get('role') or '').upper(),
    )


@main_bp.route('/analysis/fi/data', methods=['GET'])
@feature_required('analysis_fi_daily')
def fi_daily_data():
    return _daily_data(fetch_fi_reports)


@main_bp.route('/analysis/fi/saved', methods=['GET', 'POST', 'PUT'])
@feature_required('analysis_fi_daily')
def fi_saved_queries():
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    if request.method == 'GET':
        data, error = fetch_saved_fi_queries()
        if error:
            abort(500, description=error)
        return jsonify(data)

    payload = request.get_json() or {}
    keys = ["name", "description", "start_date", "end_date", "params"]
    payload = {k: payload.get(k) for k in keys if k in payload}
    overwrite = request.method == 'PUT' or request.args.get('overwrite')
    if overwrite:
        name = payload.get('name')
        data, error = update_saved_fi_query(name, payload)
        status = 200
    else:
        data, error = insert_saved_fi_query(payload)
        status = 201
    if error:
        abort(500, description=error)
    return jsonify(data), status


@main_bp.route('/analysis/aoi/saved', methods=['GET', 'POST', 'PUT'])
@feature_required('analysis_aoi_daily')
def aoi_saved_queries():
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    if request.method == 'GET':
        data, error = fetch_saved_aoi_queries()
        if error:
            abort(500, description=error)
        return jsonify(data)

    payload = request.get_json() or {}
    keys = ["name", "description", "start_date", "end_date", "params"]
    payload = {k: payload.get(k) for k in keys if k in payload}
    overwrite = request.method == 'PUT' or request.args.get('overwrite')
    if overwrite:
        name = payload.get('name')
        data, error = update_saved_aoi_query(name, payload)
        status = 200
    else:
        data, error = insert_saved_aoi_query(payload)
        status = 201
    if error:
        abort(500, description=error)
    return jsonify(data), status


@main_bp.route('/api/tracking/session/start', methods=['POST'])
def tracking_session_start():
    if 'username' not in session:
        return jsonify({'ok': False, 'error': 'unauthenticated'}), 401

    tracker = _get_tracker()
    payload = request.get_json(silent=True) or {}

    requested_token = (
        payload.get('session_id')
        or payload.get('sessionId')
        or session.get('tracking_session_id')
    )
    timestamp = payload.get('timestamp')
    started_at = timestamp or datetime.now(tz=timezone.utc)

    token = tracker.start_session(
        session.get('user_id'),
        session.get('role'),
        username=session.get('username'),
        session_token=requested_token,
        started_at=started_at,
    )
    session['tracking_session_id'] = token

    return jsonify({'ok': True, 'session_id': token})


@main_bp.route('/api/tracking/session/end', methods=['POST'])
def tracking_session_end():
    if 'username' not in session:
        return jsonify({'ok': False, 'error': 'unauthenticated'}), 401

    tracker = _get_tracker()
    payload = request.get_json(silent=True) or {}
    requested_token = (
        payload.get('session_id')
        or payload.get('sessionId')
        or session.get('tracking_session_id')
    )
    timestamp = payload.get('timestamp')
    ended_at = timestamp or datetime.now(tz=timezone.utc)
    reason = payload.get('reason') or 'client'

    updated = tracker.end_session(requested_token, ended_at=ended_at)
    if updated:
        tracker.record_click(
            requested_token,
            session.get('user_id'),
            session.get('role'),
            'session_end',
            context={'reason': reason},
            occurred_at=ended_at,
        )

    if not session.get('tracking_session_id'):
        session['tracking_session_id'] = requested_token

    return jsonify({'ok': True, 'session_id': requested_token, 'closed': updated})


@main_bp.route('/api/tracking/click', methods=['POST'])
def tracking_click_event():
    if 'username' not in session:
        return jsonify({'ok': False, 'error': 'unauthenticated'}), 401

    tracker = _get_tracker()
    payload = request.get_json(silent=True) or {}

    event_name = payload.get('event') or payload.get('event_name')
    if not event_name:
        return (
            jsonify({'ok': False, 'error': 'event name required'}),
            400,
        )

    requested_token = (
        payload.get('session_id')
        or payload.get('sessionId')
        or session.get('tracking_session_id')
    )

    timestamp = payload.get('timestamp') or datetime.now(tz=timezone.utc)
    context = payload.get('context')
    metadata = payload.get('metadata')

    tracker.record_click(
        requested_token,
        session.get('user_id'),
        session.get('role'),
        event_name,
        context=context,
        metadata=metadata,
        occurred_at=timestamp,
    )

    return jsonify({'ok': True})
