import os
import io
import pytest
from openpyxl import Workbook

os.environ.setdefault("USER_PASSWORD", "pw")
os.environ.setdefault("ADMIN_PASSWORD", "pw")

import app as app_module
from app import create_app
from app.main import routes


@pytest.fixture
def app_instance(monkeypatch):
    monkeypatch.setattr(app_module, "create_client", lambda url, key: object())
    os.environ.setdefault("SECRET_KEY", "test")
    os.environ.setdefault("SUPABASE_URL", "http://localhost")
    os.environ.setdefault("SUPABASE_SERVICE_KEY", "service")
    app = create_app()
    return app


def make_workbook():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=7, column=2, value="ModelX")
    ws.cell(row=7, column=3, value=1)
    ws.cell(row=7, column=4, value=1)
    ws.cell(row=7, column=5, value=1)
    ws.cell(row=7, column=6, value=1)
    ws.cell(row=7, column=7, value=1)
    ws.cell(row=7, column=8, value=1)
    ws.cell(row=7, column=9, value=1)
    ws.cell(row=8, column=2, value="Total")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def make_formatted_workbook():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=7, column=2, value="ModelY")
    ws.cell(row=7, column=3, value="1,234")
    ws.cell(row=7, column=4, value="2,500")
    ws.cell(row=7, column=5, value="3,750")
    ws.cell(row=7, column=6, value="4,000")
    ws.cell(row=7, column=7, value="5.5%")
    ws.cell(row=7, column=8, value="6,250")
    ws.cell(row=7, column=9, value="7%")
    ws.cell(row=8, column=2, value="Total")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_upload_dpm_single_date(app_instance, monkeypatch):
    client = app_instance.test_client()
    captured = {}

    def fake_insert(rows):
        captured["rows"] = rows
        return None, None

    monkeypatch.setattr(routes, "insert_moat_dpm_bulk", fake_insert)
    with app_instance.app_context():
        data = {"file": (make_workbook(), "DPMReportControl 2024-07-01 L1.xlsx")}
        with client.session_transaction() as sess:
            sess["username"] = "ADMIN"
        resp = client.post("/dpm_reports/upload", data=data, content_type="multipart/form-data")
    assert resp.status_code == 201
    assert resp.get_json()["inserted"] == 1
    assert captured["rows"][0]["line"] == "L1"
    assert captured["rows"][0]["report_date"] == "2024-07-01"
    assert captured["rows"][0]["windows_per_board"] == 1.0
    assert captured["rows"][0]["total_windows"] == 1
    assert captured["rows"][0]["ng_windows"] == 1
    assert captured["rows"][0]["dpm"] == 1.0
    assert captured["rows"][0]["falsecall_windows"] == 1
    assert captured["rows"][0]["fc_dpm"] == 1.0


def test_upload_dpm_date_range(app_instance, monkeypatch):
    client = app_instance.test_client()
    captured = {}

    def fake_insert(rows):
        captured["rows"] = rows
        return None, None

    monkeypatch.setattr(routes, "insert_moat_dpm_bulk", fake_insert)
    with app_instance.app_context():
        data = {
            "file": (make_workbook(), "DPMReportControl 2024-07-01 to 2024-07-02 L2.xlsx")
        }
        with client.session_transaction() as sess:
            sess["username"] = "ADMIN"
        resp = client.post("/dpm_reports/upload", data=data, content_type="multipart/form-data")
    assert resp.status_code == 201
    assert resp.get_json()["inserted"] == 1
    assert captured["rows"][0]["line"] == "L2"
    assert captured["rows"][0]["report_date"] == "2024-07-01"


def test_upload_dpm_mixed_case_filename(app_instance, monkeypatch):
    client = app_instance.test_client()
    captured = {}

    def fake_insert(rows):
        captured["rows"] = rows
        return None, None

    monkeypatch.setattr(routes, "insert_moat_dpm_bulk", fake_insert)
    with app_instance.app_context():
        data = {
            "file": (
                make_workbook(),
                "dPmRePorTCoNtRoL 2024-07-01 To 2024-07-02 l3.XLSX",
            )
        }
        with client.session_transaction() as sess:
            sess["username"] = "ADMIN"
        resp = client.post(
            "/dpm_reports/upload",
            data=data,
            content_type="multipart/form-data",
        )
    assert resp.status_code == 201
    assert resp.get_json()["inserted"] == 1
    assert captured["rows"][0]["line"] == "l3"
    assert captured["rows"][0]["report_date"] == "2024-07-01"


def test_upload_dpm_coerces_formatted_numbers(app_instance, monkeypatch):
    client = app_instance.test_client()
    captured = {}

    def fake_insert(rows):
        captured["rows"] = rows
        return None, None

    monkeypatch.setattr(routes, "insert_moat_dpm_bulk", fake_insert)
    with app_instance.app_context():
        data = {"file": (make_formatted_workbook(), "DPMReportControl 2024-07-01 L4.xlsx")}
        with client.session_transaction() as sess:
            sess["username"] = "ADMIN"
        resp = client.post("/dpm_reports/upload", data=data, content_type="multipart/form-data")

    assert resp.status_code == 201
    assert resp.get_json()["inserted"] == 1
    row = captured["rows"][0]
    assert row["model_name"] == "ModelY"
    assert row["total_boards"] == 1234
    assert row["windows_per_board"] == 2500.0
    assert row["total_windows"] == 3750
    assert row["ng_windows"] == 4000
    assert row["dpm"] == 5.5
    assert row["falsecall_windows"] == 6250
    assert row["fc_dpm"] == 7.0
